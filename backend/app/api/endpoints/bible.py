import io
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.database import get_db
from app.domain.timing_bible_data import DEFAULT_BIBLE
from app.models.db_models import CustomBibleEntry as DBEntry
from app.models.timing_bible import BibleEntry, TimingBible

router = APIRouter()


# ---------------------------------------------------------------------------
# Official (hardcoded) bible
# ---------------------------------------------------------------------------

@router.get("/bible", response_model=TimingBible)
async def get_bible():
    """Return the built-in Timing Bible."""
    return DEFAULT_BIBLE


@router.get("/bible/lookup/{code}", response_model=BibleEntry | None)
async def lookup_bible_entry(code: str):
    """Look up a single built-in bible entry by account code."""
    return DEFAULT_BIBLE.get_entry(code)


@router.get("/bible/codes", response_model=list[str])
async def get_bible_codes():
    """Return all account codes covered by the built-in bible."""
    return DEFAULT_BIBLE.get_codes()


# ---------------------------------------------------------------------------
# Custom bible (user-defined, stored in the database)
# ---------------------------------------------------------------------------

@router.get("/bible/custom", response_model=list[BibleEntry])
def get_custom_bible(db: Session = Depends(get_db)):
    """Return all user-saved custom bible entries."""
    rows = db.query(DBEntry).order_by(DBEntry.account_code).all()
    return [
        BibleEntry(
            account_code=r.account_code,
            description=r.description,
            timing_pattern=r.timing_pattern,
            timing_title=r.timing_title,
            timing_details=r.timing_details,
            is_custom=True,
        )
        for r in rows
    ]


@router.put("/bible/custom/{code}", response_model=BibleEntry)
def upsert_custom_bible_entry(
    code: str,
    entry: BibleEntry,
    db: Session = Depends(get_db),
):
    """Create or update a custom bible entry for the given account code."""
    if entry.account_code != code:
        raise HTTPException(
            status_code=400,
            detail="account_code in the request body must match the URL code",
        )
    row = db.query(DBEntry).filter(DBEntry.account_code == code).first()
    if row:
        row.description = entry.description
        row.timing_pattern = entry.timing_pattern
        row.timing_title = entry.timing_title
        row.timing_details = entry.timing_details
        row.updated_at = datetime.now(timezone.utc)
    else:
        db.add(
            DBEntry(
                account_code=code,
                description=entry.description,
                timing_pattern=entry.timing_pattern,
                timing_title=entry.timing_title,
                timing_details=entry.timing_details,
            )
        )
    db.commit()
    return BibleEntry(
        account_code=entry.account_code,
        description=entry.description,
        timing_pattern=entry.timing_pattern,
        timing_title=entry.timing_title,
        timing_details=entry.timing_details,
        is_custom=True,
    )


@router.delete("/bible/custom/{code}", status_code=204)
def delete_custom_bible_entry(code: str, db: Session = Depends(get_db)):
    """Remove a custom bible entry."""
    row = db.query(DBEntry).filter(DBEntry.account_code == code).first()
    if row:
        db.delete(row)
        db.commit()


# ---------------------------------------------------------------------------
# Export — merged bible as Excel
# ---------------------------------------------------------------------------

@router.get("/bible/export")
def export_bible(db: Session = Depends(get_db)):
    """Download the full bible (official + custom overrides) as an Excel file."""
    # Build merged map: official first, custom entries override by code
    merged: dict[str, BibleEntry] = {e.account_code: e for e in DEFAULT_BIBLE.entries}
    for row in db.query(DBEntry).order_by(DBEntry.account_code).all():
        merged[row.account_code] = BibleEntry(
            account_code=row.account_code,
            description=row.description,
            timing_pattern=row.timing_pattern,
            timing_title=row.timing_title,
            timing_details=row.timing_details,
            is_custom=True,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Timing Bible"

    headers = ["Account Code", "Description", "Timing Pattern", "Timing Title", "Timing Details", "Source"]
    ws.append(headers)

    # Bold header row
    from openpyxl.styles import Font, PatternFill, Alignment
    header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
    for col, cell in enumerate(ws[1], 1):
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    custom_fill = PatternFill(start_color="DBEAFE", end_color="DBEAFE", fill_type="solid")

    for entry in sorted(merged.values(), key=lambda e: e.account_code):
        ws.append([
            entry.account_code,
            entry.description,
            entry.timing_pattern,
            entry.timing_title,
            entry.timing_details,
            "custom" if entry.is_custom else "official",
        ])
        if entry.is_custom:
            for cell in ws[ws.max_row]:
                cell.fill = custom_fill

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"timing_bible_{datetime.now(timezone.utc).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
