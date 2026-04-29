import io

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import Response
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.cashflow import CashflowOutput, GenerateRequest
from app.models.db_models import CashflowTemplateDistribution
from app.models.distribution import CurveType, LineItemDistribution, PhaseAssignment
from app.services.phase_mapper import get_default_distributions
from app.services.cashflow_engine import generate_cashflow
from app.services.excel_writer import write_cashflow_excel

router = APIRouter()


@router.post("/cashflow/preview", response_model=CashflowOutput)
async def preview_cashflow(request: GenerateRequest):
    output = generate_cashflow(
        budget=request.budget,
        parameters=request.parameters,
        distributions=request.distributions,
    )
    return output


@router.post("/cashflow/generate")
async def generate_cashflow_excel(request: GenerateRequest):
    output = generate_cashflow(
        budget=request.budget,
        parameters=request.parameters,
        distributions=request.distributions,
    )
    buffer = write_cashflow_excel(output, request.parameters, budget=request.budget)
    filename = f"{request.parameters.title.replace(' ', '_')}_cashflow.xlsx"

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _to_dist(row: CashflowTemplateDistribution) -> LineItemDistribution:
    return LineItemDistribution(
        budget_code=row.budget_code,
        phase=PhaseAssignment(row.phase),
        curve=CurveType(row.curve),
        auto_assigned=False,
        timing_pattern_override=row.timing_pattern_override,
    )


@router.get("/cashflow/templates", response_model=list[str])
async def list_cashflow_templates(db: Session = Depends(get_db)):
    rows = db.query(CashflowTemplateDistribution.template_name).distinct().all()
    return sorted({r[0] for r in rows if r[0]})


@router.get("/cashflow/templates/{template_name}", response_model=list[LineItemDistribution])
async def get_cashflow_template(
    template_name: str,
    codes: list[str] = Query(default=[]),
    db: Session = Depends(get_db),
):
    saved = (
        db.query(CashflowTemplateDistribution)
        .filter(CashflowTemplateDistribution.template_name == template_name)
        .all()
    )
    by_code = {r.budget_code: r for r in saved}
    if not codes:
        return [_to_dist(r) for r in sorted(saved, key=lambda r: r.budget_code)]
    defaults = get_default_distributions(codes)
    merged: list[LineItemDistribution] = []
    for d in defaults:
        row = by_code.get(d.budget_code)
        merged.append(_to_dist(row) if row else d)
    return merged


@router.put("/cashflow/templates/{template_name}", response_model=list[LineItemDistribution])
async def save_cashflow_template(
    template_name: str,
    body: list[LineItemDistribution],
    db: Session = Depends(get_db),
):
    for d in body:
        row_id = f"{template_name}::{d.budget_code}"
        existing = db.query(CashflowTemplateDistribution).filter(
            CashflowTemplateDistribution.id == row_id
        ).first()
        if existing:
            existing.phase = d.phase.value
            existing.curve = d.curve.value
            existing.timing_pattern_override = d.timing_pattern_override
        else:
            db.add(CashflowTemplateDistribution(
                id=row_id,
                template_name=template_name,
                budget_code=d.budget_code,
                phase=d.phase.value,
                curve=d.curve.value,
                timing_pattern_override=d.timing_pattern_override,
            ))
    db.commit()
    return body


@router.get("/cashflow/templates/{template_name}/excel")
async def download_cashflow_template_excel(template_name: str, db: Session = Depends(get_db)):
    rows = (
        db.query(CashflowTemplateDistribution)
        .filter(CashflowTemplateDistribution.template_name == template_name)
        .order_by(CashflowTemplateDistribution.budget_code)
        .all()
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Template not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cashflow Template"
    ws.append(["Account", "Phase", "Curve", "Timing Pattern Override"])
    for row in rows:
        ws.append([row.budget_code, row.phase, row.curve, row.timing_pattern_override or ""])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe = template_name.strip().replace(" ", "_") or "template"
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="cashflow_template_{safe}.xlsx"'},
    )


@router.post("/cashflow/templates/{template_name}/upload", response_model=list[LineItemDistribution])
async def upload_cashflow_template_excel(
    template_name: str,
    file: UploadFile,
    db: Session = Depends(get_db),
):
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")
    payload = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    ws = wb.active
    header = [str(c.value or "").strip().lower() for c in ws[1]]

    def _idx(candidates: list[str]) -> int | None:
        for c in candidates:
            if c in header:
                return header.index(c)
        return None

    code_idx = _idx(["account", "account code", "budget_code", "code"])
    phase_idx = _idx(["phase"])
    curve_idx = _idx(["curve"])
    timing_idx = _idx(["timing pattern override", "timing_pattern_override"])
    if None in (code_idx, phase_idx, curve_idx):
        raise HTTPException(status_code=400, detail="Template must include Account, Phase, and Curve columns")

    parsed: list[LineItemDistribution] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_idx] or "").strip()  # type: ignore[index]
        if not code:
            continue
        phase_raw = str(row[phase_idx] or "").strip()  # type: ignore[index]
        curve_raw = str(row[curve_idx] or "").strip()  # type: ignore[index]
        try:
            phase = PhaseAssignment(phase_raw)
            curve = CurveType(curve_raw)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Invalid phase/curve for account {code}") from exc
        timing = ""
        if timing_idx is not None:
            timing = str(row[timing_idx] or "").strip()  # type: ignore[index]
        parsed.append(LineItemDistribution(
            budget_code=code,
            phase=phase,
            curve=curve,
            auto_assigned=False,
            timing_pattern_override=timing or None,
        ))

    if not parsed:
        raise HTTPException(status_code=400, detail="No template rows were found")
    return await save_cashflow_template(template_name, parsed, db)
