import io
from datetime import datetime, timezone

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.database import get_db
from app.models.budget import ParsedBudget
from app.models.db_models import BreakoutBibleEntry, TaxCreditOverride
from app.services.tax_credit_writer import (
    BREAKOUT_BIBLE,
    write_bible_excel,
    write_tax_credit_excel,
)

router = APIRouter()


# ---------------------------------------------------------------------------
# Pydantic schemas
# ---------------------------------------------------------------------------

class BreakoutOverride(BaseModel):
    account_code: str
    description: str = ""
    is_foreign: bool | None = None
    is_non_prov: bool | None = None
    fed_labour_pct: float | None = None
    fed_svc_labour_pct: float | None = None
    prov_labour_pct: float | None = None
    prov_svc_labour_pct: float | None = None
    svc_property_pct: float | None = None


class ProjectOverridesResponse(BaseModel):
    project_name: str
    overrides: list[BreakoutOverride]


class SaveOverridesRequest(BaseModel):
    overrides: list[BreakoutOverride]


class BibleEntrySchema(BaseModel):
    account_code: str
    description: str = ""
    is_non_prov: bool
    prov_labour_pct: float
    fed_labour_pct: float
    prov_svc_labour_pct: float
    svc_property_pct: float
    fed_svc_labour_pct: float
    is_customized: bool = False  # True when a DB row overrides/creates the entry
    is_standard: bool = True     # False when account_code is not in BREAKOUT_BIBLE


class TaxCreditRequest(BaseModel):
    budget: ParsedBudget
    title: str = "Untitled"
    overrides: list[BreakoutOverride] | None = None


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _override_key(project_name: str, account_code: str) -> str:
    return f"{project_name}::{account_code}"


def _template_project_name(template_name: str) -> str:
    return f"template::{template_name.strip()}"


def _list_template_names(db: Session) -> list[str]:
    rows = (
        db.query(TaxCreditOverride.project_name)
        .filter(TaxCreditOverride.project_name.like("template::%"))
        .distinct()
        .all()
    )
    names = sorted(
        {
            row[0].split("template::", 1)[1]
            for row in rows
            if row[0].startswith("template::") and row[0].split("template::", 1)[1].strip()
        }
    )
    return names


def _bible_defaults(account_code: str, description: str = "") -> BreakoutOverride:
    entry = BREAKOUT_BIBLE.get(account_code)
    if entry:
        non_prov_out, pl, fl, psl, sp, fsl = entry
        return BreakoutOverride(
            account_code=account_code,
            description=description,
            is_foreign=None,
            is_non_prov=non_prov_out,
            fed_labour_pct=fl,
            fed_svc_labour_pct=fsl,
            prov_labour_pct=pl,
            prov_svc_labour_pct=psl,
            svc_property_pct=sp,
        )
    return BreakoutOverride(
        account_code=account_code,
        description=description,
        is_foreign=None,
        is_non_prov=False,
        fed_labour_pct=0.0,
        fed_svc_labour_pct=0.0,
        prov_labour_pct=0.0,
        prov_svc_labour_pct=0.0,
        svc_property_pct=0.0,
    )


def _db_row_to_override(row: TaxCreditOverride, description: str = "") -> BreakoutOverride:
    return BreakoutOverride(
        account_code=row.account_code,
        description=description,
        is_foreign=row.is_foreign,
        is_non_prov=row.is_non_prov,
        fed_labour_pct=row.fed_labour_pct,
        fed_svc_labour_pct=row.fed_svc_labour_pct,
        prov_labour_pct=row.prov_labour_pct,
        prov_svc_labour_pct=row.prov_svc_labour_pct,
        svc_property_pct=row.svc_property_pct,
    )


def _load_global_bible(db: Session) -> dict:
    """Return global bible customisations as a tuple-dict for write_tax_credit_excel."""
    rows = db.query(BreakoutBibleEntry).all()
    return {
        r.account_code: (
            r.is_non_prov,
            r.prov_labour_pct,
            r.fed_labour_pct,
            r.prov_svc_labour_pct,
            r.svc_property_pct,
            r.fed_svc_labour_pct,
        )
        for r in rows
    }


def _load_overrides(
    project_name: str,
    account_codes: list[str],
    descriptions: list[str],
    db: Session,
) -> ProjectOverridesResponse:
    desc_map = dict(zip(account_codes, descriptions)) if descriptions else {}

    saved_rows = (
        db.query(TaxCreditOverride)
        .filter(TaxCreditOverride.project_name == project_name)
        .all()
    )
    saved_by_code = {r.account_code: r for r in saved_rows}

    results: list[BreakoutOverride] = []
    for code in account_codes:
        desc = desc_map.get(code, "")
        if code in saved_by_code:
            ov = _db_row_to_override(saved_by_code[code], desc)
        else:
            ov = _bible_defaults(code, desc)
        results.append(ov)

    return ProjectOverridesResponse(project_name=project_name, overrides=results)


def _save_overrides(
    project_name: str,
    body: SaveOverridesRequest,
    db: Session,
) -> ProjectOverridesResponse:
    for ov in body.overrides:
        key = _override_key(project_name, ov.account_code)
        existing = db.query(TaxCreditOverride).filter(TaxCreditOverride.id == key).first()
        if existing:
            existing.is_foreign = ov.is_foreign
            existing.is_non_prov = ov.is_non_prov
            existing.fed_labour_pct = ov.fed_labour_pct
            existing.fed_svc_labour_pct = ov.fed_svc_labour_pct
            existing.prov_labour_pct = ov.prov_labour_pct
            existing.prov_svc_labour_pct = ov.prov_svc_labour_pct
            existing.svc_property_pct = ov.svc_property_pct
        else:
            db.add(TaxCreditOverride(
                id=key,
                project_name=project_name,
                account_code=ov.account_code,
                is_foreign=ov.is_foreign,
                is_non_prov=ov.is_non_prov,
                fed_labour_pct=ov.fed_labour_pct,
                fed_svc_labour_pct=ov.fed_svc_labour_pct,
                prov_labour_pct=ov.prov_labour_pct,
                prov_svc_labour_pct=ov.prov_svc_labour_pct,
                svc_property_pct=ov.svc_property_pct,
            ))
    db.commit()
    return ProjectOverridesResponse(project_name=project_name, overrides=body.overrides)


# ---------------------------------------------------------------------------
# Bible endpoints
# ---------------------------------------------------------------------------

@router.get("/tax-credit/bible/excel")
async def download_bible_excel(db: Session = Depends(get_db)):
    """Download the full effective breakout bible as a formatted Excel file."""
    db_rows = {r.account_code: r for r in db.query(BreakoutBibleEntry).all()}

    entries = []
    for code, defaults in BREAKOUT_BIBLE.items():
        non_prov, pl, fl, psl, sp, fsl = defaults
        if code in db_rows:
            r = db_rows[code]
            entries.append({
                "account_code": code,
                "is_non_prov": r.is_non_prov,
                "prov_labour_pct": r.prov_labour_pct,
                "fed_labour_pct": r.fed_labour_pct,
                "prov_svc_labour_pct": r.prov_svc_labour_pct,
                "svc_property_pct": r.svc_property_pct,
                "fed_svc_labour_pct": r.fed_svc_labour_pct,
                "is_customized": True,
            })
        else:
            entries.append({
                "account_code": code,
                "is_non_prov": non_prov,
                "prov_labour_pct": pl,
                "fed_labour_pct": fl,
                "prov_svc_labour_pct": psl,
                "svc_property_pct": sp,
                "fed_svc_labour_pct": fsl,
                "is_customized": False,
            })

    buffer = write_bible_excel(entries)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="breakout_bible.xlsx"'},
    )


# ---------------------------------------------------------------------------
# Project overrides endpoints
# ---------------------------------------------------------------------------

@router.get("/tax-credit/overrides/{project_name}", response_model=ProjectOverridesResponse)
async def get_overrides(
    project_name: str,
    account_codes: list[str] = Query(default=[]),
    descriptions: list[str] = Query(default=[]),
    db: Session = Depends(get_db),
):
    return _load_overrides(project_name, account_codes, descriptions, db)


@router.put("/tax-credit/overrides/{project_name}", response_model=ProjectOverridesResponse)
async def save_overrides(
    project_name: str,
    body: SaveOverridesRequest,
    db: Session = Depends(get_db),
):
    return _save_overrides(project_name, body, db)


@router.get("/tax-credit/templates", response_model=list[str])
async def list_templates(db: Session = Depends(get_db)):
    """List saved breakout templates."""
    return _list_template_names(db)


@router.get("/tax-credit/templates/{template_name}", response_model=ProjectOverridesResponse)
async def get_template(
    template_name: str,
    account_codes: list[str] = Query(default=[]),
    descriptions: list[str] = Query(default=[]),
    db: Session = Depends(get_db),
):
    """Load one saved breakout template by name."""
    project_key = _template_project_name(template_name)
    return _load_overrides(project_key, account_codes, descriptions, db)


@router.put("/tax-credit/templates/{template_name}", response_model=ProjectOverridesResponse)
async def save_template(
    template_name: str,
    body: SaveOverridesRequest,
    db: Session = Depends(get_db),
):
    """Save one breakout template by name."""
    project_key = _template_project_name(template_name)
    return _save_overrides(project_key, body, db)


@router.get("/tax-credit/templates/{template_name}/excel")
async def download_template_excel(template_name: str, db: Session = Depends(get_db)):
    """Download a template in the breakout bible excel format."""
    project_key = _template_project_name(template_name)
    rows = (
        db.query(TaxCreditOverride)
        .filter(TaxCreditOverride.project_name == project_key)
        .all()
    )
    if not rows:
        raise HTTPException(status_code=404, detail="Template not found")

    data = []
    for row in sorted(rows, key=lambda r: r.account_code):
        defaults = BREAKOUT_BIBLE.get(row.account_code, (False, 0.0, 0.0, 0.0, 0.0, 0.0))
        non_prov_default, pl_default, fl_default, psl_default, sp_default, fsl_default = defaults
        data.append({
            "account_code": row.account_code,
            "is_non_prov": row.is_non_prov if row.is_non_prov is not None else non_prov_default,
            "prov_labour_pct": row.prov_labour_pct if row.prov_labour_pct is not None else pl_default,
            "fed_labour_pct": row.fed_labour_pct if row.fed_labour_pct is not None else fl_default,
            "prov_svc_labour_pct": row.prov_svc_labour_pct if row.prov_svc_labour_pct is not None else psl_default,
            "svc_property_pct": row.svc_property_pct if row.svc_property_pct is not None else sp_default,
            "fed_svc_labour_pct": row.fed_svc_labour_pct if row.fed_svc_labour_pct is not None else fsl_default,
            "is_customized": True,
        })

    buffer = write_bible_excel(data)
    safe = template_name.strip().replace(" ", "_") or "template"
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="breakout_template_{safe}.xlsx"'},
    )


@router.post("/tax-credit/templates/{template_name}/upload", response_model=ProjectOverridesResponse)
async def upload_template_excel(
    template_name: str,
    file: UploadFile,
    db: Session = Depends(get_db),
):
    """Upload a breakout template from an Excel sheet using breakout bible format."""
    if not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload an .xlsx file")

    payload = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
    ws = wb.active

    header = [str(c.value or "").strip().lower() for c in ws[1]]
    def _idx(names: list[str]) -> int | None:
        for n in names:
            if n in header:
                return header.index(n)
        return None

    code_idx = _idx(["account code", "account", "account_code", "code"])
    out_idx = _idx(["out", "is_non_prov", "is non prov", "is non-prov"])
    pl_idx = _idx(["prov labour %", "prov_labour_pct"])
    fl_idx = _idx(["fed labour %", "fed_labour_pct"])
    psl_idx = _idx(["prov svc %", "prov_svc_labour_pct", "prov svc labour %"])
    sp_idx = _idx(["svc property %", "svc prop %", "svc_property_pct"])
    fsl_idx = _idx(["fed svc %", "fed_svc_labour_pct", "fed svc labour %"])
    desc_idx = _idx(["description"])

    if None in (code_idx, out_idx, pl_idx, fl_idx, psl_idx, sp_idx, fsl_idx):
        raise HTTPException(status_code=400, detail="Template format is missing required breakout columns")

    def _to_pct(value: object) -> float | None:
        if value is None or str(value).strip() == "":
            return None
        if isinstance(value, (int, float)):
            return float(value) if float(value) <= 1 else float(value) / 100.0
        raw = str(value).strip().replace("%", "")
        if raw == "":
            return None
        parsed = float(raw)
        return parsed if parsed <= 1 else parsed / 100.0

    overrides: list[BreakoutOverride] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[code_idx] or "").strip()  # type: ignore[index]
        if not code:
            continue
        defaults = BREAKOUT_BIBLE.get(code, (False, 0.0, 0.0, 0.0, 0.0, 0.0))
        non_prov_default, pl_default, fl_default, psl_default, sp_default, fsl_default = defaults
        out_raw = row[out_idx]  # type: ignore[index]
        if isinstance(out_raw, str):
            is_non_prov = out_raw.strip().lower() in {"1", "true", "yes", "y", "out"}
        elif out_raw is None:
            is_non_prov = non_prov_default
        else:
            is_non_prov = bool(out_raw)
        overrides.append(BreakoutOverride(
            account_code=code,
            description=str(row[desc_idx] or "").strip() if desc_idx is not None else "",
            is_foreign=None,
            is_non_prov=is_non_prov,
            prov_labour_pct=_to_pct(row[pl_idx]) if row[pl_idx] is not None else pl_default,  # type: ignore[index]
            fed_labour_pct=_to_pct(row[fl_idx]) if row[fl_idx] is not None else fl_default,  # type: ignore[index]
            prov_svc_labour_pct=_to_pct(row[psl_idx]) if row[psl_idx] is not None else psl_default,  # type: ignore[index]
            svc_property_pct=_to_pct(row[sp_idx]) if row[sp_idx] is not None else sp_default,  # type: ignore[index]
            fed_svc_labour_pct=_to_pct(row[fsl_idx]) if row[fsl_idx] is not None else fsl_default,  # type: ignore[index]
        ))

    if not overrides:
        raise HTTPException(status_code=400, detail="No template rows were found in the uploaded file")

    body = SaveOverridesRequest(overrides=overrides)
    return _save_overrides(_template_project_name(template_name), body, db)


@router.get("/tax-credit/bible", response_model=list[BibleEntrySchema])
async def get_bible(db: Session = Depends(get_db)):
    """Return all breakout bible entries: BREAKOUT_BIBLE defaults merged with DB overrides,
    plus any custom (non-standard) accounts stored only in the DB."""
    db_rows: dict[str, BreakoutBibleEntry] = {
        r.account_code: r
        for r in db.query(BreakoutBibleEntry).all()
    }

    results: list[BibleEntrySchema] = []

    # Standard entries from BREAKOUT_BIBLE
    for code, entry in BREAKOUT_BIBLE.items():
        non_prov, pl, fl, psl, sp, fsl = entry
        row = db_rows.get(code)
        if row:
            results.append(BibleEntrySchema(
                account_code=code,
                description=row.description,
                is_non_prov=row.is_non_prov,
                prov_labour_pct=row.prov_labour_pct,
                fed_labour_pct=row.fed_labour_pct,
                prov_svc_labour_pct=row.prov_svc_labour_pct,
                svc_property_pct=row.svc_property_pct,
                fed_svc_labour_pct=row.fed_svc_labour_pct,
                is_customized=True,
                is_standard=True,
            ))
        else:
            results.append(BibleEntrySchema(
                account_code=code,
                description="",
                is_non_prov=non_prov,
                prov_labour_pct=pl,
                fed_labour_pct=fl,
                prov_svc_labour_pct=psl,
                svc_property_pct=sp,
                fed_svc_labour_pct=fsl,
                is_customized=False,
                is_standard=True,
            ))

    # Custom (non-standard) entries that exist only in the DB
    for code, row in db_rows.items():
        if code not in BREAKOUT_BIBLE:
            results.append(BibleEntrySchema(
                account_code=code,
                description=row.description,
                is_non_prov=row.is_non_prov,
                prov_labour_pct=row.prov_labour_pct,
                fed_labour_pct=row.fed_labour_pct,
                prov_svc_labour_pct=row.prov_svc_labour_pct,
                svc_property_pct=row.svc_property_pct,
                fed_svc_labour_pct=row.fed_svc_labour_pct,
                is_customized=True,
                is_standard=False,
            ))

    results.sort(key=lambda e: e.account_code)
    return results


@router.put("/tax-credit/bible/{account_code}", response_model=BibleEntrySchema)
async def upsert_bible_entry(
    account_code: str,
    body: BibleEntrySchema,
    db: Session = Depends(get_db),
):
    """Create or update a breakout bible entry in the database."""
    row = db.query(BreakoutBibleEntry).filter(
        BreakoutBibleEntry.account_code == account_code
    ).first()
    if row:
        row.description         = body.description
        row.is_non_prov         = body.is_non_prov
        row.prov_labour_pct     = body.prov_labour_pct
        row.fed_labour_pct      = body.fed_labour_pct
        row.prov_svc_labour_pct = body.prov_svc_labour_pct
        row.svc_property_pct    = body.svc_property_pct
        row.fed_svc_labour_pct  = body.fed_svc_labour_pct
    else:
        db.add(BreakoutBibleEntry(
            account_code        = account_code,
            description         = body.description,
            is_non_prov         = body.is_non_prov,
            prov_labour_pct     = body.prov_labour_pct,
            fed_labour_pct      = body.fed_labour_pct,
            prov_svc_labour_pct = body.prov_svc_labour_pct,
            svc_property_pct    = body.svc_property_pct,
            fed_svc_labour_pct  = body.fed_svc_labour_pct,
        ))
    db.commit()
    is_standard = account_code in BREAKOUT_BIBLE
    return BibleEntrySchema(**body.model_dump(), is_customized=True, is_standard=is_standard)


@router.delete("/tax-credit/bible/{account_code}", status_code=204)
async def delete_bible_entry(account_code: str, db: Session = Depends(get_db)):
    """Delete the DB override for a bible entry (reverts standard entries to defaults,
    fully removes custom entries)."""
    row = db.query(BreakoutBibleEntry).filter(
        BreakoutBibleEntry.account_code == account_code
    ).first()
    if row:
        db.delete(row)
        db.commit()
    return Response(status_code=204)


@router.post("/tax-credit/generate")
async def generate_tax_credit_excel(
    request: TaxCreditRequest,
    db: Session = Depends(get_db),
):
    overrides_map: dict[str, BreakoutOverride] = {}
    if request.overrides:
        overrides_map = {ov.account_code: ov for ov in request.overrides}

    global_bible = _load_global_bible(db)

    buffer = write_tax_credit_excel(
        request.budget,
        request.title,
        overrides_map,
        global_bible or None,
    )
    filename = f"{request.title.replace(' ', '_')}_tax_credit_budget.xlsx"

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
