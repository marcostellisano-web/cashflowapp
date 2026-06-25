"""Generate a formatted Tax Credit Filing Budget Excel workbook from a ParsedBudget."""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.models.budget import ParsedBudget

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------
CURRENCY_FORMAT = '#,##0'
# Accounting style: no currency symbol, zero shows as " - ", negatives in parens
_ACCOUNTING_FORMAT = '_(* #,##0_);_(* (#,##0);_(* "-"_);_(@_)'

_THIN = Side(style="thin")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOTTOM_BORDER = Border(bottom=_THIN)
_NO_BORDER = Border()

_BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
_SECTION_HEADER_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_LIGHT_GRAY_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_TOTAL_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_GRAND_TOTAL_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")

_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=10)
_BOLD = Font(bold=True, size=10)
_NORMAL = Font(size=10)
_TITLE_FONT = Font(bold=True, size=11)
_BOLD_ITALIC = Font(bold=True, italic=True, size=10)

_TITLE_GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------------------------
# CAVCO Topsheet account definitions
# ---------------------------------------------------------------------------
# Each entry is one of:
#   ("XX.00", "Category name")          — a data row with an account code
#   ("TOTAL", "label", "section_key")   — a section-total row
#   ("HEADER", "label")                 — a section header label (gray)
#   ("BLANK",)                          — an empty spacer row

TOPSHEET_STRUCTURE: list[tuple] = [
    # Above the line
    ("01.00", "Story rights/Acquisitions"),
    ("02.00", "Script"),
    ("03.00", "Development costs"),
    ("04.00", "Producer(s)"),
    ("05.00", "Director(s)"),
    ("06.00", "Stars"),
    ("TOTAL", 'TOTAL "A" \u2013 ABOVE THE LINE', "A"),
    # Section B
    ("HEADER", '"B" \u2013 PRODUCTION'),
    ("10.00", "Cast"),
    ("11.00", "Background Performers (Extras)"),
    ("12.00", "Production labour"),
    ("13.00", "Production Design/Art Department labour"),
    ("14.00", "Construction labour"),
    ("15.00", "Set Dressing labour"),
    ("16.00", "Props labour"),
    ("17.00", "Special Effects labour"),
    ("18.00", "Animal Wrangling labour"),
    ("19.00", "Wardrobe labour"),
    ("20.00", "Makeup/Hair labour"),
    ("21.00", "Video Technical crew"),
    ("22.00", "Camera labour"),
    ("23.00", "Electrical labour"),
    ("24.00", "Grip labour"),
    ("25.00", "Production Sound labour"),
    ("26.00", "Transportation labour"),
    ("27.00", "Fringe benefits"),
    ("28.00", "Production office expenses"),
    ("29.00", "Studio expenses"),
    ("30.00", "Location office expenses"),
    ("31.00", "Location expenses"),
    ("32.00", "Unit expenses"),
    ("33.00", "Travel & Living expenses"),
    ("34.00", "Transportation"),
    ("35.00", "Construction materials"),
    ("36.00", "Art supplies"),
    ("37.00", "Set dressing"),
    ("38.00", "Props"),
    ("39.00", "Special effects"),
    ("40.00", "Animals"),
    ("41.00", "Wardrobe supplies"),
    ("42.00", "Makeup/Hair supplies"),
    ("43.00", "Videotape studio"),
    ("44.00", "Mobile video unit"),
    ("45.00", "Camera equipment"),
    ("46.00", "Electrical equipment"),
    ("47.00", "Grip equipment"),
    ("48.00", "Sound equipment"),
    ("49.00", "Second unit"),
    ("50.00", "Video stock"),
    ("51.00", "Production laboratory"),
    ("52.00", "Voice recording \u2013 Animation"),
    ("53.00", "Production unit \u2013 Animation"),
    ("54.00", "Art & Design unit \u2013 Animation"),
    ("55.00", "2D Animation unit"),
    ("56.00", "3D Animation unit"),
    ("57.00", "Live Animation (MOCAP) unit"),
    ("58.00", "Fringe benefits \u2013 Animation"),
    ("59.00", "Animation materials & supplies"),
    ("TOTAL", 'TOTAL PRODUCTION "B"', "B"),
    # Section C
    ("HEADER", '"C" \u2013 POST-PRODUCTION'),
    ("60.00", "Post Production - Edit labour"),
    ("61.00", "Editing equipment"),
    ("62.00", "Video post production (picture)"),
    ("63.00", "Video post production (sound)"),
    ("64.00", "Film post production (picture)"),
    ("65.00", "Film post production (sound)"),
    ("66.00", "Music"),
    ("67.00", "Titles/Stock footage/Visual effects"),
    ("68.00", "Versioning"),
    ("69.00", "Amortization (series)"),
    ("TOTAL", 'TOTAL POST-PRODUCTION "C"', "C"),
    ("TOTAL_MULTI", 'TOTAL "B" + "C"\n(PRODUCTION AND POST PRODUCTION)', ("B", "C")),
    # Section D
    ("HEADER", '"D" \u2013 OTHER'),
    ("70.00", "Unit publicity"),
    ("71.00", "General expenses"),
    ("72.00", "Indirect costs"),
    ("TOTAL", 'TOTAL OTHER "D"', "D"),
    ("TOTAL_MULTI", 'TOTAL "A" + "B" + "C" + "D"', ("A", "B", "C", "D")),
    # Final items
    ("80.00", "Contingency"),
    ("81.00", "Completion guarantee"),
    ("GRAND_TOTAL", "GRAND TOTAL"),
]


def _cavco_to_mm_prefix(cavco_code: str) -> str:
    """Convert CAVCO code like '01.00' to 4-char account prefix '0100'."""
    integer_part = cavco_code.split(".")[0]  # "01", "10", "80"
    return integer_part + "00"              # "0100", "1000", "8000"


def _get_account_total(budget: "ParsedBudget", cavco_code: str) -> float:
    """Return the total for a CAVCO account code.

    Prefers pre-aggregated topsheet_totals from the source file's Topsheet tab.
    Falls back to summing matching line items when topsheet_totals is empty.
    """
    prefix = _cavco_to_mm_prefix(cavco_code)
    if budget.topsheet_totals:
        return budget.topsheet_totals.get(prefix, 0.0)
    # Fallback: sum line items whose stripped code starts with the prefix
    total = 0.0
    for item in budget.line_items:
        code = item.code.replace(".", "").replace(" ", "")
        if code.startswith(prefix):
            total += item.total
    return total


def _build_section_totals(budget: "ParsedBudget") -> dict[str, float]:
    """Pre-compute section totals A, B, C, D for the topsheet."""
    section_ranges = {
        "A": [f"{n:02d}.00" for n in range(1, 7)],
        "B": [f"{n:02d}.00" for n in range(10, 60)],
        "C": [f"{n:02d}.00" for n in range(60, 70)],
        "D": [f"{n:02d}.00" for n in range(70, 73)],
    }
    return {
        section: sum(_get_account_total(budget, code) for code in codes)
        for section, codes in section_ranges.items()
    }


# ---------------------------------------------------------------------------
# Topsheet worksheet builder
# ---------------------------------------------------------------------------

def _write_topsheet(ws, budget: ParsedBudget, title: str) -> None:
    ws.title = "Topsheet"

    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 16

    # ── Title row ──────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    title_cell = ws.cell(row=1, column=1, value="Title")
    title_cell.font = _BOLD
    title_cell.alignment = _LEFT

    value_cell = ws.cell(row=1, column=2, value=title)
    value_cell.font = _TITLE_FONT
    value_cell.alignment = _LEFT
    value_cell.border = _BOTTOM_BORDER

    ws.row_dimensions[2].height = 8  # spacer

    # ── Column headers ─────────────────────────────────────────────────────
    header_row = 3
    ws.row_dimensions[header_row].height = 18
    for col, label in enumerate(["Account", "Category", "Total"], start=1):
        cell = ws.cell(row=header_row, column=col, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER if col != 2 else _LEFT
        cell.border = _THIN_BORDER

    # ── Pre-compute values ─────────────────────────────────────────────────
    section_totals = _build_section_totals(budget)

    # Grand total = sum of all sections + 80.00 + 81.00
    grand_total = (
        sum(section_totals.values())
        + _get_account_total(budget, "80.00")
        + _get_account_total(budget, "81.00")
    )

    # ── Data rows ──────────────────────────────────────────────────────────
    current_row = header_row + 1

    for entry in TOPSHEET_STRUCTURE:
        kind = entry[0]
        ws.row_dimensions[current_row].height = 16

        if kind == "BLANK":
            current_row += 1
            continue

        elif kind == "HEADER":
            # Gray section header, spans A-C
            label = entry[1]
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row, end_column=3,
            )
            cell = ws.cell(row=current_row, column=1, value=label)
            cell.font = _BOLD
            cell.fill = _SECTION_HEADER_FILL
            cell.alignment = _LEFT
            cell.border = _THIN_BORDER
            current_row += 1

        elif kind == "TOTAL":
            label = entry[1]
            section_key = entry[2]
            amount = section_totals.get(section_key, 0.0)

            # Black row, white bold text
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = _BLACK_FILL
                cell.border = _THIN_BORDER
                cell.font = _WHITE_BOLD
                if col == 1:
                    cell.alignment = _LEFT
                elif col == 2:
                    cell.value = label
                    cell.alignment = _LEFT
                else:
                    cell.value = amount
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = _RIGHT
            current_row += 1

        elif kind == "TOTAL_MULTI":
            label = entry[1]
            section_keys = entry[2]
            amount = sum(section_totals.get(k, 0.0) for k in section_keys)

            ws.row_dimensions[current_row].height = 28
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = _BLACK_FILL
                cell.border = _THIN_BORDER
                cell.font = _WHITE_BOLD
                cell.alignment = Alignment(
                    horizontal="left" if col <= 2 else "right",
                    vertical="center",
                    wrap_text=True,
                )
                if col == 2:
                    cell.value = label
                elif col == 3:
                    cell.value = amount
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = Alignment(
                        horizontal="right", vertical="center", wrap_text=True
                    )
            current_row += 1

        elif kind == "GRAND_TOTAL":
            label = entry[1]
            ws.row_dimensions[current_row].height = 18
            for col in range(1, 4):
                cell = ws.cell(row=current_row, column=col)
                cell.fill = _BLACK_FILL
                cell.border = _THIN_BORDER
                cell.font = _WHITE_BOLD
                if col == 2:
                    cell.value = label
                    cell.alignment = _LEFT
                elif col == 3:
                    cell.value = grand_total
                    cell.number_format = CURRENCY_FORMAT
                    cell.alignment = _RIGHT
                else:
                    cell.alignment = _LEFT
            current_row += 1

        else:
            # Regular data row: ("XX.00", "Category name")
            cavco_code = kind
            label = entry[1]
            amount = _get_account_total(budget, cavco_code)

            row_fill = _LIGHT_GRAY_FILL if current_row % 2 == 0 else None

            code_cell = ws.cell(row=current_row, column=1, value=cavco_code)
            code_cell.font = _NORMAL
            code_cell.alignment = _CENTER
            code_cell.border = _THIN_BORDER
            if row_fill:
                code_cell.fill = row_fill

            desc_cell = ws.cell(row=current_row, column=2, value=label)
            desc_cell.font = _NORMAL
            desc_cell.alignment = _LEFT
            desc_cell.border = _THIN_BORDER
            if row_fill:
                desc_cell.fill = row_fill

            amt_cell = ws.cell(row=current_row, column=3, value=amount)
            amt_cell.font = _NORMAL
            amt_cell.number_format = CURRENCY_FORMAT
            amt_cell.alignment = _RIGHT
            amt_cell.border = _THIN_BORDER
            if row_fill:
                amt_cell.fill = row_fill

            current_row += 1

    # Freeze panes below header row
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Budget Lines detail worksheet builder
# ---------------------------------------------------------------------------

def _write_budget_lines(ws, budget: ParsedBudget) -> None:
    ws.title = "Budget Lines"

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 55
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20

    # Header row
    headers = ["Account Code", "Description", "Total", "CAVCO Category"]
    ws.row_dimensions[1].height = 18
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _BOLD
        cell.alignment = _CENTER if col != 2 else _LEFT
        cell.border = _THIN_BORDER
        cell.fill = _SECTION_HEADER_FILL

    # Build a lookup: mm_prefix → CAVCO label for tooltip column
    cavco_map: dict[str, str] = {}
    for entry in TOPSHEET_STRUCTURE:
        if len(entry) == 2 and entry[0] not in ("HEADER", "BLANK", "GRAND_TOTAL"):
            cavco_code = entry[0]
            cavco_label = entry[1]
            mm_prefix = _cavco_to_mm_prefix(cavco_code)
            cavco_map[mm_prefix] = f"{cavco_code} {cavco_label}"

    # Sort line items by code
    sorted_items = sorted(budget.line_items, key=lambda x: x.code)

    for row_idx, item in enumerate(sorted_items, start=2):
        ws.row_dimensions[row_idx].height = 15
        row_fill = _LIGHT_GRAY_FILL if row_idx % 2 == 0 else None

        clean_code = item.code.replace(".", "").replace(" ", "")
        # Find matching CAVCO prefix (first 4 chars of clean code)
        cavco_label = ""
        for prefix, label in cavco_map.items():
            if clean_code.startswith(prefix):
                cavco_label = label
                break

        cells_data = [
            (item.code, _CENTER),
            (item.description, _LEFT),
            (item.total, _RIGHT),
            (cavco_label, _LEFT),
        ]
        for col, (value, align) in enumerate(cells_data, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = _NORMAL
            cell.alignment = align
            cell.border = _THIN_BORDER
            if row_fill:
                cell.fill = row_fill
            if col == 3:
                cell.number_format = CURRENCY_FORMAT

    # Total row
    total_row = len(sorted_items) + 2
    ws.row_dimensions[total_row].height = 18
    for col in range(1, 5):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = _BLACK_FILL
        cell.border = _THIN_BORDER
        cell.font = _WHITE_BOLD
        if col == 2:
            cell.value = "TOTAL"
            cell.alignment = _LEFT
        elif col == 3:
            cell.value = budget.total_budget
            cell.number_format = CURRENCY_FORMAT
            cell.alignment = _RIGHT
        else:
            cell.alignment = _LEFT

    ws.freeze_panes = "A2"



def _normalize_account_code(code: str) -> str:
    return code.replace(".", "").replace(" ", "").strip()


def _topsheet_prefix_from_account(account: str) -> str:
    clean = _normalize_account_code(account)
    if len(clean) < 2 or not clean[:2].isdigit():
        return ""
    return f"{clean[:2]}00"


def _format_topsheet_code(prefix: str) -> str:
    if len(prefix) == 4 and prefix.isdigit():
        return f"{prefix[:2]}.00"
    return prefix


def _prefix_sort_key(prefix: str) -> tuple[int, str]:
    if prefix.isdigit():
        return (int(prefix), prefix)
    return (9999, prefix)


def _write_detail_budget(ws, budget: ParsedBudget) -> None:
    ws.title = "Detail Budget"

    headers = [
        "Account",
        "Account Description",
        "Description",
        "Amount",
        "Unit",
        "x",
        "Unit 2",
        "Currency",
        "Rate",
        "Unit 3",
        "4X",
        "Unit 4",
        "Subtotal",
    ]

    widths = [12, 34, 40, 10, 10, 6, 10, 10, 12, 10, 6, 10, 14]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _BOLD
        cell.alignment = _LEFT if col in (1, 2, 3) else _CENTER
        cell.border = _NO_BORDER
        cell.fill = _LIGHT_GRAY_FILL

    for col in range(1, 14):
        top_cell = ws.cell(row=1, column=col)
        top_cell.border = Border(
            left=top_cell.border.left,
            right=top_cell.border.right,
            top=_THIN,
            bottom=_THIN,
        )
    ws.cell(row=1, column=1).border = Border(
        left=_THIN,
        right=ws.cell(row=1, column=1).border.right,
        top=ws.cell(row=1, column=1).border.top,
        bottom=ws.cell(row=1, column=1).border.bottom,
    )
    ws.cell(row=1, column=13).border = Border(
        left=ws.cell(row=1, column=13).border.left,
        right=_THIN,
        top=ws.cell(row=1, column=13).border.top,
        bottom=ws.cell(row=1, column=13).border.bottom,
    )

    category_by_account: dict[str, str] = {}
    for item in budget.line_items:
        category_by_account[_normalize_account_code(item.code)] = item.description

    detail_rows = [r for r in budget.detail_rows if r.subtotal > 0]

    topsheet_name_by_prefix = {
        _cavco_to_mm_prefix(entry[0]): entry[1]
        for entry in TOPSHEET_STRUCTURE
        if len(entry) == 2 and entry[0] not in ("HEADER", "BLANK", "GRAND_TOTAL")
    }

    grouped: dict[str, list] = {}
    for row in detail_rows:
        prefix = _topsheet_prefix_from_account(row.account)
        if not prefix:
            continue
        grouped.setdefault(prefix, []).append(row)

    def _set_outline_border(start_row: int, end_row: int) -> None:
        for col in range(1, 14):
            top_cell = ws.cell(row=start_row, column=col)
            top_cell.border = Border(
                left=top_cell.border.left,
                right=top_cell.border.right,
                top=_THIN,
                bottom=top_cell.border.bottom,
            )
            bottom_cell = ws.cell(row=end_row, column=col)
            bottom_cell.border = Border(
                left=bottom_cell.border.left,
                right=bottom_cell.border.right,
                top=bottom_cell.border.top,
                bottom=_THIN,
            )
        for row in range(start_row, end_row + 1):
            left_cell = ws.cell(row=row, column=1)
            left_cell.border = Border(
                left=_THIN,
                right=left_cell.border.right,
                top=left_cell.border.top,
                bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=row, column=13)
            right_cell.border = Border(
                left=right_cell.border.left,
                right=_THIN,
                top=right_cell.border.top,
                bottom=right_cell.border.bottom,
            )

    row_idx = 2
    section_total_rows_by_prefix: dict[str, int] = {}

    section_group_end_prefixes = {
        "A": 600,
        "B": 5900,
        "C": 6900,
        "D": 7200,
    }
    group_labels = {
        "A": 'TOTAL "A" – ABOVE THE LINE',
        "B": 'TOTAL PRODUCTION "B"',
        "C": 'TOTAL POST-PRODUCTION "C"',
        "D": 'TOTAL OTHER "D"',
    }
    emitted_groups: set[str] = set()

    def _emit_group_total(group_key: str) -> None:
        nonlocal row_idx
        if group_key in emitted_groups:
            return

        min_prefix = {
            "A": 100,
            "B": 1000,
            "C": 6000,
            "D": 7000,
        }[group_key]
        max_prefix = section_group_end_prefixes[group_key]

        rows_for_group = [
            row
            for prefix, row in section_total_rows_by_prefix.items()
            if prefix.isdigit() and min_prefix <= int(prefix) <= max_prefix
        ]
        rows_for_group.sort()

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
        label_cell = ws.cell(row=row_idx, column=1, value=group_labels[group_key])
        label_cell.font = _BOLD
        label_cell.alignment = _LEFT
        label_cell.fill = _LIGHT_GRAY_FILL

        if rows_for_group:
            amount_formula = f"=SUM({','.join(f'M{r}' for r in rows_for_group)})"
        else:
            amount_formula = "=0"

        amount_cell = ws.cell(row=row_idx, column=13, value=amount_formula)
        amount_cell.font = _BOLD
        amount_cell.alignment = _RIGHT
        amount_cell.fill = _LIGHT_GRAY_FILL
        amount_cell.number_format = CURRENCY_FORMAT

        _set_outline_border(row_idx, row_idx)
        row_idx += 2
        emitted_groups.add(group_key)

    for prefix in sorted(grouped.keys(), key=_prefix_sort_key):
        if prefix.isdigit():
            prefix_num = int(prefix)
            for group_key in ("A", "B", "C", "D"):
                if group_key in emitted_groups:
                    continue
                if prefix_num > section_group_end_prefixes[group_key]:
                    _emit_group_total(group_key)

        section_start = row_idx

        label = topsheet_name_by_prefix.get(prefix, "")
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=13)
        section_cell = ws.cell(
            row=row_idx,
            column=1,
            value=f"{_format_topsheet_code(prefix)}  {label}".strip(),
        )
        section_cell.font = _BOLD
        section_cell.alignment = _LEFT
        section_cell.fill = _LIGHT_GRAY_FILL
        row_idx += 1

        section_detail_start = row_idx
        for detail in sorted(grouped[prefix], key=lambda r: (_normalize_account_code(r.account), r.description)):
            normalized = _normalize_account_code(detail.account)
            account_desc = category_by_account.get(normalized, "")

            row_data = [
                detail.account,
                account_desc,
                detail.description,
                detail.amount,
                detail.unit,
                detail.x or "x",
                detail.unit2,
                detail.currency,
                detail.rate,
                detail.unit3,
                "x",
                detail.unit4,
                detail.subtotal,
            ]

            for col, value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = _NORMAL
                cell.border = _NO_BORDER
                if col in (1, 2, 3):
                    cell.alignment = _LEFT
                elif col in (4, 13):
                    cell.alignment = _RIGHT
                else:
                    cell.alignment = _CENTER
                if col in (4, 9, 13) and isinstance(value, (int, float)):
                    cell.number_format = CURRENCY_FORMAT

            row_idx += 1

        section_detail_end = row_idx - 1

        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
        total_label = ws.cell(row=row_idx, column=1, value=f"{_format_topsheet_code(prefix)} TOTAL")
        total_label.font = _BOLD
        total_label.alignment = _LEFT
        total_label.fill = _LIGHT_GRAY_FILL

        total_formula = f"=SUM(M{section_detail_start}:M{section_detail_end})"
        total_value = ws.cell(row=row_idx, column=13, value=total_formula)
        total_value.font = _BOLD
        total_value.alignment = _RIGHT
        total_value.fill = _LIGHT_GRAY_FILL
        total_value.number_format = CURRENCY_FORMAT
        section_total_rows_by_prefix[prefix] = row_idx

        _set_outline_border(section_start, row_idx)
        row_idx += 2

    for group_key in ("A", "B", "C", "D"):
        _emit_group_total(group_key)

    # ── Grand Total row ──────────────────────────────────────────────────────
    all_section_rows = sorted(section_total_rows_by_prefix.values())
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=12)
    gt_label = ws.cell(row=row_idx, column=1, value="GRAND TOTAL")
    gt_label.font = _WHITE_BOLD
    gt_label.alignment = _LEFT
    gt_label.fill = _GRAND_TOTAL_FILL

    if all_section_rows:
        refs = ",".join(f"M{r}" for r in all_section_rows)
        gt_val = ws.cell(row=row_idx, column=13, value=f"=SUM({refs})")
    else:
        gt_val = ws.cell(row=row_idx, column=13, value=0)
    gt_val.font = _WHITE_BOLD
    gt_val.alignment = _RIGHT
    gt_val.fill = _GRAND_TOTAL_FILL
    gt_val.number_format = CURRENCY_FORMAT

    for col in range(1, 14):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER

    ws.freeze_panes = "A2"


# ---------------------------------------------------------------------------
# Breakout Budget worksheet builder
# ---------------------------------------------------------------------------

_PERCENTAGE_FORMAT = '0.00%'

# ---------------------------------------------------------------------------
# Breakout Budget bible
# ---------------------------------------------------------------------------
# Maps 4-digit account code to a 6-tuple:
#   (non_prov_out, prov_labour, fed_labour, prov_svc_labour, svc_property, fed_svc_labour)
#
#   non_prov_out  – True  → entire Grand Total is Non-Provincial Spend
#   others        – float → that fraction of Grand Total qualifies for the column
#                   0.0  → not eligible (blank or explicit "-" in source bible)
BREAKOUT_BIBLE: dict[str, tuple] = {
    "0201": (False, 0.65, 0.65, 0.10, 0.0, 0.10),
    "0220": (False, 1.00, 1.00, 0.10, 0.0, 0.10),
    "0225": (False, 1.00, 1.00, 0.0,  0.0, 1.00),
    "0227": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "0295": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "0301": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "0395": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "0401": (False, 0.58, 1.00, 0.58, 0.0, 1.00),
    "0405": (False, 0.58, 1.00, 0.58, 0.0, 1.00),
    "0407": (False, 0.65, 0.65, 0.65, 0.0, 0.65),
    "0408": (False, 0.65, 0.65, 0.65, 0.0, 0.65),
    "0410": (False, 0.65, 0.65, 0.65, 0.0, 0.65),
    "0415": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "0417": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "0460": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "0465": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "0501": (False, 0.65, 0.65, 0.65, 0.0, 0.65),
    "0560": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "0565": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "0660": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "0665": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "1001": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1010": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1025": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1070": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "1075": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "1090": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "1092": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "1095": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "1110": (False, 1.00, 1.00, 0.0,  0.0, 1.00),
    "1170": (False, 1.00, 1.00, 0.0,  0.0, 1.00),
    "1201": (False, 0.80, 0.85, 0.80, 0.0, 0.85),
    "1205": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1210": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1215": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1220": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1223": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1228": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1235": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1240": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1243": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1245": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1248": (False, 0.85, 0.85, 0.85, 0.0, 0.85),
    "1250": (False, 0.85, 0.90, 0.85, 0.0, 0.90),
    "1252": (False, 0.85, 0.90, 0.85, 0.0, 0.90),
    "1261": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1262": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1270": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1301": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1310": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1312": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1320": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1335": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1350": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1420": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1425": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1501": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1505": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1510": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1515": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1530": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1601": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1610": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1693": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "1701": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1710": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1905": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1910": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "1993": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2001": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2010": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2060": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2070": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2093": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2101": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2110": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2112": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2170": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2201": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2205": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2210": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2211": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2212": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2250": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2260": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "2265": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "2270": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2301": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2310": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2320": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2350": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2401": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2410": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2501": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2801": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2810": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2815": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2820": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2830": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2835": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "2840": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2901": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2905": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "2955": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "3105": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3106": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3110": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3150": (False, 1.00, 1.00, 1.00, 1.00, 1.00),
    "3152": (False, 1.00, 1.00, 1.00, 1.00, 1.00),
    "3160": (False, 1.00, 1.00, 1.00, 1.00, 1.00),
    "3195": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3201": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3210": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3215": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3218": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3225": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3245": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3301": (True,  0.0,  0.0,  0.0,  0.0,  0.0),
    "3310": (True,  0.0,  0.0,  0.0,  0.0,  0.0),
    "3320": (True,  0.0,  0.0,  0.0,  0.0,  0.0),
    "3330": (False, 0.0,  0.0,  0.0,  0.0,  0.0),
    "3335": (False, 0.0,  0.0,  0.0,  0.0,  0.0),
    "3350": (True,  0.0,  0.0,  0.0,  0.0,  0.0),
    "3395": (True,  0.0,  0.0,  0.0,  0.0,  0.0),
    "3401": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3405": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3430": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3440": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3445": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3447": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3510": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3515": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3545": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3710": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3730": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3740": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3810": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3830": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3850": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3910": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "3930": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4110": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4130": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4140": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4148": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4210": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4212": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4222": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4240": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4510": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4512": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4515": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4525": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4530": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4595": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4610": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4612": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4630": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4710": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4712": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4795": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4810": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4812": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4816": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4828": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "4830": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "5001": (False, 0.0,  0.0,  0.0,  1.00, 0.0),
    "6001": (False, 0.75, 0.75, 0.75, 0.0,  0.75),
    "6002": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6003": (False, 0.75, 0.75, 0.75, 0.0,  0.75),
    "6010": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6012": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6020": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6042": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6070": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6101": (False, 0.0,  0.0,  0.0,  0.0,  0.0),
    "6110": (False, 0.0,  0.0,  0.0,  0.0,  0.0),
    "6215": (False, 1.00, 1.00, 1.00, 0.0,  1.00),
    "6221": (False, 0.13, 0.13, 0.13, 0.0,  0.13),
    "6240": (False, 0.13, 0.13, 0.13, 0.0,  0.13),
    "6260": (False, 0.13, 0.13, 0.13, 0.0,  0.13),
    "6264": (False, 0.13, 0.13, 0.13, 0.0,  0.13),
    "6310": (False, 0.13, 0.13, 0.0,  1.00, 0.13),
    "6325": (False, 0.13, 0.13, 0.0,  1.00, 0.13),
    "6610": (False, 1.00, 1.00, 1.00, 0.0, 1.00),
    "6670": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "6695": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "6701": (False, 0.13, 0.13, 0.0,  0.0, 0.13),
    "6710": (False, 0.13, 0.13, 0.0,  0.0, 0.13),
    "6730": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "6795": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "6801": (False, 0.13, 0.13, 0.13, 0.0, 0.13),
    "6890": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "6892": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "7001": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "7025": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "7040": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "7095": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "7101": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "7110": (True,  0.65, 0.0,  0.0,  0.0, 0.65),
    "7120": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "7125": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "7130": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "7201": (False, 0.37, 0.70, 0.0,  0.0, 0.70),
    "7210": (True,  1.00, 0.0,  0.0,  0.0, 1.00),
    "7220": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "7230": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
    "7295": (True,  0.0,  0.0,  0.0,  0.0, 0.0),
    "8001": (False, 0.0,  0.0,  0.0,  0.0, 0.0),
}

# ---------------------------------------------------------------------------
# Default descriptions for standard BREAKOUT_BIBLE accounts.
# Maps account_code → human-readable label shown in the Bible Editor when no
# DB override has been saved. Populated from user-provided descriptions.
# ---------------------------------------------------------------------------
BIBLE_DESCRIPTIONS: dict[str, str] = {
    "0201": "Writer(S)",
    "0220": "Story Editor(S)",
    "0225": "Research",
    "0227": "Research/Rights Acquisition (Searches/Clearances)",
    "0295": "Other",
    "0301": "Preliminary Breakdown/Budget/Timing",
    "0395": "Other",
    "0401": "Executive Producer(S)",
    "0405": "Producer(S)",
    "0407": "Line Producer(S)",
    "0408": "Supervising Producer(S)",
    "0410": "Co-Producer(S)",
    "0415": "Associate Producer(S)",
    "0417": "Producer (Other)",
    "0460": "Travel Expenses",
    "0465": "Living Expenses",
    "0501": "Director",
    "0560": "Travel Expenses",
    "0565": "Living Expenses",
    "0660": "Travel Expenses",
    "0665": "Living Expenses",
    "1001": "Principal'S",
    "1010": "Actor'S",
    "1025": "Voice/Off Camera Performances",
    "1070": "Casting Director",
    "1075": "Casting Expenses",
    "1090": "Fringes",
    "1092": "Permits",
    "1095": "Other",
    "1110": "General Background Performers (Detail)",
    "1170": "Background Casting Director",
    "1201": "Production Supervisor",
    "1205": "Production Manager",
    "1210": "Unit Manager",
    "1215": "Location Manager",
    "1220": "1St Assistant Director(S)",
    "1223": "2Nd Assistant Director(S)",
    "1228": "3Rd Assistant Director(S)",
    "1235": "Production Assistants",
    "1240": "Production Coordinator",
    "1243": "Assistant Production Coordinator",
    "1245": "Production Secretary",
    "1248": "Office Production Assistant(S)",
    "1250": "Production Accountant",
    "1252": "Assistant Production Accountant",
    "1261": "Sustainability/Green Advisor",
    "1262": "Advisor(S), E.G. Technical, Edi, Accessibility, Community",
    "1270": "Craft Services/Caterer",
    "1301": "Production Designer",
    "1310": "Art Director",
    "1312": "1St Assistant Art Director",
    "1320": "Art Department Assistant(S)",
    "1335": "Graphic Artist(S)",
    "1350": "Art Department Coordinator",
    "1420": "Head Carpenter",
    "1425": "Carpenter(S)",
    "1501": "Key Set Decorator",
    "1505": "Set Decorator(S)",
    "1510": "Assistant Set Decorator(S)",
    "1515": "Set Dresser(S)",
    "1530": "Labourer(S)",
    "1601": "Key Props (Formerly Property Master)",
    "1610": "Assistant Key Props",
    "1693": "Kit Fees",
    "1701": "Special Effects Supervisor",
    "1710": "Special Effects Assistants(S)",
    "1905": "Key Wardrobe/Set Supervisor",
    "1910": "Assistant Costumer",
    "1993": "Kit Fees",
    "2001": "Key Makeup Artist",
    "2010": "Assistant Makeup Artist",
    "2060": "Assistant Hairstylist",
    "2070": "Special Effects Makeup/Hair",
    "2093": "Kit Fees",
    "2101": "Technical Supervisor",
    "2110": "Lighting Director",
    "2112": "Board Operator",
    "2170": "Teleprompter Operator(S)",
    "2201": "Director Of Photography",
    "2205": "Camera Operator",
    "2210": "1St Assistant Camera",
    "2211": "Digital Imaging Technicans (Dit/Dut/Dmt)",
    "2212": "2Nd Assistant Camera",
    "2250": "Special Equipment Operators",
    "2260": "Additional Camera Operators",
    "2265": "Additional 1St Assistant Camera",
    "2270": "Still Photographer",
    "2301": "Key Lighting Technician/Gaffer",
    "2310": "Second Electric/Best Electric (Formerly Best Boy)",
    "2320": "Electrician(S)",
    "2350": "Generator Operator",
    "2401": "Key Grip",
    "2410": "Second Grip/Best Grip",
    "2501": "Mixer/Sound Recordist",
    "2801": "Office Rentals",
    "2810": "Digital Distribution/Photocopy",
    "2815": "Stationery/Office Supplies",
    "2820": "Telephone/Mobile",
    "2830": "Courior/Postage",
    "2835": "Computer Services/Rentals/Software",
    "2840": "Office Craft Service",
    "2901": "Studio/Backlot Rentals:",
    "2905": "Power",
    "2955": "Other",
    "3105": "Site Rental",
    "3106": "Relocation Expenses",
    "3110": "Site Access:",
    "3150": "Security System/Personnel",
    "3152": "Police/Fire Services",
    "3160": "Public Relations",
    "3195": "Other",
    "3201": "Meal Payment(S)",
    "3210": "Catering",
    "3215": "Craft Service",
    "3218": "Unit Equipment, E.G. Tables/Chairs",
    "3225": "First Aid",
    "3245": "Medical/Insurance/Visa Expenses",
    "3301": "Fares",
    "3310": "Accomodations",
    "3320": "Per Diem(S)",
    "3330": "Transfers (Taxis/Limousines)",
    "3335": "Overweight Baggage",
    "3350": "Customs/Brokerage",
    "3395": "Other",
    "3401": "Production Cars",
    "3405": "Trucks/Vans",
    "3430": "Ev Charging/Fuel:",
    "3440": "Taxis/Rideshare/Shuttle",
    "3445": "Parking",
    "3447": "Mileage Allowance",
    "3510": "Carpentry Equipment Rentals",
    "3515": "Carpentry Purchases",
    "3545": "Green Screen/Backdrop/Cyclorama",
    "3710": "Rentals",
    "3730": "Purchases",
    "3740": "Fabrication",
    "3810": "Rentals",
    "3830": "Purchases",
    "3850": "Picture Vehicles Rentals",
    "3910": "Rentals",
    "3930": "Purchases",
    "4110": "Rentals",
    "4130": "Purchases",
    "4140": "Fabrication",
    "4148": "Repairs/Cleaning",
    "4210": "Makeup Rentals",
    "4212": "Makeup Purchases",
    "4222": "Hair Purchases",
    "4240": "Special Effects",
    "4510": "Basic Package Rentals",
    "4512": "Daily Rentals",
    "4515": "Specialty Rentals",
    "4525": "Video Assist Equipment, E.G. Dit",
    "4530": "Purchases",
    "4595": "Other",
    "4610": "Basic Package Rental",
    "4612": "Daily Rentals",
    "4630": "Purchases",
    "4710": "Basic Package Rental",
    "4712": "Daily Rentals",
    "4795": "Other",
    "4810": "Basic Package Rentals",
    "4812": "Daily Rentals",
    "4816": "Wireless Microphone(S)",
    "4828": "Walkie-Talkies",
    "4830": "Purchases",
    "5001": "Original Scenes",
    "6001": "Post Production Supervisor",
    "6002": "Post Production Coordinator",
    "6003": "Post Production Staff",
    "6010": "Editor",
    "6012": "Assistant Editor(S)",
    "6020": "Sound Designer/Supervisor",
    "6042": "Other Post Production Labour",
    "6070": "Dialogue Transcription",
    "6101": "Editing Rooms",
    "6110": "Editing Equipment (Linear/Non-Linear)",
    "6215": "Online Editing",
    "6221": "Colour Correction",
    "6240": "Graphics",
    "6260": "Backup/Protection Copies",
    "6264": "Distribution Copies",
    "6310": "Voice Over Recording",
    "6325": "Mixing",
    "6610": "Composer(S)",
    "6670": "Music Rights",
    "6695": "Other",
    "6701": "Titles",
    "6710": "Graphics",
    "6730": "Stock Footage",
    "6795": "Other",
    "6801": "Preparation",
    "6890": "Closed Captioning",
    "6892": "Described Video",
    "7001": "Unit Publicist",
    "7025": "Photo Processing And Prints",
    "7040": "Promotion",
    "7095": "Other",
    "7101": "Insurance",
    "7110": "Legal Fees",
    "7120": "Post Production Accounting",
    "7125": "Audit Fee",
    "7130": "Bank Charges",
    "7201": "Corporate Overhead (Administrative Costs)",
    "7210": "Tax Credit Administration",
    "7220": "Interim Financing",
    "7230": "Other Financing",
    "7295": "Other",
    "8001": "Contingency",
}

# Maps account prefix numeric range to a group label for the Groups column
def _derive_group_label(prefix: str) -> str:
    """Return the A/B/C/D group label for a given 4-digit prefix (e.g. '0200')."""
    if not prefix.isdigit():
        return ""
    n = int(prefix)
    if 100 <= n <= 600:
        return 'A \u2013 Above the Line'
    elif 1000 <= n <= 5900:
        return 'B \u2013 Production'
    elif 6000 <= n <= 6900:
        return 'C \u2013 Post-Production'
    elif 7000 <= n <= 7200:
        return 'D \u2013 Other'
    return ""


def _write_breakout_budget(
    ws,
    budget: ParsedBudget,
    overrides: dict | None = None,
    effective_bible: dict | None = None,
) -> None:
    """Generate the Breakout Budget tab.

    Fixed columns (A–I):
      A: Account
      B: Account Description (from line_items / categories)
      C: Description
      D: Agg%
      E: Groups  (from source Excel, or derived A/B/C/D label)
      F: Currency
      G: Subtotal
      H: Fringes  (= G × D for rows with an Agg%)
      I: Grand Total (= G + H)

    Fixed analysis columns (J–AA, columns 10–27; currencies at end):
      J  (10): Foreign              – "FOR" when currency is not CAD/CA
      K  (11): Foreign Spend        – Grand Total if Foreign = "FOR"
      L  (12): Canadian Spend       – Grand Total minus Foreign Spend
      M  (13): Fed Labour %         – basis %
      N  (14): Federal Labour       – calc $
      O  (15): Fed Svc Labour %     – basis %
      P  (16): Federal Svc Labour   – calc $
      Q  (17): Non-Prov             – "OUT" when account is non-provincial
      R  (18): Non-Provincial Spend – calc $
      S  (19): Provincial Spend     – Grand Total minus Non-Provincial Spend
      T  (20): Prov Labour %        – basis %
      U  (21): Provincial Labour    – calc $
      V  (22): Prov Svc Labour %    – basis %
      W  (23): Svc Property %       – basis %
      X  (24): Provincial Svc Labour– calc $
      Y  (25): Services Property    – calc $
      Z  (26): Internals            – Grand Total for Internal OH rows
      AA (27): Meals                – Grand Total for meal/per-diem rows

    Dynamic currency columns (28+):
      One "XXX Grand Total" column per distinct currency found in the data.
    """
    ws.title = "Breakout Budget"

    # ── Build category lookup ────────────────────────────────────────────────
    category_by_account: dict[str, str] = {}
    for item in budget.line_items:
        category_by_account[_normalize_account_code(item.code)] = item.description

    # ── Filter & group detail rows ───────────────────────────────────────────
    # Exclude zero-subtotal rows and "Total Fringes" rows (fringes are calculated in col H)
    detail_rows = [
        r for r in budget.detail_rows
        if r.subtotal > 0 and "total fringes" not in r.description.lower()
    ]

    # ── Discover distinct currencies (sorted) ────────────────────────────────
    seen_currencies: list[str] = []
    for r in detail_rows:
        cur = (r.currency or "").strip().upper()
        if cur and cur not in seen_currencies:
            seen_currencies.append(cur)
    seen_currencies.sort()

    # Fixed columns (A–Q = 1–17; analysis columns 18–35; currencies at end)
    #
    #  7: Amount               – raw amount from budget
    #  8: Unit                 – unit of measure
    #  9: x                    – literal "x"
    # 10: Unit 2               – secondary unit
    # 11: Rate                 – rate per unit
    # 12: Unit 3               – tertiary unit
    # 13: 4X                   – second "times" separator
    # 14: Unit 4               – quaternary unit
    # 18: Foreign              – "FOR" indicator
    # 19: Foreign Spend        – Grand Total when Foreign = "FOR"
    # 20: Canadian Spend       – Grand Total minus Foreign Spend
    # 21: Fed Labour %         – basis %
    # 22: Federal Labour       – calc $
    # 23: Fed Svc Labour %     – basis %
    # 24: Federal Svc Labour   – calc $
    # 25: Non-Prov             – "OUT" indicator
    # 26: Non-Provincial Spend – calc $
    # 27: Provincial Spend     – Grand Total minus Non-Provincial Spend
    # 28: Prov Labour %        – basis %
    # 29: Provincial Labour    – calc $
    # 30: Prov Svc Labour %    – basis %
    # 31: Svc Property %       – basis %
    # 32: Provincial Svc Labour– calc $
    # 33: Services Property    – calc $
    # 34: Internals            – Grand Total for Internal OH rows
    # 35: Meals                – Grand Total for meal/per-diem rows
    # 36+: one column per distinct currency
    foreign_col: int                  = 18
    foreign_spend_calc_col: int       = 19
    canadian_spend_calc_col: int      = 20
    fed_labour_basis_col: int         = 21
    fed_labour_calc_col: int          = 22
    fed_svc_basis_col: int            = 23
    fed_svc_calc_col: int             = 24
    non_prov_basis_col: int           = 25
    non_prov_calc_col: int            = 26
    provincial_spend_calc_col: int    = 27
    prov_labour_basis_col: int        = 28
    prov_labour_calc_col: int         = 29
    prov_svc_basis_col: int           = 30
    svc_property_basis_col: int       = 31
    prov_svc_calc_col: int            = 32
    svc_property_calc_col: int        = 33
    internals_col: int                = 34
    meals_col: int                    = 35

    # Currency grand-total columns come after all fixed columns
    currency_col_map: dict[str, int] = {
        cur: 35 + i + 1 for i, cur in enumerate(seen_currencies)
    }

    # basis_cols order must match raw_basis tuple from BREAKOUT_BIBLE:
    #   [non_prov_out, prov_labour, fed_labour, prov_svc, svc_property, fed_svc]
    basis_cols: list[int] = [
        non_prov_basis_col, prov_labour_basis_col, fed_labour_basis_col,
        prov_svc_basis_col, svc_property_basis_col, fed_svc_basis_col,
    ]
    # calc_cols order must match calc_formulas list in the per-row section below
    calc_cols: list[int] = [
        non_prov_calc_col, prov_labour_calc_col, fed_labour_calc_col,
        prov_svc_calc_col, svc_property_calc_col, fed_svc_calc_col,
        foreign_spend_calc_col,
    ]
    # Pre-compute column letters once (used in per-row formula strings)
    basis_letters = [get_column_letter(c) for c in basis_cols]

    # ── Headers & widths ─────────────────────────────────────────────────────
    headers = [
        "Account",
        "Account Description",
        "Description",
        "Agg%",
        "Groups",
        "Currency",
        "Amount",
        "Unit",
        "x",
        "Unit 2",
        "Rate",
        "Unit 3",
        "4X",
        "Unit 4",
        "Subtotal",
        "Fringes",
        "Grand Total",
        # cols 10–12: Foreign indicator, Foreign Spend, Canadian Spend
        "Foreign",
        "Foreign Spend",
        "Canadian Spend",
        # cols 13–16: Federal
        "Fed Labour %",
        "Federal Labour",
        "Fed Svc Labour %",
        "Federal Services Labour",
        # cols 17–19: Non-Provincial
        "Non-Prov",
        "Non-Provincial Spend",
        "Provincial Spend",
        # cols 20–25: Provincial
        "Prov Labour %",
        "Provincial Labour",
        "Prov Svc Labour %",
        "Svc Property %",
        "Provincial Services Labour",
        "Services Property",
        # cols 26–27: Internals, Meals
        "Internals",
        "Meals",
    ] + [f"{cur} Grand Total" for cur in seen_currencies]

    num_cols = len(headers)
    widths = (
        [12, 34, 40, 8, 28, 10, 10, 8, 4, 10, 10, 8, 4, 10, 14, 14, 14]    # A–Q
        + [10, 18, 18]                           # Foreign, Foreign Spend, Canadian Spend
        + [13, 18, 16, 24]                       # Fed Labour %, Federal Labour, Fed Svc Labour %, Federal Services Labour
        + [10, 22, 22]                           # Non-Prov, Non-Provincial Spend, Provincial Spend
        + [13, 20, 16, 13, 26, 20]              # Prov Labour %, Provincial Labour, Prov Svc Labour %, Svc Property %, Provincial Services Labour, Services Property
        + [16, 14]                               # Internals, Meals
        + [16] * len(seen_currencies)            # currency grand total cols
    )

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

    # ── Column groups (collapsed by default, expandable) ─────────────────────
    _HIDDEN_GROUPS = [
        [4, 5],                          # D–E:   Agg%, Groups
        [7, 8, 9, 10, 11, 12, 13, 14],  # G–N:   Amount, Unit, x, Unit 2, Rate, Unit 3, 4X, Unit 4
        [23, 24],                        # W–X:   Fed Svc Labour %, Federal Services Labour
        [30, 31, 32, 33],                # AD–AG: Prov Svc Labour %, Svc Property %, Provincial Services Labour, Services Property
    ]
    for group in _HIDDEN_GROUPS:
        for col in group:
            cd = ws.column_dimensions[get_column_letter(col)]
            cd.outlineLevel = 1
            cd.hidden = True

    # ── Header row ──────────────────────────────────────────────────────────
    for col, label in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _BOLD
        cell.alignment = _LEFT if col in (1, 2, 3, 5) else _CENTER
        cell.border = _NO_BORDER
        cell.fill = _LIGHT_GRAY_FILL

    # Outer border on header row
    for col in range(1, num_cols + 1):
        c = ws.cell(row=1, column=col)
        c.border = Border(left=c.border.left, right=c.border.right, top=_THIN, bottom=_THIN)
    ws.cell(row=1, column=1).border = Border(
        left=_THIN,
        right=ws.cell(row=1, column=1).border.right,
        top=ws.cell(row=1, column=1).border.top,
        bottom=ws.cell(row=1, column=1).border.bottom,
    )
    ws.cell(row=1, column=num_cols).border = Border(
        left=ws.cell(row=1, column=num_cols).border.left,
        right=_THIN,
        top=ws.cell(row=1, column=num_cols).border.top,
        bottom=ws.cell(row=1, column=num_cols).border.bottom,
    )

    topsheet_name_by_prefix = {
        _cavco_to_mm_prefix(entry[0]): entry[1]
        for entry in TOPSHEET_STRUCTURE
        if len(entry) == 2 and entry[0] not in ("HEADER", "BLANK", "GRAND_TOTAL")
    }

    grouped: dict[str, list] = {}
    for row in detail_rows:
        prefix = _topsheet_prefix_from_account(row.account)
        if not prefix:
            continue
        grouped.setdefault(prefix, []).append(row)

    # ── Outline border helper ────────────────────────────────────────────────
    def _set_outline_border_bb(start_row: int, end_row: int) -> None:
        for col in range(1, num_cols + 1):
            top_cell = ws.cell(row=start_row, column=col)
            top_cell.border = Border(
                left=top_cell.border.left, right=top_cell.border.right,
                top=_THIN, bottom=top_cell.border.bottom,
            )
            bottom_cell = ws.cell(row=end_row, column=col)
            bottom_cell.border = Border(
                left=bottom_cell.border.left, right=bottom_cell.border.right,
                top=bottom_cell.border.top, bottom=_THIN,
            )
        for row in range(start_row, end_row + 1):
            left_cell = ws.cell(row=row, column=1)
            left_cell.border = Border(
                left=_THIN, right=left_cell.border.right,
                top=left_cell.border.top, bottom=left_cell.border.bottom,
            )
            right_cell = ws.cell(row=row, column=num_cols)
            right_cell.border = Border(
                left=right_cell.border.left, right=_THIN,
                top=right_cell.border.top, bottom=right_cell.border.bottom,
            )

    # ── Group totals logic ───────────────────────────────────────────────────
    row_idx = 4  # rows 2–3 reserved for pinned summary (written after grand total is known)
    section_total_rows_by_prefix: dict[str, int] = {}

    section_group_end_prefixes = {"A": 600, "B": 5900, "C": 6900, "D": 7200}
    group_labels = {
        "A": 'TOTAL "A" \u2013 ABOVE THE LINE',
        "B": 'TOTAL PRODUCTION "B"',
        "C": 'TOTAL POST-PRODUCTION "C"',
        "D": 'TOTAL OTHER "D"',
    }
    emitted_groups: set[str] = set()

    def _emit_group_total_bb(group_key: str) -> None:
        nonlocal row_idx
        if group_key in emitted_groups:
            return

        min_prefix = {"A": 100, "B": 1000, "C": 6000, "D": 7000}[group_key]
        max_prefix = section_group_end_prefixes[group_key]

        rows_for_group = sorted(
            row
            for prefix, row in section_total_rows_by_prefix.items()
            if prefix.isdigit() and min_prefix <= int(prefix) <= max_prefix
        )

        # Merge A-F for the label
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        label_cell = ws.cell(row=row_idx, column=1, value=group_labels[group_key])
        label_cell.font = _BOLD
        label_cell.alignment = _LEFT
        label_cell.fill = _LIGHT_GRAY_FILL

        if rows_for_group:
            refs_o = ','.join(f'O{r}' for r in rows_for_group)
            refs_p = ','.join(f'P{r}' for r in rows_for_group)
            refs_q = ','.join(f'Q{r}' for r in rows_for_group)
            subtotal_formula   = f"=SUM({refs_o})"
            fringes_formula    = f"=SUM({refs_p})"
            grandtotal_formula = f"=SUM({refs_q})"
        else:
            subtotal_formula = fringes_formula = grandtotal_formula = "=0"

        for col, formula in zip((15, 16, 17), (subtotal_formula, fringes_formula, grandtotal_formula)):
            c = ws.cell(row=row_idx, column=col, value=formula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        # Currency grand total columns
        for cur, col in currency_col_map.items():
            letter = get_column_letter(col)
            if rows_for_group:
                refs = ','.join(f'{letter}{r}' for r in rows_for_group)
                formula = f"=SUM({refs})"
            else:
                formula = "=0"
            c = ws.cell(row=row_idx, column=col, value=formula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        # Internals column for group total
        internals_letter = get_column_letter(internals_col)
        if rows_for_group:
            refs = ','.join(f'{internals_letter}{r}' for r in rows_for_group)
            internals_formula = f"=SUM({refs})"
        else:
            internals_formula = "=0"
        c = ws.cell(row=row_idx, column=internals_col, value=internals_formula)
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Meals column for group total
        meals_letter = get_column_letter(meals_col)
        if rows_for_group:
            refs = ','.join(f'{meals_letter}{r}' for r in rows_for_group)
            meals_formula = f"=SUM({refs})"
        else:
            meals_formula = "=0"
        c = ws.cell(row=row_idx, column=meals_col, value=meals_formula)
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Bible basis cols + Foreign: blank at aggregate rows (% not meaningful)
        for bcol in [foreign_col] + basis_cols:
            c = ws.cell(row=row_idx, column=bcol, value=None)
            c.fill = _LIGHT_GRAY_FILL

        # Canadian Spend: Grand Total minus Foreign Spend at this aggregate row
        fs_letter = get_column_letter(foreign_spend_calc_col)
        c = ws.cell(row=row_idx, column=canadian_spend_calc_col,
                    value=f"=Q{row_idx}-{fs_letter}{row_idx}")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Provincial Spend: Grand Total minus Non-Provincial Spend at this aggregate row
        np_calc_letter = get_column_letter(non_prov_calc_col)
        c = ws.cell(row=row_idx, column=provincial_spend_calc_col,
                    value=f"=Q{row_idx}-{np_calc_letter}{row_idx}")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Bible calc cols: SUM of section total rows
        for ccol in calc_cols:
            cletter = get_column_letter(ccol)
            if rows_for_group:
                refs = ','.join(f'{cletter}{r}' for r in rows_for_group)
                cformula = f"=SUM({refs})"
            else:
                cformula = "=0"
            c = ws.cell(row=row_idx, column=ccol, value=cformula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        _set_outline_border_bb(row_idx, row_idx)
        row_idx += 2
        emitted_groups.add(group_key)

    # ── Main iteration over prefix groups ───────────────────────────────────
    for prefix in sorted(grouped.keys(), key=_prefix_sort_key):
        if prefix.isdigit():
            prefix_num = int(prefix)
            for group_key in ("A", "B", "C", "D"):
                if group_key in emitted_groups:
                    continue
                if prefix_num > section_group_end_prefixes[group_key]:
                    _emit_group_total_bb(group_key)

        section_start = row_idx
        label = topsheet_name_by_prefix.get(prefix, "")

        # Section header (no merge — fill applied to every cell in the row)
        section_cell = ws.cell(
            row=row_idx,
            column=1,
            value=f"{_format_topsheet_code(prefix)}  {label}".strip(),
        )
        section_cell.font = _BOLD
        section_cell.alignment = _LEFT
        section_cell.fill = _LIGHT_GRAY_FILL
        for col in range(2, num_cols + 1):
            ws.cell(row=row_idx, column=col).fill = _LIGHT_GRAY_FILL
        row_idx += 1

        section_detail_start = row_idx

        for detail in sorted(grouped[prefix], key=lambda r: (_normalize_account_code(r.account), r.description)):
            normalized = _normalize_account_code(detail.account)
            account_desc = category_by_account.get(normalized, "")
            group_label = detail.groups if detail.groups else _derive_group_label(prefix)
            is_fringes_row = detail.agg is not None and detail.agg > 0

            subtotal_col = f"O{row_idx}"
            agg_col = f"D{row_idx}"

            row_data = [
                (detail.account,     _LEFT,   None),
                (account_desc,       _LEFT,   None),
                (detail.description, _LEFT,   None),
                (detail.agg,         _CENTER, _PERCENTAGE_FORMAT),
                (group_label,        _LEFT,   None),
                (detail.currency,    _CENTER, None),
                (detail.amount,      _RIGHT,  _ACCOUNTING_FORMAT),
                (detail.unit,        _CENTER, None),
                (detail.x or "x",   _CENTER, None),
                (detail.unit2,       _CENTER, None),
                (detail.rate,        _RIGHT,  _ACCOUNTING_FORMAT),
                (detail.unit3,       _CENTER, None),
                ("x",                _CENTER, None),
                (detail.unit4,       _CENTER, None),
                (detail.subtotal,    _RIGHT,  _ACCOUNTING_FORMAT),
                (f"={subtotal_col}*{agg_col}" if is_fringes_row else 0, _RIGHT, _ACCOUNTING_FORMAT),
                (f"=O{row_idx}+P{row_idx}", _RIGHT, _ACCOUNTING_FORMAT),
            ]

            for col, (value, align, num_fmt) in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx, column=col, value=value)
                cell.font = _NORMAL
                cell.border = _NO_BORDER
                cell.alignment = align
                if num_fmt:
                    cell.number_format = num_fmt

            # Currency grand total columns: =Q{row} if matching currency, else 0
            row_currency = (detail.currency or "").strip().upper()
            for cur, col in currency_col_map.items():
                value = f"=Q{row_idx}" if row_currency == cur else 0
                c = ws.cell(row=row_idx, column=col, value=value)
                c.font = _NORMAL
                c.border = _NO_BORDER
                c.alignment = _RIGHT
                c.number_format = _ACCOUNTING_FORMAT

            # Internals column: if "Internal OH" appears anywhere in the Groups cell (E), return grand total
            internals_value = f'=IF(ISNUMBER(SEARCH("Internal OH",E{row_idx})),Q{row_idx},0)'
            c = ws.cell(row=row_idx, column=internals_col, value=internals_value)
            c.font = _NORMAL
            c.border = _NO_BORDER
            c.alignment = _RIGHT
            c.number_format = _ACCOUNTING_FORMAT

            # Meals column: account in {2840,3201,3210,3215,3320} OR "Diem" in Description (col C)
            meals_value = (
                f'=IF(OR(ISNUMBER(SEARCH("Diem",C{row_idx})),'
                f'A{row_idx}="2840",A{row_idx}="3201",A{row_idx}="3210",A{row_idx}="3215",A{row_idx}="3320"),'
                f'Q{row_idx},0)'
            )
            c = ws.cell(row=row_idx, column=meals_col, value=meals_value)
            c.font = _NORMAL
            c.border = _NO_BORDER
            c.alignment = _RIGHT
            c.number_format = _ACCOUNTING_FORMAT

            # ── Bible basis columns: visible raw treatment from bible ──
            # Overrides (if any) supersede the bible; None fields fall back to bible.
            _bible = effective_bible if effective_bible is not None else BREAKOUT_BIBLE
            ov = (overrides or {}).get(normalized)
            bible_entry = _bible.get(normalized)
            if bible_entry:
                b_non_prov_out, b_pl, b_fl, b_psl, b_sp, b_fsl = bible_entry
            else:
                b_non_prov_out, b_pl, b_fl, b_psl, b_sp, b_fsl = False, 0.0, 0.0, 0.0, 0.0, 0.0

            def _ov_val(override_val, bible_val):
                """Return override if set, else bible value."""
                return bible_val if override_val is None else override_val

            if ov is not None:
                # Support both Pydantic model and plain dict
                _get = (lambda f: getattr(ov, f)) if hasattr(ov, "__fields__") else (lambda f: ov.get(f))
                non_prov_out = _ov_val(_get("is_non_prov"), b_non_prov_out)
                pl  = _ov_val(_get("prov_labour_pct"),     b_pl)
                fl  = _ov_val(_get("fed_labour_pct"),      b_fl)
                psl = _ov_val(_get("prov_svc_labour_pct"), b_psl)
                sp  = _ov_val(_get("svc_property_pct"),    b_sp)
                fsl = _ov_val(_get("fed_svc_labour_pct"),  b_fsl)
                is_foreign_override = _get("is_foreign")  # None / True / False
            else:
                non_prov_out, pl, fl, psl, sp, fsl = b_non_prov_out, b_pl, b_fl, b_psl, b_sp, b_fsl
                is_foreign_override = None

            raw_basis = [
                "OUT" if non_prov_out else None,
                pl   if pl  > 0 else None,
                fl   if fl  > 0 else None,
                psl  if psl > 0 else None,
                sp   if sp  > 0 else None,
                fsl  if fsl > 0 else None,
            ]

            for bcol, bval in zip(basis_cols, raw_basis):
                c = ws.cell(row=row_idx, column=bcol, value=bval)
                c.font = _NORMAL
                c.border = _NO_BORDER
                c.alignment = _CENTER
                if isinstance(bval, float):
                    c.number_format = _PERCENTAGE_FORMAT

            # Foreign column: Excel formula by default; hard-coded when overridden
            if is_foreign_override is True:
                foreign_value = "FOR"
            elif is_foreign_override is False:
                foreign_value = ""
            else:
                foreign_value = (
                    f'=IF(AND(F{row_idx}<>"",F{row_idx}<>"CAD",F{row_idx}<>"CA"),"FOR","")'
                )
            c = ws.cell(row=row_idx, column=foreign_col, value=foreign_value)
            c.font = _NORMAL
            c.border = _NO_BORDER
            c.alignment = _CENTER

            # Canadian Spend: Grand Total minus Foreign Spend (auditable formula)
            fs_letter = get_column_letter(foreign_spend_calc_col)
            canadian_formula = f"=Q{row_idx}-{fs_letter}{row_idx}"
            c = ws.cell(row=row_idx, column=canadian_spend_calc_col, value=canadian_formula)
            c.font = _NORMAL
            c.border = _NO_BORDER
            c.alignment = _RIGHT
            c.number_format = _ACCOUNTING_FORMAT

            # ── Bible calc columns: IF formulas referencing the basis columns ──
            np_l, pl_l, fl_l, psl_l, sp_l, fsl_l = basis_letters
            for_l = get_column_letter(foreign_col)
            calc_formulas = [
                # Non-Provincial Spend: triggered by either "OUT" (bible) or "FOR" (foreign currency)
                f'=IF(OR({np_l}{row_idx}="OUT",{for_l}{row_idx}="FOR"),Q{row_idx},0)',
                # Foreign rows ("FOR") are ineligible for labour and services property credits
                f'=IF({for_l}{row_idx}="FOR",0,IF({pl_l}{row_idx}>0,O{row_idx}*{pl_l}{row_idx},0))',
                f'=IF({for_l}{row_idx}="FOR",0,IF({fl_l}{row_idx}>0,O{row_idx}*{fl_l}{row_idx},0))',
                f'=IF({for_l}{row_idx}="FOR",0,IF({psl_l}{row_idx}>0,O{row_idx}*{psl_l}{row_idx},0))',
                f'=IF({for_l}{row_idx}="FOR",0,IF({sp_l}{row_idx}>0,Q{row_idx}*{sp_l}{row_idx},0))',
                f'=IF({for_l}{row_idx}="FOR",0,IF({fsl_l}{row_idx}>0,O{row_idx}*{fsl_l}{row_idx},0))',
                # Foreign Spend: Grand Total when the Foreign column reads "FOR"
                f'=IF({for_l}{row_idx}="FOR",Q{row_idx},0)',
            ]
            for ccol, cval in zip(calc_cols, calc_formulas):
                c = ws.cell(row=row_idx, column=ccol, value=cval)
                c.font = _NORMAL
                c.border = _NO_BORDER
                c.alignment = _RIGHT
                c.number_format = _ACCOUNTING_FORMAT

            # Provincial Spend: Grand Total minus Non-Provincial Spend (auditable formula)
            np_calc_letter = get_column_letter(non_prov_calc_col)
            c = ws.cell(row=row_idx, column=provincial_spend_calc_col,
                        value=f"=Q{row_idx}-{np_calc_letter}{row_idx}")
            c.font = _NORMAL
            c.border = _NO_BORDER
            c.alignment = _RIGHT
            c.number_format = _ACCOUNTING_FORMAT

            row_idx += 1

        section_detail_end = row_idx - 1

        # Prefix total row
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        total_label_cell = ws.cell(row=row_idx, column=1, value=f"{_format_topsheet_code(prefix)} TOTAL")
        total_label_cell.font = _BOLD
        total_label_cell.alignment = _LEFT
        total_label_cell.fill = _LIGHT_GRAY_FILL

        for col, letter in zip((15, 16, 17), ("O", "P", "Q")):
            formula = f"=SUM({letter}{section_detail_start}:{letter}{section_detail_end})"
            c = ws.cell(row=row_idx, column=col, value=formula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        # Currency grand total columns for this section total
        for cur, col in currency_col_map.items():
            letter = get_column_letter(col)
            formula = f"=SUM({letter}{section_detail_start}:{letter}{section_detail_end})"
            c = ws.cell(row=row_idx, column=col, value=formula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        # Internals column for this section total
        internals_letter = get_column_letter(internals_col)
        formula = f"=SUM({internals_letter}{section_detail_start}:{internals_letter}{section_detail_end})"
        c = ws.cell(row=row_idx, column=internals_col, value=formula)
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Meals column for this section total
        meals_letter = get_column_letter(meals_col)
        formula = f"=SUM({meals_letter}{section_detail_start}:{meals_letter}{section_detail_end})"
        c = ws.cell(row=row_idx, column=meals_col, value=formula)
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Bible basis cols + Foreign: blank at aggregate rows
        for bcol in [foreign_col] + basis_cols:
            c = ws.cell(row=row_idx, column=bcol, value=None)
            c.fill = _LIGHT_GRAY_FILL

        # Canadian Spend: Grand Total minus Foreign Spend at this section row
        fs_letter = get_column_letter(foreign_spend_calc_col)
        c = ws.cell(row=row_idx, column=canadian_spend_calc_col,
                    value=f"=Q{row_idx}-{fs_letter}{row_idx}")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Provincial Spend: Grand Total minus Non-Provincial Spend at this section row
        np_calc_letter = get_column_letter(non_prov_calc_col)
        c = ws.cell(row=row_idx, column=provincial_spend_calc_col,
                    value=f"=Q{row_idx}-{np_calc_letter}{row_idx}")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Bible calc cols: SUM of detail rows in this section
        for ccol in calc_cols:
            cletter = get_column_letter(ccol)
            formula = f"=SUM({cletter}{section_detail_start}:{cletter}{section_detail_end})"
            c = ws.cell(row=row_idx, column=ccol, value=formula)
            c.font = _BOLD
            c.alignment = _RIGHT
            c.fill = _LIGHT_GRAY_FILL
            c.number_format = _ACCOUNTING_FORMAT

        section_total_rows_by_prefix[prefix] = row_idx
        _set_outline_border_bb(section_start, row_idx)
        row_idx += 2

    # Emit any remaining group totals
    for group_key in ("A", "B", "C", "D"):
        _emit_group_total_bb(group_key)

    # ── Grand Total row ──────────────────────────────────────────────────────
    all_section_rows = sorted(section_total_rows_by_prefix.values())
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
    gt_label = ws.cell(row=row_idx, column=1, value="GRAND TOTAL")
    gt_label.font = _WHITE_BOLD
    gt_label.alignment = _LEFT
    gt_label.fill = _GRAND_TOTAL_FILL

    if all_section_rows:
        for col, letter in zip((15, 16, 17), ("O", "P", "Q")):
            refs = ",".join(f"{letter}{r}" for r in all_section_rows)
            c = ws.cell(row=row_idx, column=col, value=f"=SUM({refs})")
            c.font = _WHITE_BOLD
            c.alignment = _RIGHT
            c.fill = _GRAND_TOTAL_FILL
            c.number_format = _ACCOUNTING_FORMAT
        for cur, col in currency_col_map.items():
            letter = get_column_letter(col)
            refs = ",".join(f"{letter}{r}" for r in all_section_rows)
            c = ws.cell(row=row_idx, column=col, value=f"=SUM({refs})")
            c.font = _WHITE_BOLD
            c.alignment = _RIGHT
            c.fill = _GRAND_TOTAL_FILL
            c.number_format = _ACCOUNTING_FORMAT
        # Internals column for grand total
        internals_letter = get_column_letter(internals_col)
        refs = ",".join(f"{internals_letter}{r}" for r in all_section_rows)
        c = ws.cell(row=row_idx, column=internals_col, value=f"=SUM({refs})")
        c.font = _WHITE_BOLD
        c.alignment = _RIGHT
        c.fill = _GRAND_TOTAL_FILL
        c.number_format = _ACCOUNTING_FORMAT
        # Meals column for grand total
        meals_letter = get_column_letter(meals_col)
        refs = ",".join(f"{meals_letter}{r}" for r in all_section_rows)
        c = ws.cell(row=row_idx, column=meals_col, value=f"=SUM({refs})")
        c.font = _WHITE_BOLD
        c.alignment = _RIGHT
        c.fill = _GRAND_TOTAL_FILL
        c.number_format = _ACCOUNTING_FORMAT
        # Bible basis cols + Foreign: blank on grand total row
        for bcol in [foreign_col] + basis_cols:
            c = ws.cell(row=row_idx, column=bcol, value=None)
            c.font = _WHITE_BOLD
            c.fill = _GRAND_TOTAL_FILL

        # Canadian Spend: Grand Total minus Foreign Spend at the grand total row
        fs_letter = get_column_letter(foreign_spend_calc_col)
        c = ws.cell(row=row_idx, column=canadian_spend_calc_col,
                    value=f"=Q{row_idx}-{fs_letter}{row_idx}")
        c.font = _WHITE_BOLD
        c.alignment = _RIGHT
        c.fill = _GRAND_TOTAL_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Provincial Spend: Grand Total minus Non-Provincial Spend at the grand total row
        np_calc_letter = get_column_letter(non_prov_calc_col)
        c = ws.cell(row=row_idx, column=provincial_spend_calc_col,
                    value=f"=Q{row_idx}-{np_calc_letter}{row_idx}")
        c.font = _WHITE_BOLD
        c.alignment = _RIGHT
        c.fill = _GRAND_TOTAL_FILL
        c.number_format = _ACCOUNTING_FORMAT

        # Bible calc cols: SUM of section total rows
        for ccol in calc_cols:
            cletter = get_column_letter(ccol)
            refs = ",".join(f"{cletter}{r}" for r in all_section_rows)
            c = ws.cell(row=row_idx, column=ccol, value=f"=SUM({refs})")
            c.font = _WHITE_BOLD
            c.alignment = _RIGHT
            c.fill = _GRAND_TOTAL_FILL
            c.number_format = _ACCOUNTING_FORMAT
    else:
        for col in range(17, num_cols + 1):
            c = ws.cell(row=row_idx, column=col, value=0)
            c.font = _WHITE_BOLD
            c.alignment = _RIGHT
            c.fill = _GRAND_TOTAL_FILL
            c.number_format = _ACCOUNTING_FORMAT

    for col in range(1, num_cols + 1):
        ws.cell(row=row_idx, column=col).border = _THIN_BORDER

    # ── Pinned summary rows (written now that grand total row is known) ───────
    grand_total_row = row_idx

    # All dollar-amount columns (accounting format) — used in both summary rows
    accounting_cols = [
        15, 16, 17,
        foreign_spend_calc_col, canadian_spend_calc_col,
        fed_labour_calc_col, fed_svc_calc_col,
        non_prov_calc_col, provincial_spend_calc_col,
        prov_labour_calc_col, prov_svc_calc_col, svc_property_calc_col,
        internals_col, meals_col,
    ] + list(currency_col_map.values())

    # Apply background fill to every cell in both summary rows first
    for col in range(1, num_cols + 1):
        ws.cell(row=2, column=col).fill = _LIGHT_GRAY_FILL
        ws.cell(row=3, column=col).fill = _LIGHT_GRAY_FILL

    # Row 2 — TOTAL: mirror of grand total row
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    lbl2 = ws.cell(row=2, column=1, value="TOTAL")
    lbl2.font = _BOLD
    lbl2.alignment = _LEFT
    lbl2.fill = _LIGHT_GRAY_FILL
    for col in accounting_cols:
        letter = get_column_letter(col)
        c = ws.cell(row=2, column=col, value=f"={letter}{grand_total_row}")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _ACCOUNTING_FORMAT
    _set_outline_border_bb(2, 2)

    # Row 3 — % OF GRAND TOTAL: each accounting col as % of Grand Total (col I)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
    lbl3 = ws.cell(row=3, column=1, value="% OF GRAND TOTAL")
    lbl3.font = _BOLD
    lbl3.alignment = _LEFT
    lbl3.fill = _LIGHT_GRAY_FILL
    for col in accounting_cols:
        letter = get_column_letter(col)
        c = ws.cell(row=3, column=col,
                    value=f"=IFERROR({letter}{grand_total_row}/Q{grand_total_row},0)")
        c.font = _BOLD
        c.alignment = _RIGHT
        c.fill = _LIGHT_GRAY_FILL
        c.number_format = _PERCENTAGE_FORMAT
    _set_outline_border_bb(3, 3)

    # ── Conditional formatting: highlight indicator cells ────────────────────
    # Subtle rose fill for "FOR" (Foreign) and "OUT" (Non-Prov) flags
    _FLAG_FILL = PatternFill(start_color="FFDAD6", end_color="FFDAD6", fill_type="solid")
    _for_col  = get_column_letter(foreign_col)
    _nprov_col = get_column_letter(non_prov_basis_col)
    max_row = grand_total_row
    ws.conditional_formatting.add(
        f"{_for_col}1:{_for_col}{max_row}",
        CellIsRule(operator="equal", formula=['"FOR"'], fill=_FLAG_FILL),
    )
    ws.conditional_formatting.add(
        f"{_nprov_col}1:{_nprov_col}{max_row}",
        CellIsRule(operator="equal", formula=['"OUT"'], fill=_FLAG_FILL),
    )

    # ── Column-group outside borders ─────────────────────────────────────────
    # A medium border box is drawn around each logical column group, running
    # from the header row all the way down to the grand total row.
    _MED = Side(style="medium")
    col_groups = [
        (foreign_col,        fed_labour_calc_col),   # Foreign … Federal Labour     (10–14)
        (non_prov_basis_col, prov_labour_calc_col),  # Non-Prov … Provincial Labour (17–21)
        (internals_col,      num_cols),              # Internals … last currency col (26+)
    ]
    for first_col, last_col in col_groups:
        # Left and right edges — full height of the data
        for row in range(1, grand_total_row + 1):
            lc = ws.cell(row=row, column=first_col)
            lc.border = Border(left=_MED, right=lc.border.right,
                               top=lc.border.top, bottom=lc.border.bottom)
            rc = ws.cell(row=row, column=last_col)
            rc.border = Border(left=rc.border.left, right=_MED,
                               top=rc.border.top, bottom=rc.border.bottom)
        # Top edge — header row
        for col in range(first_col, last_col + 1):
            tc = ws.cell(row=1, column=col)
            tc.border = Border(left=tc.border.left, right=tc.border.right,
                               top=_MED, bottom=tc.border.bottom)
        # Bottom edge — grand total row
        for col in range(first_col, last_col + 1):
            bc = ws.cell(row=grand_total_row, column=col)
            bc.border = Border(left=bc.border.left, right=bc.border.right,
                               top=bc.border.top, bottom=_MED)

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

# Cross-sheet references to Breakout Budget's pinned Row 2 ("TOTAL")
_BB_GRAND_TOTAL     = "='Breakout Budget'!Q2"   # col Q  (17) Grand Total
_BB_PROV_LABOUR     = "='Breakout Budget'!AC2"  # col AC (29) Provincial Labour
_BB_FED_LABOUR      = "='Breakout Budget'!V2"   # col V  (22) Federal Labour
_BB_PROV_SVC_LABOUR = "='Breakout Budget'!AF2"  # col AF (32) Provincial Services Labour
_BB_SVC_PROPERTY    = "='Breakout Budget'!AG2"  # col AG (33) Services Property
_BB_FED_SVC_LABOUR  = "='Breakout Budget'!X2"   # col X  (24) Federal Services Labour
_BB_MEALS           = "='Breakout Budget'!AI2"  # col AI (35) Meals

# Light yellow fill for user-editable input cells
_INPUT_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
_PCT_FORMAT = '0.00%'


def _write_ofttc_sheet(ws, title: str) -> None:
    """Ontario – Full (OFTTC) calculation sheet, linked to Breakout Budget Row 2.

    Layout matches the reference Excel design:
    - Green-fill title block at top
    - Borders ONLY on grey section-header / total rows; data rows borderless
    - Uniform row height throughout
    - Footer summary (Total PC + % of Total Credits) in bold italic
    """
    ws.title = "Ontario - OFTTC"

    ROW_H = 16
    _GFI = _SECTION_HEADER_FILL   # D9D9D9 light grey

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16

    # ── low-level helpers ──────────────────────────────────────────
    def _plain(row, col, value=None, font=None, fill=None, align=None, fmt=None):
        """Cell with NO border."""
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _NORMAL
        c.border    = _NO_BORDER
        c.alignment = align or _LEFT
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    def _lined(row, col, value=None, font=None, fill=None, align=None, fmt=None):
        """Cell on a grey row — fill/font only; borders applied in post-processing."""
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _BOLD
        c.border    = _NO_BORDER
        c.fill      = fill or _GFI
        c.alignment = align or _LEFT
        if fmt: c.number_format = fmt
        return c

    # ── row-level helpers ──────────────────────────────────────────
    def blank_row(row):
        for col in range(1, 4):
            _plain(row, col)

    def grey_row(row, label="", c_val=None, c_fmt=_ACCOUNTING_FORMAT):
        _lined(row, 1, label, fill=_GFI)
        _lined(row, 2, fill=_GFI)
        _lined(row, 3, c_val, fill=_GFI, align=_RIGHT, fmt=c_fmt)

    def label_row(row, text, font=None):
        _plain(row, 1, text, font=font or _NORMAL)
        _plain(row, 2)
        _plain(row, 3)

    def data_row(row, label, b_val=None, c_val=None, bold=False,
                 c_fmt=_ACCOUNTING_FORMAT, b_input=False, c_input=False):
        _plain(row, 1, label, font=_BOLD if bold else _NORMAL)
        b = _plain(row, 2, b_val, align=_CENTER)
        if b_input: b.fill = _INPUT_FILL
        c = _plain(row, 3, c_val, font=_BOLD if bold else _NORMAL,
                   align=_RIGHT, fmt=c_fmt)
        if c_input: c.fill = _INPUT_FILL

    # ══════════════════════════════════════════════════════════════
    # Title block
    # ══════════════════════════════════════════════════════════════
    # Row 1: sheet title with green fill
    for col in range(1, 4):
        _plain(1, col, fill=_TITLE_GREEN_FILL)
    _plain(1, 1, "ONTARIO \u2013 FULL (OFTTC)",
           font=_BOLD_ITALIC, fill=_TITLE_GREEN_FILL)

    blank_row(2)

    # Row 3: production name
    _plain(3, 1, title, font=_BOLD)
    _plain(3, 2); _plain(3, 3)

    # Row 4: sub-title
    label_row(4, "Tax Credit Calculation", font=_BOLD)

    blank_row(5)

    # ══════════════════════════════════════════════════════════════
    # ONTARIO PROVINCIAL TAX CREDIT
    # ══════════════════════════════════════════════════════════════
    grey_row(6, "ONTARIO PROVINCIAL TAX CREDIT")
    blank_row(7)
    label_row(8, "A")

    R_PC = 9
    data_row(R_PC, "Total Production Cost", c_val=_BB_GRAND_TOTAL)

    R_ONT_LAB = 10
    data_row(R_ONT_LAB, "Estimate of Total Ont. Labour", c_val=_BB_PROV_LABOUR)

    data_row(11, "Proportion of labour",
             c_val=f"=C{R_ONT_LAB}/C{R_PC}", c_fmt=_PCT_FORMAT)

    blank_row(12)
    label_row(13, "B")
    blank_row(14)

    R_B_LAB = 15
    data_row(R_B_LAB, "Estimate of total Labour expenditure",
             c_val=f"=C{R_ONT_LAB}")

    R_EQUITY = 16; R_DEFS_P = 17; R_OTHERS = 18
    data_row(R_EQUITY, "Reduction", b_val="Equity",    c_input=True)
    data_row(R_DEFS_P, "",          b_val="Deferrals",  c_input=True)
    data_row(R_OTHERS, "",          b_val="Others",     c_input=True)

    blank_row(19)

    R_NET_P = 20
    data_row(R_NET_P, "Net Production cost", bold=True,
             c_val=(f"=C{R_B_LAB}"
                    f"-IF(ISNUMBER(C{R_EQUITY}),C{R_EQUITY},0)"
                    f"-IF(ISNUMBER(C{R_DEFS_P}),C{R_DEFS_P},0)"
                    f"-IF(ISNUMBER(C{R_OTHERS}),C{R_OTHERS},0)"))

    blank_row(21)
    label_row(22, "C")
    blank_row(23)

    R_ONT_LAB_C = 24
    data_row(R_ONT_LAB_C, "Ontario Labour", c_val=f"=C{R_NET_P}")

    R_GENERAL = 25
    data_row(R_GENERAL, "General OFTTC (\u00d735%)", bold=True,
             c_val=f"=C{R_ONT_LAB_C}*0.35")

    blank_row(26)

    R_REGIONAL = 27
    data_row(R_REGIONAL, "Regional Bonus \u2013 10%", bold=True,
             b_val="y", b_input=True,
             c_val=f'=IF(LOWER(B{R_REGIONAL})="y",C{R_ONT_LAB_C}*0.1,0)')

    blank_row(28)

    R_OFTTC = 29
    grey_row(R_OFTTC, "TOTAL OFTTC", c_val=f"=C{R_GENERAL}+C{R_REGIONAL}")

    blank_row(30)

    data_row(31, "Percentage of budget",
             c_val=f"=C{R_OFTTC}/C{R_PC}", c_fmt=_PCT_FORMAT)

    # ══════════════════════════════════════════════════════════════
    # FEDERAL TAX CREDIT
    # ══════════════════════════════════════════════════════════════
    grey_row(32, "FEDERAL TAX CREDIT")
    blank_row(33)
    blank_row(34)

    R_FED_PC = 35
    data_row(R_FED_PC, "Total Production cost", c_val=f"=C{R_PC}")

    blank_row(36)

    R_ON_TAX = 37
    data_row(R_ON_TAX, "ON Tax Credits", c_val=f"=-C{R_OFTTC}")

    R_FED_DEFS = 38
    data_row(R_FED_DEFS, "Deferrals", c_input=True)

    R_ME = 39
    _plain(R_ME, 1, "50% Meals & Entertainment")
    _plain(R_ME, 2, _BB_MEALS, align=_RIGHT, fmt=_ACCOUNTING_FORMAT)
    _plain(R_ME, 3, f"=IF(ISNUMBER(B{R_ME}),-B{R_ME}*0.5,0)",
           align=_RIGHT, fmt=_ACCOUNTING_FORMAT)

    R_ASSIST = 40
    data_row(R_ASSIST, "Assistance", c_input=True)

    R_NET_F = 41
    data_row(R_NET_F, "Net Production Cost", bold=True,
             c_val=(f"=C{R_FED_PC}+C{R_ON_TAX}"
                    f"-IF(ISNUMBER(C{R_FED_DEFS}),C{R_FED_DEFS},0)"
                    f"+C{R_ME}"
                    f"-IF(ISNUMBER(C{R_ASSIST}),C{R_ASSIST},0)"))

    R_ELIG_A = 42
    data_row(R_ELIG_A, "(A) Eligible production cost", bold=True,
             c_val=f"=C{R_NET_F}*0.6")

    blank_row(43)

    R_FED_LAB = 44
    data_row(R_FED_LAB, "Labour expenditure", c_val=_BB_FED_LABOUR)

    blank_row(45)

    R_LAB_DEFS = 46
    data_row(R_LAB_DEFS, "Deferrals", c_input=True)

    R_SUB = 47
    data_row(R_SUB, "Sub-total",
             c_val=(f"=C{R_FED_LAB}"
                    f"-IF(ISNUMBER(C{R_LAB_DEFS}),C{R_LAB_DEFS},0)"))

    R_OWN = 48
    data_row(R_OWN, "Percentage of ownership",
             c_val=1.0, c_input=True, c_fmt="0%")

    R_NET_LAB_B = 49
    data_row(R_NET_LAB_B, "(B) Net labour expenditure", bold=True,
             c_val=f"=C{R_SUB}*C{R_OWN}")

    blank_row(50)

    R_ELIG_FED = 51
    data_row(R_ELIG_FED, "Eligible cost for Fed. Tax Credit", bold=True,
             c_val=f"=MIN(C{R_ELIG_A},C{R_NET_LAB_B})")

    blank_row(52)

    R_FED_CR = 53
    data_row(R_FED_CR, "Total Federal Tax Credit", bold=True,
             c_val=f"=C{R_ELIG_FED}*0.25")

    data_row(54, "Percentage of budget",
             c_val=f"=C{R_FED_CR}/C{R_PC}", c_fmt=_PCT_FORMAT)

    R_TOTAL = 55
    grey_row(R_TOTAL, "TOTAL TAX CREDIT",
             c_val=f"=C{R_OFTTC}+C{R_FED_CR}")

    # ── Post-processing: outline borders ──────────────────────────────
    # Grey rows: 6 (ONTARIO PROV), R_OFTTC (TOTAL OFTTC),
    #            32 (FEDERAL TAX), R_TOTAL (TOTAL TAX CREDIT)
    CALC_START = 6
    CALC_END   = R_TOTAL
    _GREY_ROWS = {CALC_START, R_OFTTC, 32, CALC_END}

    # For every cell in the calculation block, compute its border from
    # two rules:
    #   1. Grey rows   → top + bottom across full width
    #   2. Outer box   → left on col-A, right on col-C, top on first row,
    #                    bottom on last row
    for row in range(CALC_START, CALC_END + 1):
        is_grey  = row in _GREY_ROWS
        is_start = row == CALC_START
        is_end   = row == CALC_END
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(
                top    = _THIN if (is_grey or is_start) else None,
                bottom = _THIN if (is_grey or is_end)   else None,
                left   = _THIN if col == 1               else None,
                right  = _THIN if col == 3               else None,
            )

    blank_row(56)

    # Footer: Total Production Cost + % of Total Tax Credits (bold italic)
    _plain(57, 1, "Total Production Cost", font=_BOLD_ITALIC)
    _plain(57, 2)
    _plain(57, 3, f"=C{R_PC}", font=_BOLD_ITALIC,
           align=_RIGHT, fmt=_ACCOUNTING_FORMAT)

    _plain(58, 1, "Percentage of Total Tax Credits", font=_BOLD_ITALIC)
    _plain(58, 2)
    _plain(58, 3, f"=C{R_TOTAL}/C{R_PC}", font=_BOLD_ITALIC,
           align=_RIGHT, fmt=_PCT_FORMAT)


def _write_opstc_sheet(ws, title: str) -> None:
    """Ontario – SVS (OPSTC + PSTC) calculation sheet, linked to Breakout Budget Row 2.

    Sections:
      A  Ontario Services Tax Credit (OPSTC @ 21.5% of Ont. Labour + Services)
      B  Federal Services Tax Credit – Ontario Producer (PSTC @ 16% of net labour)
      Total Tax Credit = OPSTC + PSTC
    """
    ws.title = "Ontario - OPSTC"

    ROW_H = 16
    _GFI = _SECTION_HEADER_FILL

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 16

    # ── low-level helpers ──────────────────────────────────────────
    def _plain(row, col, value=None, font=None, fill=None, align=None, fmt=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _NORMAL
        c.border    = _NO_BORDER
        c.alignment = align or _LEFT
        if fill: c.fill = fill
        if fmt:  c.number_format = fmt
        return c

    def _lined(row, col, value=None, font=None, fill=None, align=None, fmt=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _BOLD
        c.border    = _NO_BORDER
        c.fill      = fill or _GFI
        c.alignment = align or _LEFT
        if fmt: c.number_format = fmt
        return c

    # ── row-level helpers ──────────────────────────────────────────
    def blank_row(row):
        for col in range(1, 4):
            _plain(row, col)

    def grey_row(row, label="", c_val=None, c_fmt=_ACCOUNTING_FORMAT):
        _lined(row, 1, label, fill=_GFI)
        _lined(row, 2, fill=_GFI)
        _lined(row, 3, c_val, fill=_GFI, align=_RIGHT, fmt=c_fmt)

    def label_row(row, text, font=None):
        _plain(row, 1, text, font=font or _NORMAL)
        _plain(row, 2)
        _plain(row, 3)

    def data_row(row, label, b_val=None, c_val=None, bold=False,
                 c_fmt=_ACCOUNTING_FORMAT, b_input=False, c_input=False):
        _plain(row, 1, label, font=_BOLD if bold else _NORMAL)
        b = _plain(row, 2, b_val, align=_CENTER)
        if b_input: b.fill = _INPUT_FILL
        c = _plain(row, 3, c_val, font=_BOLD if bold else _NORMAL,
                   align=_RIGHT, fmt=c_fmt)
        if c_input: c.fill = _INPUT_FILL

    # ══════════════════════════════════════════════════════════════
    # Title block
    # ══════════════════════════════════════════════════════════════
    for col in range(1, 4):
        _plain(1, col, fill=_TITLE_GREEN_FILL)
    _plain(1, 1, "ONTARIO \u2013 SVS (OPSTC + PSTC)",
           font=_BOLD_ITALIC, fill=_TITLE_GREEN_FILL)

    blank_row(2)

    _plain(3, 1, title, font=_BOLD)
    _plain(3, 2); _plain(3, 3)

    label_row(4, "Tax Credit Calculation", font=_BOLD)

    blank_row(5)

    # ══════════════════════════════════════════════════════════════
    # ONTARIO SERVICES TAX CREDIT
    # ══════════════════════════════════════════════════════════════
    grey_row(6, "ONTARIO SERVICES TAX CREDIT")
    blank_row(7)
    label_row(8, "A")

    R_PC = 9
    data_row(R_PC, "Total Production Cost", c_val=_BB_GRAND_TOTAL)

    R_ONT_LAB = 10
    data_row(R_ONT_LAB, "Estimate of Total Ont. Labour", c_val=_BB_PROV_SVC_LABOUR)

    R_ONT_SVC = 11
    data_row(R_ONT_SVC, "Estimate of Total Ont. Services", c_val=_BB_SVC_PROPERTY)

    R_SUBTOTAL = 12
    data_row(R_SUBTOTAL, "Subtotal", c_val=f"=C{R_ONT_LAB}+C{R_ONT_SVC}")

    data_row(13, "Proportion of labour",
             c_val=f"=C{R_ONT_LAB}/C{R_PC}", c_fmt=_PCT_FORMAT)

    blank_row(14)
    label_row(15, "B")

    R_B_LAB_SVC = 16
    data_row(R_B_LAB_SVC, "Estimate of total Labour+Services expenditure",
             c_val=f"=C{R_SUBTOTAL}")

    R_EQUITY = 17; R_DEFS_P = 18; R_OTHERS = 19
    data_row(R_EQUITY, "Reduction", b_val="Equity",    c_input=True)
    data_row(R_DEFS_P, "",          b_val="Deferrals",  c_input=True)
    data_row(R_OTHERS, "",          b_val="Others",     c_input=True)

    R_NET_P = 20
    data_row(R_NET_P, "Net Production cost", bold=True,
             c_val=(f"=C{R_B_LAB_SVC}"
                    f"-IF(ISNUMBER(C{R_EQUITY}),C{R_EQUITY},0)"
                    f"-IF(ISNUMBER(C{R_DEFS_P}),C{R_DEFS_P},0)"
                    f"-IF(ISNUMBER(C{R_OTHERS}),C{R_OTHERS},0)"))

    blank_row(21)
    label_row(22, "C")
    blank_row(23)

    R_ONT_LAB_SVC_C = 24
    data_row(R_ONT_LAB_SVC_C, "Ontario Labour + Services", c_val=f"=C{R_NET_P}")

    R_OPSTC = 25
    data_row(R_OPSTC, "General OPSTC (\u00d721.5%)", bold=True,
             c_val=f"=C{R_ONT_LAB_SVC_C}*0.215")

    data_row(26, "Percentage of budget",
             c_val=f"=C{R_OPSTC}/C{R_PC}", c_fmt=_PCT_FORMAT)

    # ══════════════════════════════════════════════════════════════
    # FEDERAL SERVICES TAX CREDIT – ONTARIO PRODUCER
    # ══════════════════════════════════════════════════════════════
    grey_row(27, "FEDERAL SERVICES TAX CREDIT \u2013 ONTARIO PRODUCER")
    blank_row(28)

    R_FED_PC = 29
    data_row(R_FED_PC, "Total Production cost", c_val=f"=C{R_PC}")

    blank_row(30)

    R_FED_SVC_LAB = 31
    data_row(R_FED_SVC_LAB, "Labour expenditure", c_val=_BB_FED_SVC_LABOUR)

    R_ASSIST = 32
    data_row(R_ASSIST, "less  Assistance",
             c_val=f"=C{R_ONT_LAB}*0.215")

    R_SUB = 33
    data_row(R_SUB, "Sub-Total",
             c_val=f"=C{R_FED_SVC_LAB}-C{R_ASSIST}")

    R_OWN = 34
    data_row(R_OWN, "Percentage of ownership",
             c_val=1.0, c_input=True, c_fmt="0%")

    R_NET_LAB = 35
    data_row(R_NET_LAB, "Net Labour Expenditure", bold=True,
             c_val=f"=C{R_SUB}*C{R_OWN}")

    blank_row(36)

    R_FED_CR = 37
    data_row(R_FED_CR, "Total Federal Services Tax Credit", bold=True,
             c_val=f"=C{R_NET_LAB}*0.16")

    data_row(38, "Percentage of budget",
             c_val=f"=C{R_FED_CR}/C{R_PC}", c_fmt=_PCT_FORMAT)

    R_TOTAL = 39
    grey_row(R_TOTAL, "TOTAL TAX CREDIT",
             c_val=f"=C{R_OPSTC}+C{R_FED_CR}",
             c_fmt='#,##0" $"')

    # ── Post-processing: outline borders ──────────────────────────
    CALC_START = 6
    CALC_END   = R_TOTAL
    _GREY_ROWS = {CALC_START, 27, CALC_END}

    for row in range(CALC_START, CALC_END + 1):
        is_grey  = row in _GREY_ROWS
        is_start = row == CALC_START
        is_end   = row == CALC_END
        for col in range(1, 4):
            ws.cell(row=row, column=col).border = Border(
                top    = _THIN if (is_grey or is_start) else None,
                bottom = _THIN if (is_grey or is_end)   else None,
                left   = _THIN if col == 1               else None,
                right  = _THIN if col == 3               else None,
            )

    blank_row(40)

    _plain(41, 1, "Total Production Cost", font=_BOLD_ITALIC)
    _plain(41, 2)
    _plain(41, 3, f"=C{R_PC}", font=_BOLD_ITALIC,
           align=_RIGHT, fmt=_ACCOUNTING_FORMAT)

    _plain(42, 1, "Percentage of Total Tax Credits", font=_BOLD_ITALIC)
    _plain(42, 2)
    _plain(42, 3, f"=C{R_TOTAL}/C{R_PC}", font=_BOLD_ITALIC,
           align=_RIGHT, fmt=_PCT_FORMAT)


def write_bible_excel(entries: list[dict]) -> BytesIO:
    """Export the full breakout bible as a formatted Excel workbook.

    ``entries`` is a list of dicts with keys:
        account_code, description, is_non_prov, prov_labour_pct, fed_labour_pct,
        prov_svc_labour_pct, svc_property_pct, fed_svc_labour_pct,
        is_customized (bool – True if differs from hardcoded default)
    """
    _CUSTOM_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    _HDR_FILL    = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    _HDR_FONT    = Font(bold=True, color="FFFFFF", size=10)
    _PCT_FMT     = "0%"

    wb = Workbook()
    ws = wb.active
    ws.title = "Breakout Bible"

    # ── column widths ────────────────────────────────────────────
    widths = [12, 44, 8, 14, 14, 18, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── title rows ───────────────────────────────────────────────
    ws.row_dimensions[1].height = 20
    t = ws.cell(row=1, column=1, value="Tax Credit Breakout Bible")
    t.font = Font(bold=True, size=13)
    ws.row_dimensions[2].height = 14
    ws.cell(row=2, column=1, value="Yellow rows have been customised from the default values")
    ws.cell(row=2, column=1).font = Font(italic=True, size=9, color="806000")
    ws.row_dimensions[3].height = 8

    # ── header row ───────────────────────────────────────────────
    headers = [
        "Account", "Description", "OUT",
        "Prov Labour %", "Fed Labour %",
        "Prov Svc Labour %", "Svc Property %", "Fed Svc Labour %",
    ]
    ws.row_dimensions[4].height = 18
    for col, hdr in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=hdr)
        c.font      = _HDR_FONT
        c.fill      = _HDR_FILL
        c.alignment = _CENTER
        c.border    = _THIN_BORDER

    # ── data rows ────────────────────────────────────────────────
    for r_offset, entry in enumerate(entries):
        row = r_offset + 5
        ws.row_dimensions[row].height = 15
        fill = _CUSTOM_FILL if entry.get("is_customized") else None

        def _dc(col, value, fmt=None, align=None):
            c = ws.cell(row=row, column=col, value=value)
            c.font      = _NORMAL
            c.border    = _THIN_BORDER
            c.alignment = align or _LEFT
            if fill:    c.fill = fill
            if fmt:     c.number_format = fmt
            return c

        _dc(1, entry["account_code"])
        _dc(2, entry.get("description", "") or "")
        out_cell = _dc(3, "OUT" if entry["is_non_prov"] else "", align=_CENTER)
        if entry["is_non_prov"]:
            out_cell.font = Font(bold=True, size=10, color="C00000")

        for col, key in [
            (4, "prov_labour_pct"),
            (5, "fed_labour_pct"),
            (6, "prov_svc_labour_pct"),
            (7, "svc_property_pct"),
            (8, "fed_svc_labour_pct"),
        ]:
            val = entry.get(key, 0.0) or 0.0
            _dc(col, val if val else None, fmt=_PCT_FMT, align=_RIGHT)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_breakdown_sheet(ws, title: str, num_episodes: int | None = None) -> None:
    """Breakdown overview sheet – clean summary linked to Breakout Budget and Topsheet."""
    ws.title = "Breakdown"

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3    # blank indent
    ws.column_dimensions["B"].width = 32   # labels
    ws.column_dimensions["C"].width = 18   # primary amounts
    ws.column_dimensions["D"].width = 12   # % formulas

    ROW_H    = 16
    FMT_CAD  = CURRENCY_FORMAT   # '#,##0'
    FMT_PCT  = '0.00%'
    _ITALIC  = Font(italic=True, size=10)
    _DIV_BORDER = Border(bottom=_THIN)   # thin line used as section divider

    # ── Helpers ───────────────────────────────────────────────────────────────
    def _c(row, col, value=None, font=None, fill=None, align=None, fmt=None,
           border=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font   or _NORMAL
        c.border    = border or _NO_BORDER
        c.alignment = align  or _LEFT
        if fill   is not None: c.fill         = fill
        if fmt    is not None: c.number_format = fmt
        return c

    def _label(row, text, bold=False, italic=False):
        font = _BOLD if bold else (_ITALIC if italic else _NORMAL)
        _c(row, 2, text, font=font)

    def _amount(row, formula_or_value, pct_formula=None, bold=False,
                pct_fmt=None):
        font = _BOLD if bold else _NORMAL
        _c(row, 3, formula_or_value, font=font, align=_RIGHT, fmt=FMT_CAD)
        if pct_formula is not None:
            _c(row, 4, pct_formula, font=font, align=_RIGHT,
               fmt=pct_fmt if pct_fmt is not None else FMT_PCT)

    def _blank(row):
        ws.row_dimensions[row].height = ROW_H

    def _divider(row):
        """Thin bottom border across label + value columns to separate sections."""
        for col in (2, 3, 4):
            ws.cell(row=row, column=col).border = _DIV_BORDER

    def _sumif_acct(code: str) -> str:
        return (
            f"=SUMIF('Breakout Budget'!$A:$A,\"{code}*\","
            f"'Breakout Budget'!$Q:$Q)"
        )

    def _sumif_desc(text: str) -> str:
        return (
            f"=SUMIF('Breakout Budget'!$C:$C,\"*{text}*\","
            f"'Breakout Budget'!$Q:$Q)"
        )

    # ── Row constants (row 1 = blank breathing-room row) ─────────────────────
    R_HDR      = 2
    R_EPS      = 3
    R_VER      = 4
    R_FX       = 5
    # 6: blank
    R_TOTAL    = 7
    R_PER_EP   = 8
    # 9: blank
    R_NON_PROV = 10
    R_FOR_SP   = 11
    R_BC       = 12
    # 13: blank
    R_CAD_SP   = 14
    R_USD_SP   = 15
    # 16: blank
    R_INT      = 17
    R_TC_EST   = 18
    # 19: blank
    R_EP_FEE   = 20
    R_PR_FEE   = 21
    R_OVERHEAD = 22
    R_PROD_FEE = 23
    # 24: blank
    R_FINANC   = 25
    R_LEGAL    = 26
    R_INSUR    = 27
    # 28: Promotion
    R_PROMO    = 28
    # 29: blank
    R_ONT_CR   = 30
    R_CAVCO    = 31

    _GREEN_FILL = PatternFill(start_color="C0FFCC", end_color="C0FFCC", fill_type="solid")
    _BLUE_FILL  = PatternFill(start_color="A9F8FF", end_color="A9F8FF", fill_type="solid")
    _BOLD_ITALIC = Font(bold=True, italic=True, size=10)

    # ── Row 1: blank breathing-room row ──────────────────────────────────────
    _blank(1)

    # ── Row 2 (R_HDR): header – A2 no fill, B2:D2 black fill + white bold ────
    ws.row_dimensions[R_HDR].height = 20
    _c(R_HDR, 1)   # col A: no fill
    for col in range(2, 5):
        c = ws.cell(row=R_HDR, column=col)
        c.fill   = _BLACK_FILL
        c.border = _NO_BORDER
    _c(R_HDR, 2, title, font=_WHITE_BOLD, fill=_BLACK_FILL)

    # ── Rows 2–4: metadata (italic – user links these cells) ─────────────────
    _label(R_EPS, "Episodes", italic=True)
    if num_episodes is not None:
        _c(R_EPS, 3, num_episodes, font=_ITALIC, align=_RIGHT)
    _label(R_VER, "Budget Version", italic=True)
    _label(R_FX,  "FX",            italic=True)
    _divider(R_FX)

    # ── Row 6: blank ──────────────────────────────────────────────────────────
    _blank(6)

    # ── Row 7: Total Budget (bold) ────────────────────────────────────────────
    _label(R_TOTAL, "Total Budget", bold=True)
    _amount(R_TOTAL, "='Breakout Budget'!Q2", bold=True)

    # ── Row 7: Per Ep ─────────────────────────────────────────────────────────
    _label(R_PER_EP, "Per Ep")
    _amount(R_PER_EP, f"=C{R_TOTAL}/C{R_EPS}")
    _divider(R_PER_EP)

    # ── Row 9: blank ──────────────────────────────────────────────────────────
    _blank(9)

    # ── Row 10: Non-Provincial Spend ─────────────────────────────────────────
    _label(R_NON_PROV, "Non-Provincial Spend")
    _amount(R_NON_PROV, "='Breakout Budget'!Z2",
            pct_formula=f"=C{R_NON_PROV}/C{R_TOTAL}")

    # ── Row 10: Foreign Spend ─────────────────────────────────────────────────
    _label(R_FOR_SP, "Foreign Spend")
    _amount(R_FOR_SP, "='Breakout Budget'!S2",
            pct_formula=f"=C{R_FOR_SP}/C{R_TOTAL}")

    # ── Row 11: B+C (bold) ────────────────────────────────────────────────────
    _label(R_BC, "B+C", bold=True)
    _amount(R_BC,
        "=INDEX('Topsheet'!C:C,"
        "MATCH(\"TOTAL \"\"B\"\" + \"\"C\"\"*\",'Topsheet'!B:B,0))",
        bold=True,
    )
    _divider(R_BC)

    # ── Row 13: blank ────────────────────────────────────────────────────────
    _blank(13)

    # ── Row 14: CAD Spend ────────────────────────────────────────────────────
    _label(R_CAD_SP, "CAD Spend")
    _amount(R_CAD_SP,
        "=IFERROR("
        "INDEX('Breakout Budget'!2:2,MATCH(\"CAD Grand Total\",'Breakout Budget'!1:1,0)),"
        "INDEX('Breakout Budget'!2:2,MATCH(\"CA Grand Total\",'Breakout Budget'!1:1,0)))",
        pct_formula=f"=IFERROR(C{R_CAD_SP}/C{R_TOTAL},0)",
    )

    # ── Row 15: USD Spend ────────────────────────────────────────────────────
    _label(R_USD_SP, "USD Spend")
    _amount(R_USD_SP,
        "=IFERROR("
        "INDEX('Breakout Budget'!2:2,MATCH(\"USD Grand Total\",'Breakout Budget'!1:1,0)),"
        "INDEX('Breakout Budget'!2:2,MATCH(\"US Grand Total\",'Breakout Budget'!1:1,0)))",
        pct_formula=f"=IFERROR(C{R_USD_SP}/C{R_TOTAL},0)",
    )
    _divider(R_USD_SP)

    # ── Row 16: blank ────────────────────────────────────────────────────────
    _blank(16)

    # ── Row 17: Internals – bold+italic, green fill ───────────────────────────
    for col in range(2, 5):
        ws.cell(row=R_INT, column=col).fill = _GREEN_FILL
    _label(R_INT, "Internals", bold=True)
    ws.cell(row=R_INT, column=2).font = _BOLD_ITALIC
    _amount(R_INT,
        "='Breakout Budget'!AH2",
        pct_formula=f"=C{R_INT}/C{R_TOTAL}",
        bold=True,
    )
    for col in (3, 4):
        ws.cell(row=R_INT, column=col).font = _BOLD_ITALIC

    # ── Row 18: Tax Credit Est. – bold+italic, green fill ────────────────────
    for col in range(2, 5):
        ws.cell(row=R_TC_EST, column=col).fill = _GREEN_FILL
    _label(R_TC_EST, "Tax Credit Est.", italic=True)
    ws.cell(row=R_TC_EST, column=2).font = _BOLD_ITALIC
    _divider(R_TC_EST)

    # ── Row 19: blank ────────────────────────────────────────────────────────
    _blank(19)

    # ── Rows 20–23: Fees ─────────────────────────────────────────────────────
    _label(R_EP_FEE, "EP Fee")
    _amount(R_EP_FEE, _sumif_acct("0401"), pct_formula=f"=C{R_EP_FEE}/C{R_BC}")

    _label(R_PR_FEE, "Producer Fee")
    _amount(R_PR_FEE, _sumif_acct("0405"), pct_formula=f"=C{R_PR_FEE}/C{R_BC}")

    _label(R_OVERHEAD, "Overhead")
    _amount(R_OVERHEAD, _sumif_acct("7201"), pct_formula=f"=C{R_OVERHEAD}/C{R_BC}")

    _label(R_PROD_FEE, "Production Fee")
    _amount(R_PROD_FEE,
        _sumif_acct("8001"),
        pct_formula=f"=C{R_PROD_FEE}/(C{R_TOTAL}-C{R_PROD_FEE})",
    )
    _divider(R_PROD_FEE)

    # ── Row 24: blank ────────────────────────────────────────────────────────
    _blank(24)

    # ── Rows 25–28: OPCS + Promotion – blue fill ─────────────────────────────
    for r in (R_FINANC, R_LEGAL, R_INSUR, R_PROMO):
        for col in range(2, 5):
            ws.cell(row=r, column=col).fill = _BLUE_FILL

    _label(R_FINANC, "Interim Financing")
    _amount(R_FINANC, _sumif_acct("7220"), pct_formula=f"=C{R_FINANC}/C{R_TOTAL}")

    _label(R_LEGAL, "Legal Fees")
    _amount(R_LEGAL, _sumif_acct("7110"), pct_formula=f"=C{R_LEGAL}/C{R_TOTAL}")

    _label(R_INSUR, "Insurance")
    _amount(R_INSUR, _sumif_acct("7101"), pct_formula=f"=C{R_INSUR}/C{R_TOTAL}")

    _label(R_PROMO, "Promotion")
    _amount(R_PROMO,
        "=SUMIF('Breakout Budget'!$A:$A,\"7040*\",'Breakout Budget'!$Q:$Q)",
        pct_formula=f"=(300*C{R_EPS})+500",
        pct_fmt=FMT_CAD,
    )
    _divider(R_PROMO)

    # ── Row 29: blank ────────────────────────────────────────────────────────
    _blank(29)

    # ── Row 30: Ontario Creates ───────────────────────────────────────────────
    _label(R_ONT_CR, "Ontario Creates")
    _amount(R_ONT_CR,
        "=SUMIF('Breakout Budget'!$C:$C,\"*OMDC*\",'Breakout Budget'!$Q:$Q)"
        "+SUMIF('Breakout Budget'!$C:$C,\"*Ontario Creates*\",'Breakout Budget'!$Q:$Q)",
        pct_formula=f"=MIN(5000,0.0006*C{R_TOTAL})",
        pct_fmt=FMT_CAD,
    )

    # ── Row 31: CAVCO ─────────────────────────────────────────────────────────
    _label(R_CAVCO, "CAVCO")
    _amount(R_CAVCO,
        _sumif_desc("CAVCO"),
        pct_formula=f"=0.003*C{R_BC}",
        pct_fmt=FMT_CAD,
    )

    # ── Row 32: SODEC ─────────────────────────────────────────────────────────
    R_SODEC = 32
    _label(R_SODEC, "SODEC")
    _amount(R_SODEC,
        _sumif_desc("Sodec"),
        pct_formula=f"=MIN(25000,(C{R_TOTAL}/1000)*4)",
        pct_fmt=FMT_CAD,
    )
    _divider(R_SODEC)

    ws.freeze_panes = "B3"


# ---------------------------------------------------------------------------
# Form 6 sheet
# ---------------------------------------------------------------------------

def _write_form6_sheet(ws) -> None:
    """CAVCO Form 6 (7540-CH-040-0102): Breakdown of Costs – Production.

    Row positions are fixed to match the formula references the user provided;
    do not renumber them.
    """
    ws.title = "Form6"

    ws.column_dimensions["A"].width = 7
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14

    ROW_H = 15
    FMT   = CURRENCY_FORMAT

    _DARK_FILL   = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
    _ITALIC_SM   = Font(italic=True, size=9)
    _WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=10)
    BB = "'Breakout Budget'"

    # ── Cell helper ──────────────────────────────────────────────────────────
    def _c(row, col, value=None, font=None, fill=None, align=None, fmt=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _NORMAL
        c.border    = _NO_BORDER
        c.alignment = align or _LEFT
        if fill is not None: c.fill          = fill
        if fmt  is not None: c.number_format = fmt
        return c

    def _fill_row(row, fill, h=ROW_H):
        ws.row_dimensions[row].height = h
        for col in range(1, 10):
            ws.cell(row=row, column=col).fill = fill

    def _blank(row, h=ROW_H):
        ws.row_dimensions[row].height = h

    # ── Section header (dark or light grey) ──────────────────────────────────
    def _shdr(row, label, acct=None, dark=False, h=16):
        fill = _DARK_FILL if dark else _SECTION_HEADER_FILL
        font = _WHITE_BOLD if dark else _BOLD
        _fill_row(row, fill, h=h)
        if acct is not None:
            _c(row, 1, acct, font=font, fill=fill, align=_CENTER)
        _c(row, 2, label, font=font, fill=fill)

    # ── Column header block (2 rows) ─────────────────────────────────────────
    def _col_hdrs(row):
        _fill_row(row, _SECTION_HEADER_FILL)
        for col, lbl, al in [
            (1, "Account",                 _CENTER),
            (2, "Category",                _LEFT),
            (3, "Personnel Key Creative",  _CENTER),
            (4, "Services Canadian",       _CENTER),
            (5, "Services Non-Canadian",   _CENTER),
            (6, "Labs Canadian",           _CENTER),
            (7, "Labs Non-Canadian",       _CENTER),
            (8, "Other",                   _CENTER),
            (9, "Total",                   _CENTER),
        ]:
            cell = ws.cell(row=row, column=col, value=lbl)
            cell.font      = _BOLD
            cell.fill      = _SECTION_HEADER_FILL
            cell.border    = _NO_BORDER
            cell.alignment = Alignment(horizontal=al.horizontal,
                                       vertical="center", wrap_text=False)

    # ── Data row ─────────────────────────────────────────────────────────────
    def _row(row, acct, label,
             c=None, d=None, e=None, f=None, g=None, h=None, i=None,
             bold=False):
        font = _BOLD if bold else _NORMAL
        ws.row_dimensions[row].height = ROW_H
        if acct is not None:
            a_val = int(acct) if str(acct).lstrip("-").isdigit() else acct
            _c(row, 1, a_val, font=font, align=_CENTER)
        if label is not None:
            _c(row, 2, label, font=font)
        for col_i, val in zip([3, 4, 5, 6, 7, 8, 9], [c, d, e, f, g, h, i]):
            if val is not None:
                cell = ws.cell(row=row, column=col_i, value=val)
                cell.font = font; cell.alignment = _RIGHT
                cell.number_format = FMT; cell.border = _NO_BORDER

    # ── Sub-total row ─────────────────────────────────────────────────────────
    def _sub(row, label, c, d, e, f, g, h, i):
        _fill_row(row, _SECTION_HEADER_FILL, h=16)
        _c(row, 2, label, font=_BOLD, fill=_SECTION_HEADER_FILL)
        for col_i, val in zip([3, 4, 5, 6, 7, 8, 9], [c, d, e, f, g, h, i]):
            cell = ws.cell(row=row, column=col_i, value=val)
            cell.font = _BOLD; cell.fill = _SECTION_HEADER_FILL
            cell.alignment = _RIGHT; cell.number_format = FMT
            cell.border = _NO_BORDER

    # ── Formula shortcuts ────────────────────────────────────────────────────
    def si_text(row_ref, src_col):
        return (f"=SUMIF({BB}!$A:$A,TEXT($A{row_ref},\"00\")&\"??\","
                f"{BB}!${src_col}:${src_col})")

    def si_exact(code, src_col):
        return f"=SUMIF({BB}!$A:$A,\"{code}\",{BB}!${src_col}:${src_col})"

    def sr(r):  # SUM row C:H
        return f"=SUM(C{r}:H{r})"

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 1-13: form title + ATL column headers
    # ══════════════════════════════════════════════════════════════════════════
    ws.row_dimensions[1].height = 18
    _c(1, 2, "Breakdown of Costs - Production", font=_TITLE_FONT)
    _c(1, 9, "Canadian Audio-Visual Certification Office",
       font=_ITALIC_SM, align=_RIGHT)

    ws.row_dimensions[2].height = 13
    _c(2, 2, "(Not required for treaty co-productions)", font=_ITALIC_SM)

    for r in range(3, 7):
        _blank(r)

    _col_hdrs(7)  # header row

    for r in range(8, 14):
        _blank(r, h=5)

    _shdr(13, "ABOVE THE LINE", dark=True)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 14 – Account 1  STORY RIGHTS / ACQUISITIONS
    # ══════════════════════════════════════════════════════════════════════════
    _row(14, 1, "STORY RIGHTS / ACQUISITIONS",
         d=si_text(14, "T"), e=si_text(14, "S"), i=sr(14))

    _blank(15)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 16-19 – Account 2  SCENARIO
    # ══════════════════════════════════════════════════════════════════════════
    _shdr(16, "SCENARIO", acct=2)
    _row(17, 2, "  a)   Remuneration",    c=si_exact("0201", "Q"), i=sr(17))
    _row(18, 2, "  b)   Travel and living expenses",                i=sr(18))
    _row(19, 2, "  c)   Other costs",
         d=si_text(19, "T"), e=si_text(19, "S"), i=sr(19))

    _blank(20)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 21 – Account 3  DEVELOPMENT COSTS
    # ══════════════════════════════════════════════════════════════════════════
    _row(21, 3, "DEVELOPMENT COSTS",
         d=si_text(21, "T"), e=si_text(21, "S"), i=sr(21))

    _blank(22)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 23-42 – Account 4  PRODUCER
    # ══════════════════════════════════════════════════════════════════════════
    _shdr(23, "PRODUCER", acct=4)

    _shdr(24, "4.1  CANADIAN PRODUCER")
    _row(25, None, "  a)   Remuneration",          c=si_exact("0405", "Q"), i=sr(25))
    _row(26, None, "  b)   Travel and living expenses",                      i=sr(26))

    _shdr(27, "4.2  CANADIAN CO-PRODUCER")
    _row(28, None, "  a)   Remuneration",           i=sr(28))
    _row(29, None, "  b)   Travel and living expenses",                      i=sr(29))

    _shdr(30, "4.3  LINE PRODUCER")
    _row(31, None, "  a)   Remuneration",
         d=si_exact("0407", "Q"), i=sr(31))
    _row(32, None, "  b)   Travel and living expenses",                      i=sr(32))

    _shdr(33, "4.4  EXECUTIVE PRODUCER")
    _row(34, None, "  a)   Remuneration",
         d=si_exact("0401", "T"), e=si_exact("0401", "S"), i=sr(34))
    _row(35, None, "  b)   Travel and living expenses",
         d=(f"=SUMIF({BB}!$A:$A,\"0460\",{BB}!$T:$T)"
            f"+SUMIF({BB}!$A:$A,\"0465\",{BB}!$T:$T)"),
         e=(f"=SUMIF({BB}!$A:$A,\"0460\",{BB}!$S:$S)"
            f"+SUMIF({BB}!$A:$A,\"0465\",{BB}!$S:$S)"),
         i=sr(35))

    _shdr(36, "4.5  ASSOCIATE PRODUCER")
    _row(37, None, "  a)   Remuneration",
         d=si_exact("0415", "T"), e=si_exact("0415", "S"), i=sr(37))
    _row(38, None, "  b)   Travel and living expenses",                      i=sr(38))

    _shdr(39, "4.6  OTHER PERSONNEL RELATED TO PRODUCTION")
    _row(40, None, "  a)   Remuneration",
         d=si_exact("0408", "T"), e=si_exact("0408", "S"), i=sr(40))
    _row(41, None, "  b)   Travel and living expenses",                      i=sr(41))

    _row(42, "4.7", "OTHER COSTS", i=sr(42))

    _blank(43)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 44-47 – Account 5  DIRECTOR
    # ══════════════════════════════════════════════════════════════════════════
    _shdr(44, "DIRECTOR", acct=5)
    _row(45, None, "  a)   Remuneration",   c=si_exact("0501", "Q"), i=sr(45))
    _row(46, None, "  b)   Travel and living expenses",
         c=(f"=SUMIF({BB}!$A:$A,\"0560\",{BB}!$Q:$Q)"
            f"+SUMIF({BB}!$A:$A,\"0565\",{BB}!$Q:$Q)"),
         i=sr(46))
    _row(47, None, "  c)   Other costs",    i=sr(47))

    _blank(48)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 49-52 – Account 6  STARS
    # ══════════════════════════════════════════════════════════════════════════
    _shdr(49, "STARS", acct=6)
    _row(50, None, "  a)   Remuneration",   c=si_exact("0601", "Q"), i=sr(50))
    _row(51, None, "  b)   Travel and living expenses",
         c=(f"=SUMIF({BB}!$A:$A,\"0660\",{BB}!$Q:$Q)"
            f"+SUMIF({BB}!$A:$A,\"0665\",{BB}!$Q:$Q)"),
         i=sr(51))
    _row(52, None, "  c)   Other costs",    i=sr(52))

    _blank(53)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 54 – SUB-TOTAL  (ATL)
    # ══════════════════════════════════════════════════════════════════════════
    _sub(54, "SUB-TOTAL",
         "=SUM(C14:C52)", "=SUM(D14:D52)", "=SUM(E14:E52)",
         "=SUM(F14:F52)", "=SUM(G14:G52)", "=SUM(H14:H52)", "=SUM(I14:I52)")

    ws.row_dimensions[55].height = 13
    _c(55, 2, "NOTE :    Type all answers.  Round off to the nearest dollar.",
       font=_ITALIC_SM)
    _c(55, 9, "7540-CH-040-0102 (E)", font=_ITALIC_SM, align=_RIGHT)

    _blank(56)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 57-62 – PRODUCTION section header block
    # ══════════════════════════════════════════════════════════════════════════
    _col_hdrs(57)
    for r in range(58, 63):
        _blank(r, h=5)
    _shdr(62, "PRODUCTION", dark=True)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 63-104 – PRODUCTION data rows
    # Standard pattern: d=T:T TEXT, e=S:S TEXT, i=SUM
    # Special rows handled separately below.
    # ══════════════════════════════════════════════════════════════════════════
    _std_prod = [
        (63, 10, "CAST"),
        (64, 11, "EXTRAS"),
        (65, 12, "PRODUCTION STAFF"),
        # 66 special (has C)
        (67, 14, "CONSTRUCTION LABOUR"),
        (68, 15, "SET DRESSING LABOUR"),
        (69, 16, "PROPERTY LABOUR"),
        (70, 17, "SPECIAL EFFECTS LABOUR"),
        (71, 18, "WRANGLING LABOUR"),
        (72, 19, "WARDROBE LABOUR"),
        (73, 20, "MAKEUP/HAIR LABOUR"),
        (74, 21, "VIDEO TECHNICAL CREW"),
        # 75 special (has C, modified D)
        (76, 23, "ELECTRICAL LABOUR"),
        (77, 24, "GRIP LABOUR"),
        (78, 25, "PRODUCTION SOUND LABOUR"),
        (79, 26, "HEALTH & SAFETY LABOUR"),
        (80, 27, "FRINGE BENEFITS"),
        (81, 28, "PRODUCTION OFFICE EXPENSES"),
        (82, 29, "STUDIO/BACKLOT EXPENSES"),
        (83, 30, "LOCATION OFFICE EXPENSES"),
        (84, 31, "SITE EXPENSES"),
        (85, 32, "UNIT EXPENSES"),
        # 86 special (has C exact)
        (87, 34, "TRANSPORTATION"),
        (88, 35, "CONSTRUCTION MATERIALS"),
        (89, 36, "ART SUPPLIES"),
        (90, 37, "SET DRESSING"),
        (91, 38, "PROPS"),
        (92, 39, "SPECIAL EFFECTS"),
        (93, 40, "HEALTH AND SAFETY PREVENTION"),
        (94, 41, "WARDROBE SUPPLIES"),
        (95, 42, "MAKEUP/HAIR SUPPLIES"),
        (96, 43, "VIDEO STUDIO FACILITIES"),
        (97, 44, "VIDEO REMOTE TECHNICAL FACILITIES"),
        (98, 45, "CAMERA EQUIPMENT"),
        (99, 46, "ELECTRICAL EQUIPMENT"),
        (100, 47, "GRIP EQUIPMENT"),
        (101, 48, "SOUND EQUIPMENT"),
        (102, 49, "SECOND UNIT"),
        (103, 50, "VIDEOTAPE STOCK"),
        (104, 51, "PRODUCTION LABORATORY"),
    ]
    for r, acct, lbl in _std_prod:
        _row(r, acct, lbl, d=si_text(r, "T"), e=si_text(r, "S"), i=sr(r))

    # Row 66 – DESIGN LABOUR (has col C)
    _row(66, 13, "DESIGN LABOUR",
         c=si_exact("1301", "Q"),
         d=si_text(66, "T"), e=si_text(66, "S"), i=sr(66))

    # Row 75 – CAMERA LABOUR (col C exact; col D = TEXT total minus col C and row 86 C)
    _row(75, 22, "CAMERA LABOUR",
         c=si_exact("2201", "Q"),
         d=(f"=SUMIF({BB}!$A:$A,TEXT($A75,\"00\")&\"??\"," \
            f"{BB}!$T:$T)-C75-C86"),
         e=si_text(75, "S"),
         i=sr(75))

    # Row 86 – TRAVEL AND LIVING EXPENSES (col C = two exact codes)
    _row(86, 33, "TRAVEL AND LIVING EXPENSES",
         c=(f"=SUMIF({BB}!$A:$A,\"2263\",{BB}!$Q:$Q)"
            f"+SUMIF({BB}!$A:$A,\"2266\",{BB}!$Q:$Q)"),
         d=si_text(86, "T"), e=si_text(86, "S"), i=sr(86))

    _blank(105)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 106 – SUB-TOTAL  PRODUCTION
    # ══════════════════════════════════════════════════════════════════════════
    _sub(106, "SUB-TOTAL  -  PRODUCTION",
         "=SUM(C63:C104)", "=SUM(D63:D104)", "=SUM(E63:E104)",
         "=SUM(F63:F104)", "=SUM(G63:G104)", "=SUM(H63:H104)", "=SUM(I63:I104)")

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 107-113 – POST-PRODUCTION section header block
    # ══════════════════════════════════════════════════════════════════════════
    _col_hdrs(107)
    for r in range(108, 113):
        _blank(r, h=5)
    _shdr(112, "POST-PRODUCTION", dark=True)
    _shdr(113, "60  EDITORIAL LABOUR")

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 114-134 – POST-PRODUCTION data rows
    # ══════════════════════════════════════════════════════════════════════════

    # 60a  Remuneration – C=exact 6010 Q:Q; F=TEXT(60)?? T:T minus C; G=S:S
    _row(114, 60, "  a)   Remuneration",
         c=si_exact("6010", "Q"),
         f=(f"=SUMIF({BB}!$A:$A,TEXT($A114,\"00\")&\"??\"," \
            f"{BB}!$T:$T)-C114"),
         g=si_text(114, "S"),
         i=sr(114))
    _row(115, 60, "  b)   Travel and living expenses", i=sr(115))
    _row(116, 60, "  c)   Other costs",                i=sr(116))

    # 61  EDITORIAL EQUIPMENT – F and G use TEXT
    _row(117, 61, "EDITORIAL EQUIPMENT",
         f=si_text(117, "T"), g=si_text(117, "S"), i=sr(117))

    # 62  VIDEO POST-PRODUCTION (PICTURE)
    _shdr(118, "62  VIDEO POST-PRODUCTION (PICTURE)")
    _row(119, 62, "  a)   Remuneration",
         f=si_text(119, "T"), g=si_text(119, "S"), i=sr(119))
    _row(120, 62, "  b)   Travel and living expenses", i=sr(120))
    _row(121, 62, "  c)   Other costs",                i=sr(121))

    # 63-65 direct rows
    _row(122, 63, "VIDEO POST-PRODUCTION (SOUND)",
         f=si_text(122, "T"), g=si_text(122, "S"), i=sr(122))
    _row(123, 64, "POST-PRODUCTION LABORATORY",
         f=si_text(123, "T"), g=si_text(123, "S"), i=sr(123))
    _row(124, 65, "FILM POST-PRODUCTION SOUND",
         f=si_text(124, "T"), g=si_text(124, "S"), i=sr(124))

    # 66  MUSIC
    _shdr(125, "66  MUSIC")
    # 66a  Remuneration – C=exact 6610 Q:Q; F=TEXT(66)?? T:T minus C126 and D128; G=S:S
    _row(126, 66, "  a)   Remuneration",
         c=si_exact("6610", "Q"),
         f=(f"=SUMIF({BB}!$A:$A,TEXT($A126,\"00\")&\"??\"," \
            f"{BB}!$T:$T)-C126-D128"),
         g=si_text(126, "S"),
         i=sr(126))
    _row(127, 66, "  b)   Travel and living expenses", i=sr(127))
    _row(128, 66, "  c)   Music rights",
         d=si_exact("6670", "T"), e=si_exact("6670", "S"), i=sr(128))
    _row(129, 66, "  d)   Other costs",                i=sr(129))

    # 67  TITLES/OPTICALS/STOCK FOOTAGE
    _shdr(130, "67  TITLES/OPTICALS/STOCK FOOTAGE")
    # 67a  Titles/Opticals – F=TEXT(67)?? T:T minus H132; G=S:S minus H132
    _row(131, 67, "  a)   Titles/Opticals",
         f=(f"=SUMIF({BB}!$A:$A,TEXT($A131,\"00\")&\"??\"," \
            f"{BB}!$Q:$Q)-H132"),
         g=(f"=SUMIF({BB}!$A:$A,TEXT($A131,\"00\")&\"??\"," \
            f"{BB}!$S:$S)-H132"),
         i=sr(131))
    # 67b  Stock footage – H col only (exact 6730 Q:Q)
    _row(132, None, "  b)   Stock footage",
         h=si_exact("6730", "Q"), i=sr(132))

    _row(133, 68, "VERSIONING",
         f=si_text(133, "T"), g=si_text(133, "S"), i=sr(133))
    _row(134, 69, "AMORTIZATION (SERIES)",            i=sr(134))

    _blank(135)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 136 – SUB-TOTAL  POST-PRODUCTION
    # ══════════════════════════════════════════════════════════════════════════
    _sub(136, "SUB-TOTAL  -  POST-PRODUCTION",
         "=SUM(C113:C134)", "=SUM(D113:D134)", "=SUM(E113:E134)",
         "=SUM(F113:F134)", "=SUM(G113:G134)", "=SUM(H113:H134)", "=SUM(I113:I134)")

    _blank(137)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 138 – OTHERS section header
    # ══════════════════════════════════════════════════════════════════════════
    _shdr(138, "OTHERS", dark=True)

    # ══════════════════════════════════════════════════════════════════════════
    # Rows 139-149 – OTHERS data rows
    # ══════════════════════════════════════════════════════════════════════════

    # 70  UNIT PUBLICITY
    _row(139, 70, "UNIT PUBLICITY",
         d=si_text(139, "T"), e=si_text(139, "S"), i=sr(139))

    # 71  GENERAL EXPENSES
    _shdr(140, "71  GENERAL EXPENSES")
    _row(141, None, "  a)   Insurance",     h=si_exact("7101", "Q"), i=sr(141))
    _row(142, None, "  b)   Legal fees",    h=si_exact("7110", "Q"), i=sr(142))
    _row(143, None, "  c)   Audited costs", h=si_exact("7125", "Q"), i=sr(143))
    _row(144, 71,   "  d)   Other costs",
         h=(f"=SUMIF({BB}!$A:$A,TEXT($A144,\"00\")&\"??\"," \
            f"{BB}!$Q:$Q)-SUM(H141:H143)"),
         i=sr(144))

    # 72  INDIRECT COSTS
    _shdr(145, "72  INDIRECT COSTS")
    _row(146, None, "  a)   Corporate overhead",  h=si_exact("7201", "Q"), i=sr(146))
    _row(147, None, "  b)   Interim financing",   h=si_exact("7220", "Q"), i=sr(147))
    _row(148, 72,   "  c)   Other costs",
         h=(f"=SUMIF({BB}!$A:$A,TEXT($A148,\"00\")&\"??\"," \
            f"{BB}!$Q:$Q)-SUM(H146:H147)"),
         i=sr(148))

    # 81  COMPLETION GUARANTEE
    _row(149, 81, "COMPLETION GUARANTEE", i=sr(149))

    _blank(150)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 151 – SUB-TOTAL  OTHERS
    # ══════════════════════════════════════════════════════════════════════════
    _sub(151, "SUB-TOTAL  -  OTHERS",
         "=SUM(C139:C149)", "=SUM(D139:D149)", "=SUM(E139:E149)",
         "=SUM(F139:F149)", "=SUM(G139:G149)", "=SUM(H139:H149)", "=SUM(I139:I149)")

    _blank(152)

    # ══════════════════════════════════════════════════════════════════════════
    # Row 153 – TOTAL
    # ══════════════════════════════════════════════════════════════════════════
    _sub(153, "TOTAL",
         "=SUM(C54+C106+C136+C151)",
         "=SUM(D54+D106+D136+D151)",
         "=SUM(E54+E106+E136+E151)",
         "=SUM(G54+F106+F136+F151)",   # reproduced exactly as provided
         "=SUM(G54+G106+G136+G151)",
         "=SUM(H54+H106+H136+H151)",
         "=SUM(I54+I106+I136+I151)")

    # Row 154 – aggregate Services / Labs totals (used by % row)
    ws.row_dimensions[154].height = ROW_H
    for col_i, val in [
        (4, "=SUM(D153:E153)"),
        (6, "=SUM(F153:G153)"),
    ]:
        cell = ws.cell(row=154, column=col_i, value=val)
        cell.font = _NORMAL; cell.alignment = _RIGHT
        cell.number_format = FMT; cell.border = _NO_BORDER

    _blank(155)

    # Row 156 – CANADIAN / NON-CANADIAN COSTS
    ws.row_dimensions[156].height = ROW_H
    _c(156, 2, "CANADIAN / NON-CANADIAN COSTS", font=_BOLD)
    for col_i, val in [
        (4, "=SUM(D153*1/D154)"),
        (5, "=SUM(E153*1/D154)"),
        (6, "=SUM(F153*1/F154)"),
        (7, "=SUM(G153*1/F154)"),
    ]:
        cell = ws.cell(row=156, column=col_i, value=val)
        cell.font = _NORMAL; cell.alignment = _RIGHT
        cell.number_format = "0.00%"; cell.border = _NO_BORDER

    # Row 157 – aggregate % totals
    ws.row_dimensions[157].height = ROW_H
    for col_i, val in [
        (4, "=SUM(D156:E156)"),
        (6, "=SUM(F156:G156)"),
    ]:
        cell = ws.cell(row=157, column=col_i, value=val)
        cell.font = _NORMAL; cell.alignment = _RIGHT
        cell.number_format = "0.00%"; cell.border = _NO_BORDER

    # Row 158 – Note
    _c(158, 2, "Note:  All items in italic are video budget accounts.",
       font=_ITALIC_SM)

    # ── Uniform row height ────────────────────────────────────────────────────
    for r in range(1, 159):
        ws.row_dimensions[r].height = ROW_H

    # ── Thin borders on all form body cells (col headers through last row) ────
    for r in range(7, 159):
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = _THIN_BORDER

    # ── Grey fill on non-applicable columns (Labs/Other for ATL & Production,
    #    Services for Post editorial section) ─────────────────────────────────
    _GREY = _SECTION_HEADER_FILL   # D9D9D9
    for r in range(14, 55):        # ATL data rows: Labs Canadian/Non-Can + Other
        for col in (6, 7, 8):
            ws.cell(row=r, column=col).fill = _GREY
    for r in range(63, 106):       # Production rows: Labs Canadian/Non-Can + Other
        for col in (6, 7, 8):
            ws.cell(row=r, column=col).fill = _GREY
    for r in range(114, 128):      # Post editorial rows: Services Canadian/Non-Can
        for col in (4, 5):
            ws.cell(row=r, column=col).fill = _GREY
    for r in range(131, 136):      # Titles/Opticals rows: Personnel + Services cols
        for col in (3, 4, 5):
            ws.cell(row=r, column=col).fill = _GREY
    for col in (6, 7):             # Titles row 132 Stock footage: Labs cols
        ws.cell(132, column=col).fill = _GREY
    for r in range(141, 149):      # Others General/Indirect rows: Personnel + Services + Labs
        for col in (3, 4, 5, 6, 7):
            ws.cell(row=r, column=col).fill = _GREY

    ws.freeze_panes = "C14"


def _write_fs_sheet(ws, title: str, num_episodes: int | None = None, duration_minutes: int | None = None) -> None:
    """Financing Summary (FS) sheet – Calibri font, linked to Breakdown and Ontario-OFTTC tabs."""
    ws.title = "FS"

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 3

    # ── Calibri fonts ──────────────────────────────────────────────────────────
    _CAL_NORMAL = Font(name="Calibri", size=10)
    _CAL_BOLD   = Font(name="Calibri", bold=True, size=10)
    _CAL_ITALIC = Font(name="Calibri", italic=True, size=10)
    _CAL_TITLE  = Font(name="Calibri", bold=True, size=14)

    ROW_H   = 16
    FMT_NUM = _ACCOUNTING_FORMAT   # zeros show as " - "
    FMT_PCT = '0.0%'
    _TS     = Side(style="thin")

    def _cell(row, col, value=None, font=None, align=None, fmt=None, fill=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _CAL_NORMAL
        c.alignment = align or _LEFT
        c.border    = _NO_BORDER
        if fill is not None: c.fill          = fill
        if fmt  is not None: c.number_format = fmt
        return c

    # ── Row 1: Title (from Breakdown B1) ──────────────────────────────────────
    ws.row_dimensions[1].height = 26
    ws.merge_cells("B1:E1")
    c = ws.cell(row=1, column=2, value="='Breakdown'!B2")
    c.font      = _CAL_TITLE
    c.alignment = _CENTER

    # ── Row 2: Series format (e.g. "4X60") – auto-filled if both values provided
    ws.row_dimensions[2].height = 16
    ws.merge_cells("B2:E2")
    subtitle = (f"{num_episodes}X{duration_minutes}"
                if num_episodes is not None and duration_minutes is not None else None)
    c = ws.cell(row=2, column=2, value=subtitle)
    c.font      = _CAL_NORMAL
    c.alignment = _CENTER

    # ── Row 3: spacer ─────────────────────────────────────────────────────────
    ws.row_dimensions[3].height = ROW_H

    # ── Row 4: blank padding (top of box) ─────────────────────────────────────
    ws.row_dimensions[4].height = ROW_H

    # ── Rows 5-8: summary metadata block ──────────────────────────────────────
    R_TOT_B  = 5
    R_PER_EP = 6
    R_N_EPS  = 7
    R_DUR    = 8   # thin bottom separator goes below this row

    _cell(R_TOT_B, 2, "Total budget",            font=_CAL_BOLD)
    _cell(R_TOT_B, 3)
    _cell(R_TOT_B, 4, "='Breakdown'!C7",          font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_TOT_B, 5)

    _cell(R_PER_EP, 2, "Total budget per episode", font=_CAL_BOLD)
    _cell(R_PER_EP, 3)
    _cell(R_PER_EP, 4, "='Breakdown'!C8",          font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_PER_EP, 5)

    _cell(R_N_EPS, 2, "Number of episodes",       font=_CAL_BOLD)
    _cell(R_N_EPS, 3)
    _cell(R_N_EPS, 4, "='Breakdown'!C3",           font=_CAL_BOLD, align=_RIGHT)
    _cell(R_N_EPS, 5)

    _cell(R_DUR, 2, "Duration in minutes",        font=_CAL_BOLD)
    _cell(R_DUR, 3)
    _cell(R_DUR, 4, duration_minutes, font=_CAL_BOLD, align=_RIGHT)
    _cell(R_DUR, 5)

    # ── Row 9: blank ──────────────────────────────────────────────────────────
    ws.row_dimensions[9].height = ROW_H

    # ── Row 10: column headers ─────────────────────────────────────────────────
    R_COL_HDR = 10
    _cell(R_COL_HDR, 2)
    _cell(R_COL_HDR, 3, "Per eps", align=_RIGHT)
    _cell(R_COL_HDR, 4, "Total",   align=_RIGHT)
    _cell(R_COL_HDR, 5, "%",       align=_RIGHT)

    # ── Row 11: TV LICENSES header ─────────────────────────────────────────────
    R_TV_LIC = 11
    _cell(R_TV_LIC, 2, "TV LICENSES", font=_CAL_BOLD)
    _cell(R_TV_LIC, 3)
    _cell(R_TV_LIC, 4)
    _cell(R_TV_LIC, 5)

    # ── Rows 12-14: Broadcasters – names and totals pulled from Sales tab ────────
    # If no broadcaster name is entered in Sales, all cells on that row show blank.
    R_BC1 = 12
    _cell(R_BC1, 2, '=IF(Sales!B4="","",Sales!B4)')
    _cell(R_BC1, 4, f'=IF(Sales!B4="","",IFERROR(Sales!E4,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC1, 3, f'=IF(D{R_BC1}="","",IFERROR(D{R_BC1}/\'Breakdown\'!$C$3,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC1, 5, f'=IF(D{R_BC1}="","",IFERROR(D{R_BC1}/\'Breakdown\'!$C$7,0))', align=_RIGHT, fmt=FMT_PCT)

    R_BC2 = 13
    _cell(R_BC2, 2, '=IF(Sales!B5="","",Sales!B5)')
    _cell(R_BC2, 4, f'=IF(Sales!B5="","",IFERROR(Sales!E5,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC2, 3, f'=IF(D{R_BC2}="","",IFERROR(D{R_BC2}/\'Breakdown\'!$C$3,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC2, 5, f'=IF(D{R_BC2}="","",IFERROR(D{R_BC2}/\'Breakdown\'!$C$7,0))', align=_RIGHT, fmt=FMT_PCT)

    R_BC3 = 14
    _cell(R_BC3, 2, '=IF(Sales!B6="","",Sales!B6)')
    _cell(R_BC3, 4, f'=IF(Sales!B6="","",IFERROR(Sales!E6,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC3, 3, f'=IF(D{R_BC3}="","",IFERROR(D{R_BC3}/\'Breakdown\'!$C$3,0))', align=_RIGHT, fmt=FMT_NUM)
    _cell(R_BC3, 5, f'=IF(D{R_BC3}="","",IFERROR(D{R_BC3}/\'Breakdown\'!$C$7,0))', align=_RIGHT, fmt=FMT_PCT)

    # ── Rows 15-16: blank ─────────────────────────────────────────────────────
    ws.row_dimensions[15].height = ROW_H
    ws.row_dimensions[16].height = ROW_H

    # ── Row 17: Ontario Production Tax Credit (OFTTC tab C29 = TOTAL OFTTC) ───
    R_ONT = 17
    _cell(R_ONT, 2, "Ontario Production Tax Credit")
    _cell(R_ONT, 3,
          "=IFERROR('Ontario - OFTTC'!$C$29/'Breakdown'!$C$3,0)",
          align=_RIGHT, fmt=FMT_NUM)
    _cell(R_ONT, 4, "='Ontario - OFTTC'!$C$29", align=_RIGHT, fmt=FMT_NUM)
    _cell(R_ONT, 5,
          f"=IFERROR(D{R_ONT}/'Breakdown'!$C$7,0)", align=_RIGHT, fmt=FMT_PCT)

    # ── Row 18: Federal Production Tax Credit (OFTTC tab C53 = federal only) ──
    R_FED = 18
    _cell(R_FED, 2, "Federal Production Tax Credit")
    _cell(R_FED, 3,
          "=IFERROR('Ontario - OFTTC'!$C$53/'Breakdown'!$C$3,0)",
          align=_RIGHT, fmt=FMT_NUM)
    _cell(R_FED, 4, "='Ontario - OFTTC'!$C$53", align=_RIGHT, fmt=FMT_NUM)
    _cell(R_FED, 5,
          f"=IFERROR(D{R_FED}/'Breakdown'!$C$7,0)", align=_RIGHT, fmt=FMT_PCT)

    # ── Rows 19-20: blank ─────────────────────────────────────────────────────
    ws.row_dimensions[19].height = ROW_H
    ws.row_dimensions[20].height = ROW_H

    # ── Row 21: DISTRIBUTION GUARANTEE = budget − licenses − tax credits ──────
    R_DIST = 21
    dist_d = (
        f"='Breakdown'!$C$7"
        f"-IF(ISNUMBER(D{R_BC1}),D{R_BC1},0)"
        f"-IF(ISNUMBER(D{R_BC2}),D{R_BC2},0)"
        f"-IF(ISNUMBER(D{R_BC3}),D{R_BC3},0)"
        f"-IFERROR(D{R_ONT},0)-IFERROR(D{R_FED},0)"
    )
    _cell(R_DIST, 2, "DISTRIBUTION GUARANTEE", font=_CAL_BOLD)
    _cell(R_DIST, 3,
          f"=IFERROR(D{R_DIST}/'Breakdown'!$C$3,0)",
          font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_DIST, 4, dist_d, font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_DIST, 5,
          f"=IFERROR(D{R_DIST}/'Breakdown'!$C$7,0)",
          font=_CAL_BOLD, align=_RIGHT, fmt=FMT_PCT)

    # ── Row 22: Cineflix Rights (sub-label, italic) ───────────────────────────
    R_CINEF = 22
    _cell(R_CINEF, 2, "Cineflix Rights", font=_CAL_ITALIC)
    _cell(R_CINEF, 3)
    _cell(R_CINEF, 4)
    _cell(R_CINEF, 5)

    # ── Rows 23-24: blank ─────────────────────────────────────────────────────
    ws.row_dimensions[23].height = ROW_H
    ws.row_dimensions[24].height = ROW_H

    # ── Row 25: Total ──────────────────────────────────────────────────────────
    R_TOTAL = 25
    _cell(R_TOTAL, 2, "Total",           font=_CAL_BOLD)
    _cell(R_TOTAL, 3, "='Breakdown'!C8", font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_TOTAL, 4, "='Breakdown'!C8", font=_CAL_BOLD, align=_RIGHT, fmt=FMT_NUM)
    _cell(R_TOTAL, 5, 1.0,               font=_CAL_BOLD, align=_RIGHT, fmt=FMT_PCT)

    # ── Row 26: blank bottom padding (inside box) ─────────────────────────────
    ws.row_dimensions[26].height = ROW_H

    # ── Borders: outer box (A4:F26), separator above TV LICENSES (row 10),
    #            Total separator above row 25 ─────────────────────────────────
    BOX_R1, BOX_R2 = 4, 26
    BOX_C1, BOX_C2 = 1, 6

    for row in range(BOX_R1, BOX_R2 + 1):
        for col in range(BOX_C1, BOX_C2 + 1):
            inner  = BOX_C1 < col < BOX_C2   # cols B–E (not the outer A/F edge cols)
            top    = _TS if (row == BOX_R1) or (row == R_TOTAL   and inner) else None
            bottom = _TS if (row == BOX_R2) or (row == R_COL_HDR and inner) else None
            left   = _TS if col == BOX_C1                                   else None
            right  = _TS if col == BOX_C2                                   else None
            if any((top, bottom, left, right)):
                ws.row_dimensions[row].height = ROW_H
                ws.cell(row=row, column=col).border = Border(
                    top=top, bottom=bottom, left=left, right=right
                )


def _write_sales_sheet(ws) -> None:
    """Sales sheet – Presales, Sales Projections, and Participation sections."""
    ws.title = "Sales"

    _CAL_NORMAL = Font(name="Calibri", size=10)
    _CAL_BOLD   = Font(name="Calibri", bold=True, size=10)
    _CAL_ITALIC = Font(name="Calibri", italic=True, size=9)
    _CAL_SMALL  = Font(name="Calibri", size=9)

    ROW_H   = 16
    FMT_NUM = _ACCOUNTING_FORMAT
    FMT_PCT = '0.00%'

    _YELLOW_FILL  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    _CYAN_FILL    = PatternFill(start_color="CCFFFF", end_color="CCFFFF", fill_type="solid")
    _GRAY_FILL    = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _LT_GREEN_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Column widths
    ws.column_dimensions["A"].width = 18   # Broadcaster label
    ws.column_dimensions["B"].width = 24   # Name
    ws.column_dimensions["C"].width = 12   # Currency
    ws.column_dimensions["D"].width = 16   # License fee local
    ws.column_dimensions["E"].width = 16   # License Fee CAD
    ws.column_dimensions["F"].width = 16   # Per Ep Local
    ws.column_dimensions["G"].width = 2    # spacer
    ws.column_dimensions["H"].width = 22   # FX label
    ws.column_dimensions["I"].width = 12   # FX value

    def _cell(row, col, value=None, font=None, align=None, fmt=None, fill=None, border=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font   or _CAL_NORMAL
        c.alignment = align  or _LEFT
        if fill   is not None: c.fill          = fill
        if fmt    is not None: c.number_format = fmt
        if border is not None: c.border        = border
        return c

    # ── FX Table (cols H–I, rows 1–6) ─────────────────────────────────────────
    _cell(1, 8, "FX Table (from fixed data tab)", font=_CAL_BOLD)
    for r, (ccy, rate) in enumerate(
        [("CAD", 1.00), ("USD", 1.35), ("GBP", 1.80), ("EUR", 1.60)], start=2
    ):
        _cell(r, 8, ccy,  font=_CAL_BOLD)
        c = _cell(r, 9, rate, font=_CAL_NORMAL, align=_RIGHT, fmt="0.00")
        c.fill = _GRAY_FILL

    # ── PRESALES section header ────────────────────────────────────────────────
    R_PRESALES = 1
    ws.merge_cells(f"A{R_PRESALES}:F{R_PRESALES}")
    c = ws.cell(row=R_PRESALES, column=1, value="Presales")
    c.font  = Font(name="Calibri", bold=True, size=11, color="000000")
    c.fill  = PatternFill(start_color="A8FFC1", end_color="A8FFC1", fill_type="solid")
    c.alignment = _LEFT
    ws.row_dimensions[R_PRESALES].height = ROW_H

    # ── Column headers row ────────────────────────────────────────────────────
    R_HDR = 3
    for col, label in [
        (2, "Name"), (3, "Currency"), (4, "License fee local"),
        (5, "License Fee CAD"), (6, "Per Ep Local"),
    ]:
        c = _cell(R_HDR, col, label, font=_CAL_BOLD, align=_CENTER)
        c.border = _BOTTOM_BORDER

    # ── Broadcaster rows (3 rows, yellow input cells) ─────────────────────────
    BC_ROWS = [4, 5, 6]
    for i, row in enumerate(BC_ROWS, start=1):
        _cell(row, 1, f"Broadcaster {i}", font=_CAL_BOLD)
        # Name (yellow – user input)
        _cell(row, 2, None, fill=_YELLOW_FILL)
        # Currency (yellow – user input)
        _cell(row, 3, None, fill=_YELLOW_FILL)
        # License fee local (yellow – user input)
        _cell(row, 4, None, fill=_YELLOW_FILL, fmt=FMT_NUM)
        # License Fee CAD – formula: local * FX rate matched from FX table
        # FX table: H2:H5 = CAD/USD/GBP/EUR, I2:I5 = rates
        cad_formula = (
            f"=IFERROR(D{row}*INDEX($I$2:$I$5,MATCH(C{row},$H$2:$H$5,0)),0)"
        )
        _cell(row, 5, cad_formula, fmt=FMT_NUM)
        # Per Ep Local – license fee local divided by episode count from Breakdown tab
        _cell(row, 6, f"=IFERROR(D{row}/Breakdown!$C$3,\"\")", fmt=FMT_NUM)

    # ── Total License Fee CAD ─────────────────────────────────────────────────
    R_TOTAL = 8
    ws.row_dimensions[R_TOTAL].height = ROW_H
    c = _cell(R_TOTAL, 5, f"=SUM(E{BC_ROWS[0]}:E{BC_ROWS[-1]})", fmt=FMT_NUM, font=_CAL_BOLD)
    c.border = Border(top=_THIN)

    # ── SALES PROJECTIONS section ─────────────────────────────────────────────
    R_SP = 10
    ws.merge_cells(f"A{R_SP}:F{R_SP}")
    c = ws.cell(row=R_SP, column=1, value="Sales Projections")
    c.font  = Font(name="Calibri", bold=True, size=11, color="000000")
    c.fill  = PatternFill(start_color="A8FFC1", end_color="A8FFC1", fill_type="solid")
    c.alignment = _LEFT
    ws.row_dimensions[R_SP].height = ROW_H

    R_SFC = 12
    _cell(R_SFC, 1, "Sales forecast by team (taken from sales projection tab)", font=_CAL_NORMAL)
    _cell(R_SFC, 4, None, fill=_YELLOW_FILL, fmt=FMT_NUM)

    R_SFPH = 13
    _cell(R_SFPH, 1, "Sales forecast per hour", font=_CAL_NORMAL)
    _cell(R_SFPH, 4, None, fill=_YELLOW_FILL, fmt=FMT_NUM)

    # ── PARTICIPATION section ─────────────────────────────────────────────────
    R_PART = 16
    ws.merge_cells(f"A{R_PART}:F{R_PART}")
    c = ws.cell(row=R_PART, column=1, value="Participation")
    c.font  = Font(name="Calibri", bold=True, size=11, color="000000")
    c.fill  = PatternFill(start_color="A8FFC1", end_color="A8FFC1", fill_type="solid")
    c.alignment = _LEFT
    ws.row_dimensions[R_PART].height = ROW_H

    # Commission / expense / back-end rates
    R_CR  = 18
    R_DER = 19
    R_BER = 20

    _cell(R_CR,  1, "Commission rate for CR",    font=_CAL_NORMAL)
    c = _cell(R_CR,  3, 0.25, fill=_YELLOW_FILL, fmt=FMT_PCT)
    _cell(R_CR,  5, "Default: 25%", font=_CAL_SMALL)

    _cell(R_DER, 1, "Distribution expense rate", font=_CAL_NORMAL)
    c = _cell(R_DER, 3, 0.05, fill=_YELLOW_FILL, fmt=FMT_PCT)
    _cell(R_DER, 5, "Default: 5%", font=_CAL_SMALL)

    _cell(R_BER, 1, "Back-end rate *",            font=_CAL_NORMAL)
    _cell(R_BER, 3, "=SUM(C23:C29)", fill=_YELLOW_FILL, fmt=FMT_PCT)

    # Participation table header
    R_PHDR = 22
    _cell(R_PHDR, 2, "Name", font=_CAL_BOLD, align=_CENTER)
    c = _cell(R_PHDR, 3, "Rate", font=_CAL_BOLD, align=_CENTER)

    PART_ROWS = [
        (23, "Broadcaster 1"),
        (24, "Broadcaster 2"),
        (25, "Broadcaster 3"),
        (26, "Broadcaster 4"),
        (27, "Host"),
        (28, "Other"),
        (29, None),
    ]
    for row, label in PART_ROWS:
        if label:
            _cell(row, 1, label, font=_CAL_NORMAL)
        _cell(row, 2, None, fill=_YELLOW_FILL)
        _cell(row, 3, 0.00, fill=_YELLOW_FILL, fmt=FMT_PCT)

    # Footnote
    R_NOTE = 31
    ws.merge_cells(f"A{R_NOTE}:F{R_NOTE}")
    c = ws.cell(row=R_NOTE, column=1,
                value="* rate at which partners participate in profit on title, assumed as being "
                      "applied against international sales, less advance, less commish and dist expenses.")
    c.font      = Font(name="Calibri", italic=True, size=9)
    c.alignment = _LEFT
    ws.row_dimensions[R_NOTE].height = 30


def _write_sodec_sheet(ws) -> None:
    """SODEC – Quebec Provincial + QC Producer Federal Tax Credit calculation sheet."""
    ws.title = "Sodec"

    _CAL_NORMAL  = Font(name="Calibri", size=10)
    _CAL_BOLD    = Font(name="Calibri", bold=True, size=10)
    _CAL_TITLE   = Font(name="Calibri", bold=True, size=13)
    _CAL_SUBTITLE = Font(name="Calibri", size=10)

    ROW_H    = 16
    FMT_NUM  = _ACCOUNTING_FORMAT
    FMT_PCT  = '0.00%'
    FMT_PCT1 = '0.0%'

    _YELLOW_FILL  = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    _BLACK_HDR    = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    _BOLD_WHITE   = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    _TOTAL_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 16

    def _cell(row, col, value=None, font=None, align=None, fmt=None, fill=None, border=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font  or _CAL_NORMAL
        c.alignment = align or _LEFT
        c.border    = _NO_BORDER
        if fill   is not None: c.fill          = fill
        if fmt    is not None: c.number_format = fmt
        if border is not None: c.border        = border
        return c

    def _hdr(row, label):
        """Black section header spanning both columns."""
        ws.merge_cells(f"A{row}:B{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font      = _BOLD_WHITE
        c.fill      = _BLACK_HDR
        c.alignment = _LEFT
        ws.row_dimensions[row].height = ROW_H

    def _row(row, label, value=None, bold=False, yellow=False, fmt=FMT_NUM,
             pct=False, total=False):
        font = _CAL_BOLD if (bold or total) else _CAL_NORMAL
        fill = _YELLOW_FILL if yellow else (_TOTAL_FILL if total else None)
        _cell(row, 1, label, font=font, fill=fill)
        c = _cell(row, 2, value, font=font, align=_RIGHT,
                  fmt=(FMT_PCT if pct else fmt), fill=fill)
        if total:
            c.border = Border(top=_THIN, bottom=_THIN)
        return c

    # ── Rows 1-2: Title ───────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 22
    _cell(1, 1, "='Breakdown'!B2", font=_CAL_TITLE)
    _cell(2, 1, "TAX CREDITS CALCULATION", font=_CAL_SUBTITLE)

    # ── Row 4: Quebec Provincial header ──────────────────────────────────────
    _hdr(4, "QUEBEC PROVINCIAL TAX CREDIT")

    # Row 6: Production costs — linked to Breakdown total budget
    R_PROD = 6
    _row(R_PROD, "Production costs", "='Breakdown'!$C$7")

    # Row 7: Government assistance
    R_GOVTA = 7
    _row(R_GOVTA, "Government assistance", None)

    # Row 8: Eligible production costs = prod - govt assistance
    R_ELIG_PROD = 8
    _row(R_ELIG_PROD, "Eligible production costs",
         f"=B{R_PROD}-IFERROR(B{R_GOVTA},0)")

    # Row 9: (A) Eligible cap = 65% of eligible production costs
    R_CAP_A = 9
    _row(R_CAP_A, "(A) Eligible cap", f"=B{R_ELIG_PROD}*0.65")

    # Row 11: Total labour — Provincial Labour from Breakout Budget AC2
    R_LABOUR = 11
    _row(R_LABOUR, "Total labour including deferred amounts", "='Breakout Budget'!$AC$2")

    # Row 12: Government assistance (second instance)
    R_GOVTB = 12
    _row(R_GOVTB, "Government assistance", None)

    # Row 13: Deferred amounts
    R_DEFER = 13
    _row(R_DEFER, "Deferred amounts", None)

    # Row 14: (B) Eligible labour = labour - govt assistance - deferred
    R_ELIG_LAB = 14
    _row(R_ELIG_LAB, "(B) Eligible labour",
         f"=B{R_LABOUR}-IFERROR(B{R_GOVTB},0)-IFERROR(B{R_DEFER},0)")

    # Row 15: Amount eligible for the tax credit = MIN(A cap, B labour)
    R_AMOUNT_ELIG = 15
    _row(R_AMOUNT_ELIG, "Amount eligible for the tax credit",
         f"=MIN(B{R_CAP_A},B{R_ELIG_LAB})")

    # Row 16: Total tax credit = eligible * 32%
    R_QC_TC = 16
    _row(R_QC_TC, "Total tax credit", f"=B{R_AMOUNT_ELIG}*0.32")

    # Row 17: Animation and special effects bonus (sub-label, bold)
    _cell(17, 1, "Animation and special effects bonus", font=_CAL_BOLD)
    ws.row_dimensions[17].height = ROW_H

    # Row 18: Animation and special effects labour
    R_ANIM_LAB = 18
    _row(R_ANIM_LAB, "Animation and special effects labour", None)

    # Row 19: Bonus
    R_BONUS = 19
    _row(R_BONUS, "Bonus", None)

    # Row 20: Total Quebec tax credit = TC + Bonus
    R_TOTAL_QC = 20
    _row(R_TOTAL_QC, "Total Quebec tax credit",
         f"=B{R_QC_TC}+IFERROR(B{R_BONUS},0)", bold=True)

    # Row 21: Percentage of budget = Total QC TC / Production costs
    R_PCT_QC = 21
    _row(R_PCT_QC, "Percentage of budget",
         f"=IFERROR(B{R_TOTAL_QC}/B{R_PROD},0)", pct=True)

    # ── Row 23: Federal header ────────────────────────────────────────────────
    _hdr(23, "QC PRODUCER- FEDERAL TAX CREDIT")

    # Row 25: Total Production cost (same reference)
    R_PROD_FED = 25
    _row(R_PROD_FED, "Total Production cost", f"=B{R_PROD}")

    # Row 26: QC Tax Credit (negative of Total Quebec TC)
    R_QC_TC_NEG = 26
    _row(R_QC_TC_NEG, "QC Tax Credit", f"=-B{R_TOTAL_QC}")

    # Row 27: 50% Meals & Perdiems — negative 50% of Breakout Budget AI2
    R_MEALS = 27
    _row(R_MEALS, "50% Meals & Perdiems", "=-'Breakout Budget'!$AI$2*0.5")

    # Row 28: aide gouv
    R_AIDE = 28
    _row(R_AIDE, "aide gouv.", None)

    # Row 29: Net Production cost — B26 and B27 are already negative so add them
    R_NET_PROD = 29
    _row(R_NET_PROD, "Net Production cost",
         f"=B{R_PROD_FED}+B{R_QC_TC_NEG}"
         f"+IFERROR(B{R_MEALS},0)-IFERROR(B{R_AIDE},0)")

    # Row 30: (A) Eligible production cost = net * 60%
    R_ELIG_A_FED = 30
    _row(R_ELIG_A_FED, "(A) Eligible production cost", f"=B{R_NET_PROD}*0.6")

    # Row 32: Labour expenditure — Federal Labour from Breakout Budget V2
    R_LAB_FED = 32
    _row(R_LAB_FED, "Labour expenditure", "='Breakout Budget'!$V$2")

    # Row 33: Deferrals
    R_DEFER_FED = 33
    _row(R_DEFER_FED, "Deferrals", None)

    # Row 34: Sub-total = labour - deferrals
    R_SUB = 34
    _row(R_SUB, "Sub-total",
         f"=B{R_LAB_FED}-IFERROR(B{R_DEFER_FED},0)")

    # Row 35: Percentage of ownership (default 100%)
    R_OWN = 35
    _row(R_OWN, "Percentage of ownership", 1.0, pct=True)

    # Row 36: (B) Net labour expenditure = sub-total * ownership %
    R_ELIG_B_FED = 36
    _row(R_ELIG_B_FED, "(B) Net labour expenditure",
         f"=B{R_SUB}*B{R_OWN}")

    # Row 38: Eligible cost for Fed. Tax Credit = MIN(A, B)
    R_ELIG_FED = 38
    _row(R_ELIG_FED, "Eligible cost for Fed. Tax Credit",
         f"=MIN(B{R_ELIG_A_FED},B{R_ELIG_B_FED})")

    # Row 40: Total Federal Tax Credit = eligible * 25%
    R_FED_TC = 40
    _row(R_FED_TC, "Total Federal Tax Credit", f"=B{R_ELIG_FED}*0.25", bold=True)

    # Row 41: Percentage of budget = Fed TC / Production cost
    R_PCT_FED = 41
    _row(R_PCT_FED, "Percentage of budget",
         f"=IFERROR(B{R_FED_TC}/B{R_PROD},0)", pct=True)

    # Row 42: TOTAL TAX CREDIT = Quebec + Federal
    R_TOTAL_TC = 42
    _row(R_TOTAL_TC, "TOTAL TAX CREDIT",
         f"=B{R_TOTAL_QC}+B{R_FED_TC}", bold=True, total=True,
         fmt='#,##0 "$"')
    # Match the double border on the label cell too
    ws.cell(row=R_TOTAL_TC, column=1).border = Border(top=_THIN, bottom=_THIN)

    # ── Outer box: A4:B42 ────────────────────────────────────────────────────
    R_BOX_TOP = 4
    R_BOX_BOT = R_TOTAL_TC  # 42
    for r in range(R_BOX_TOP, R_BOX_BOT + 1):
        for col in (1, 2):
            c = ws.cell(row=r, column=col)
            existing = c.border
            top    = _THIN if r == R_BOX_TOP else existing.top
            bottom = _THIN if r == R_BOX_BOT else existing.bottom
            left   = _THIN if col == 1 else existing.left
            right  = _THIN if col == 2 else existing.right
            c.border = Border(top=top, bottom=bottom, left=left, right=right)

    # Blank row 43
    ws.row_dimensions[43].height = ROW_H

    # Row 44: Total Production Cost (label row)
    _row(44, "Total Production Cost", f"=B{R_PROD}")

    # Row 45: Percentage of Total Tax Credits
    _row(45, "Percentage of Total Tax Credits",
         f"=IFERROR(B{R_TOTAL_TC}/B{R_PROD},0)", bold=True, pct=True,
         fmt=FMT_PCT1)


def _write_cto_sheet(ws) -> None:
    """Contribution To Overhead (CTO) analysis sheet."""
    ws.title = "CTO"

    _CAL_NORMAL = Font(name="Calibri", size=10)
    _CAL_BOLD   = Font(name="Calibri", bold=True, size=10)
    _CAL_TITLE  = Font(name="Calibri", bold=True, size=12)
    _CAL_SMALL  = Font(name="Calibri", size=9)

    ROW_H    = 16
    FMT_NUM  = _ACCOUNTING_FORMAT
    FMT_PCT  = '0.00%'
    FMT_PCT1 = '0.0%'

    _GRAY_FILL   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    _YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    _HDR_FILL    = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    _TITLE_FILL  = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    _GREEN_FILL  = PatternFill(start_color="A8FFC1", end_color="A8FFC1", fill_type="solid")
    _TEAL_FILL   = PatternFill(start_color="C1FFFD", end_color="C1FFFD", fill_type="solid")
    _TEAL2_FILL  = PatternFill(start_color="C1FFFE", end_color="C1FFFE", fill_type="solid")

    # Column widths
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 3    # blank spacer between P&L and fiscal table
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 16
    ws.column_dimensions["I"].width = 16   # Total column for fiscal table

    def _cell(row, col, value=None, font=None, align=None, fmt=None,
              fill=None, border=None):
        ws.row_dimensions[row].height = ROW_H
        c = ws.cell(row=row, column=col, value=value)
        c.font      = font   or _CAL_NORMAL
        c.alignment = align  or _LEFT
        c.border    = _NO_BORDER
        if fill   is not None: c.fill          = fill
        if fmt    is not None: c.number_format = fmt
        if border is not None: c.border        = border
        return c

    def _label(row, text, bold=False, indent=False):
        font = _CAL_BOLD if bold else _CAL_NORMAL
        col  = 1
        c = _cell(row, col, text, font=font)
        return c

    def _num(row, col, value, bold=False, pct=False):
        font = _CAL_BOLD if bold else _CAL_NORMAL
        fmt  = FMT_PCT if pct else FMT_NUM
        return _cell(row, col, value, font=font, align=_RIGHT, fmt=fmt)

    def _blank(row):
        ws.row_dimensions[row].height = ROW_H

    def _fill_range(row_start, row_end, col_start, col_end, fill):
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = fill

    def _border_bottom_range(row, col_start, col_end):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            existing = cell.border
            cell.border = Border(
                top=existing.top, left=existing.left, right=existing.right,
                bottom=_THIN,
            )

    # ── Row 1: Title ──────────────────────────────────────────────────────────
    c = ws.cell(row=1, column=1, value="Contribution To Overhead")
    c.font      = _CAL_TITLE
    c.alignment = _LEFT
    ws.row_dimensions[1].height = ROW_H

    # Row 2: Project name from Breakdown sheet
    _cell(2, 1, "=Breakdown!B2", font=_CAL_NORMAL)

    # ── Row 3: Column group headers ───────────────────────────────────────────
    for col, label in [
        (5, "Production"), (7, "Distribution"),
    ]:
        c = _cell(3, col, label, font=_CAL_BOLD, align=_CENTER, fill=_HDR_FILL,
                  border=Border(bottom=_THIN))
    # Merge Production E3:F3 and Distribution G3:H3
    ws.merge_cells("E3:F3")
    ws.merge_cells("G3:H3")

    # ── Row 4: Sub-column headers ─────────────────────────────────────────────
    for col, label in [
        (5, "At Green Light"), (6, "Projected"),
        (7, "Projected"), (8, "Total"), (9, "Comments"),
    ]:
        c = _cell(4, col, label, font=_CAL_BOLD, align=_CENTER, fill=_HDR_FILL,
                  border=Border(bottom=_THIN))

    # ── Rows 5–7: Revenue section ─────────────────────────────────────────────
    _label(5, "Presales")
    _num(5, 3, "=Sales!E8")
    _num(5, 5, "=C5")
    _num(5, 8, "=SUM(E5:G5)")

    _label(6, "Library")
    _num(6, 3, "=Sales!D12")
    _num(6, 6, "=C6")
    _num(6, 8, "=SUM(E6:G6)")

    _label(7, "Total", bold=True)
    for col, formula in [
        (3, "=SUM(C5:C6)"), (5, "=SUM(E5:E6)"), (6, "=SUM(F5:F6)"),
        (7, "=SUM(G5:G6)"), (8, "=SUM(E7:G7)"),
    ]:
        _num(7, col, formula, bold=True)

    _blank(8)

    # ── Rows 9–13: Cost section ───────────────────────────────────────────────
    _label(9, "Production Costs")
    _num(9, 3, "=Breakdown!C7")
    _num(9, 8, "=SUM(E9:G9)")

    _label(10, "Tax Credits")
    _num(10, 3, "=-Breakdown!C18")
    _num(10, 8, "=SUM(E10:G10)")

    _label(11, "Expected Overage / (Underage)")
    _num(11, 3, 0)
    _num(11, 8, "=SUM(E11:G11)")

    _label(12, "Internals")
    _num(12, 3, "=-Breakdown!C17")
    _num(12, 8, "=SUM(E12:G12)")

    _label(13, "Total Incremental Costs", bold=True)
    for col, formula in [
        (3, "=SUM(C9:C12)"), (5, "=-C13"), (8, "=SUM(E13:G13)"),
    ]:
        _num(13, col, formula, bold=True)

    _blank(14)

    # ── Rows 15–17: Distribution section ─────────────────────────────────────
    _label(15, "Distribution Fee")
    _cell(15, 2, "=Sales!C18", align=_RIGHT, fmt=FMT_PCT)
    _num(15, 6, "=-G15")
    _num(15, 7, "=B15*C6")

    _label(16, "Distribution Expenses")
    _cell(16, 2, "=Sales!C19", align=_RIGHT, fmt=FMT_PCT)
    _num(16, 3, "=B$16*C6")
    _num(16, 6, "=-C16")
    _num(16, 8, "=SUM(E16:G16)")

    _label(17, "Residual & Back End Estimate")
    _cell(17, 2, "=Sales!C20", align=_RIGHT, fmt=FMT_PCT)
    _num(17, 3, 0)
    _num(17, 6, "=-C17")
    _num(17, 8, "=SUM(E17:G17)")

    _blank(18)

    # ── Row 19: Contribution to Overheads ────────────────────────────────────
    _label(19, "Contribution to Overheads", bold=True)
    for col, formula in [
        (3, "=C7-C13-C15-C16-C17"), (5, "=SUM(E7:E18)"), (6, "=SUM(F7:F18)"),
        (7, "=SUM(G7:G18)"), (8, "=SUM(H7:H18)"),
    ]:
        _num(19, col, formula, bold=True)

    # Row 21: Hours
    _label(21, "Hours")
    _num(21, 3, "=Breakdown!C3")

    # Row 22: Contribution per hour
    _label(22, "Contribution per hour", bold=True)
    for col, formula in [
        (3, "=C19/$C$21"), (5, "=E19/$C$21"), (6, "=F19/$C$21"),
        (7, "=G19/$C$21"), (8, "=H19/$C$21"),
    ]:
        _num(22, col, formula, bold=True)

    # Row 23: Contribution as % of revenues
    for col, formula in [
        (3, "=C19/C7"), (5, "=E19/E7"), (6, "=F19/F7"),
        (7, "=G19/C6"), (8, "=H19/H7"),
    ]:
        _num(23, col, formula, pct=True)

    _blank(24)

    # Row 25: Internal Rate of Return
    _label(25, "Internal Rate of Return Generated")
    _num(25, 3, 0)

    _blank(26)

    # ── Rows 27–33: Financial reporting rates ────────────────────────────────
    _label(27, "Rates for financial reporting purposes", bold=True)

    _label(28, "Marginal rate of participation")
    _num(28, 3, "=C17/C7", pct=True)

    _label(29, "Amortisation rate")
    _num(29, 3, "=(C9+C10+C11+C12+(G29*C9))/C7", pct=True)
    _cell(29, 5, "assume's capitalization rate of 12%", font=_CAL_SMALL)
    _cell(29, 7, 0.12, align=_RIGHT, fmt="0%",
          fill=_YELLOW_FILL)

    _label(30, "distribution expense")
    _num(30, 3, "=C16/C7", pct=True)

    _label(31, "total cost of sales")
    _num(31, 3, "=SUM(C28:C30)", pct=True)

    _blank(32)

    _label(33, "Gross Margin", bold=True)
    _num(33, 3, "=1-C31", bold=True, pct=True)

    _blank(34)

    # ── Rows 35–50: Production P&L section ───────────────────────────────────
    ws.merge_cells("A34:C34")
    c = ws.cell(row=34, column=1, value="Production P&L")
    c.font      = _CAL_BOLD
    c.alignment = _LEFT
    ws.row_dimensions[34].height = ROW_H

    # ── Fiscal delivery table (E34:I39) ───────────────────────────────────────
    for col, label in [(5, "Fiscal"), (6, "2026"), (7, "2027"), (8, "2028"), (9, "Total")]:
        _cell(34, col, label, font=_CAL_BOLD, align=_CENTER)

    _cell(35, 5, "Deliveries", font=_CAL_BOLD)
    _cell(35, 6, 0,            align=_RIGHT, fmt="#,##0")
    _cell(35, 7, 21,           align=_RIGHT, fmt="#,##0")
    _cell(35, 8, 9,            align=_RIGHT, fmt="#,##0")
    _num(35, 9,  "=SUM(F35:H35)")

    _cell(36, 5, "=A38")
    for col, formula in [
        (6, "=$B38*(F35/$I35)"), (7, "=$B38*G35/$I35"),
        (8, "=$B38*H35/$I35"),   (9, "=SUM(F36:H36)"),
    ]:
        _num(36, col, formula)

    _cell(37, 5, "=A45")
    for col, formula in [
        (6, "=-F36*$C45"), (7, "=-G36*$C45"),
        (8, "=-H36*$C45"), (9, "=SUM(F37:H37)"),
    ]:
        _num(37, col, formula)

    _cell(38, 5, "=A47", font=_CAL_BOLD)
    for col, formula in [
        (6, "=F36+F37"), (7, "=G36+G37"),
        (8, "=H36+H37"), (9, "=SUM(F38:H38)"),
    ]:
        _num(38, col, formula, bold=True)

    for col, formula in [
        (6, "=F38/F36"), (7, "=G38/G36"),
        (8, "=H38/H36"), (9, "=I38/I36"),
    ]:
        _num(39, col, formula, pct=True)

    _label(35, "Revenues", bold=True)

    _label(36, "Presales")
    _num(36, 2, "=C5")
    _num(36, 3, "=B36/B41", pct=True)

    _label(37, "Distribution Advance")
    _num(37, 3, "=B37/B41", pct=True)

    _label(38, "Total Revenues", bold=True)
    _num(38, 2, "=SUM(B36:B37)", bold=True)
    _num(38, 3, "=C36+C37", bold=True, pct=True)

    _blank(39)

    _label(40, "Cost of sales", bold=True)

    _label(41, "Production cost")
    _num(41, 2, "=C9")
    _num(41, 3, 1, pct=True)

    _label(42, "tax credits")
    _num(42, 2, "=C10")
    _num(42, 3, "=-B42/B41", pct=True)

    _label(43, "internal production fees")
    _num(43, 2, "=C12")
    _num(43, 3, "=-B43/B$41", pct=True)

    _label(44, "capitalized overheads costs")
    _num(44, 2, "=G28*B41")
    _num(44, 3, "=-B44/B$41", pct=True)

    _label(45, "Total costs", bold=True)
    _num(45, 2, "=SUM(B41:B44)", bold=True)
    _num(45, 3, "=B45/B38", bold=True, pct=True)

    _blank(46)

    _label(47, "Gross Margin", bold=True)
    _num(47, 2, "=B38-B45", bold=True)
    _num(47, 3, "=C38-C45", bold=True, pct=True)

    _num(48, 2, "=B47/B38", pct=True)
    _num(48, 3, "=C47/C38", pct=True)

    _blank(49)

    _label(50, "Adjusted gross margin", bold=True)
    _num(50, 2, "=B47/B36", bold=True, pct=True)

    # ── Post-data styling ─────────────────────────────────────────────────────
    def _border_top_range(row, col_start, col_end):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            existing = cell.border
            cell.border = Border(
                top=_THIN, left=existing.left,
                right=existing.right, bottom=existing.bottom,
            )

    def _border_top_bottom_range(row, col_start, col_end):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=row, column=c)
            existing = cell.border
            cell.border = Border(
                top=_THIN, bottom=_THIN,
                left=existing.left, right=existing.right,
            )

    _DOUBLE = Side(style="double")
    _NO_FILL = PatternFill(fill_type=None)

    def _clear_fill_range(row_start, row_end, col_start, col_end):
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = _NO_FILL

    # A1:I1 – light grey title row
    _fill_range(1, 1, 1, 9, _TITLE_FILL)
    # Row 6 bottom border: A6:I6
    _border_bottom_range(6, 1, 9)
    # A12:I12 – bottom border
    _border_bottom_range(12, 1, 9)
    # A19:I19 – top and bottom border
    _border_top_bottom_range(19, 1, 9)
    # A19:C19 – green (Contribution to Overheads)
    _fill_range(19, 19, 1, 3, _GREEN_FILL)
    # A22:I22 – top border
    _border_top_range(22, 1, 9)
    # A22:C22 – green (Contribution per hour)
    _fill_range(22, 22, 1, 3, _GREEN_FILL)
    # A23:I23 – bottom border
    _border_bottom_range(23, 1, 9)
    # A25:C25 – green (Internal Rate of Return)
    _fill_range(25, 25, 1, 3, _GREEN_FILL)
    # A27:I27 – top and bottom border, grey fill (Rates section header)
    _border_top_bottom_range(27, 1, 9)
    _fill_range(27, 27, 1, 9, _TITLE_FILL)
    # E3:H3 and E4:M4 – no fill (remove header fill from data columns)
    _clear_fill_range(3, 3, 5, 8)
    _clear_fill_range(4, 4, 5, 13)
    # A34:C34 – bottom border (Production P&L header row)
    _border_bottom_range(34, 1, 3)
    # A34:C50 – teal (Production P&L section)
    _fill_range(34, 50, 1, 3, _TEAL_FILL)
    # E34:I39 – fiscal delivery table teal fill
    _fill_range(34, 39, 5, 9, _TEAL2_FILL)
    # B38 – bottom border
    _border_bottom_range(38, 2, 2)
    # B44 – bottom border
    _border_bottom_range(44, 2, 2)
    # B47 – top border and double bottom border
    cell_b47 = ws.cell(row=47, column=2)
    existing = cell_b47.border
    cell_b47.border = Border(
        top=_THIN, bottom=_DOUBLE,
        left=existing.left, right=existing.right,
    )


def _write_irr_sheet(ws):
    from openpyxl.utils import get_column_letter as gcl

    def _fill_range(row_start, row_end, col_start, col_end, fill):
        for r in range(row_start, row_end + 1):
            for c in range(col_start, col_end + 1):
                ws.cell(row=r, column=c).fill = fill

    # Quarter columns C..AP = cols 3..42
    Q_START, Q_END = 3, 42          # 40 quarters (10 years × 4)
    # Annual summary AR..BA = cols 44..53
    A_START, A_END = 44, 53         # 10 annual summary cols
    # Check columns
    BB_COL, BC_COL, BD_COL, BE_COL = 54, 55, 56, 57

    def _c(row, col, val=None, **kwargs):
        cell = ws.cell(row=row, column=col)
        if val is not None:
            if isinstance(val, str) and val.startswith("="):
                cell.value = val
            else:
                cell.value = val
        if "font" in kwargs:
            cell.font = kwargs["font"]
        if "align" in kwargs:
            cell.alignment = kwargs["align"]
        if "fmt" in kwargs:
            cell.number_format = kwargs["fmt"]
        if "fill" in kwargs:
            cell.fill = kwargs["fill"]
        if "border" in kwargs:
            cell.border = kwargs["border"]
        return cell

    _BOLD = Font(bold=True)
    _BOLD_ITALIC = Font(bold=True, italic=True)
    _GREY_FILL = PatternFill("solid", fgColor="F2F2F2")
    _TEAL_FILL = PatternFill("solid", fgColor="C1FFFD")
    _YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
    _HDR_FILL = PatternFill("solid", fgColor="D9E1F2")   # light blue for col headers
    _SUBHDR_FILL = PatternFill("solid", fgColor="BDD7EE") # slightly darker blue
    _GREEN_FILL = PatternFill("solid", fgColor="A8FFC1")
    _CENTER = Alignment(horizontal="center")
    _RIGHT = Alignment(horizontal="right")
    _LEFT = Alignment(horizontal="left")
    _NUM = "#,##0_);(#,##0)"          # accounting-style with parens for negatives
    _PCT = "0.00%"
    _PCT0 = "0%"
    _DATE_FMT = "MMM-YY"
    _T = Side(style="thin")
    _D = Side(style="double")

    def _border_bottom(row, c1, c2):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = Border(top=cell.border.top, left=cell.border.left,
                                 right=cell.border.right, bottom=_T)

    def _border_top(row, c1, c2):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = Border(top=_T, left=cell.border.left,
                                 right=cell.border.right, bottom=cell.border.bottom)

    def _border_top_double_bottom(row, c1, c2):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = Border(top=_T, left=cell.border.left,
                                 right=cell.border.right, bottom=_D)

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16
    # Q cols C..AP
    for col in range(Q_START, Q_END + 1):
        ws.column_dimensions[gcl(col)].width = 12
    # blank spacer AQ (col 43)
    ws.column_dimensions[gcl(43)].width = 3
    # Annual AR..BA
    for col in range(A_START, A_END + 1):
        ws.column_dimensions[gcl(col)].width = 14
    # Check cols
    for col in [BB_COL, BC_COL, BD_COL, BE_COL]:
        ws.column_dimensions[gcl(col)].width = 14

    # ── Row 1: Title ───────────────────────────────────────────────────────────
    _c(1, 1, "IRR Calculation", font=Font(bold=True, size=13))
    _fill_range(1, 1, 1, 57, _GREY_FILL)

    # ── Row 2: Basis of preparation ────────────────────────────────────────────
    _c(2, 1, "Basis of preparation")
    _c(2, 2, "exclude internals")
    _c(2, 3, "**fill in Q1")

    # ── Row 3: Date headers ────────────────────────────────────────────────────
    _c(3, 3, 46357, fmt=_DATE_FMT)
    offsets = (
        [90] +            # D3 = C3+90
        [93] * 18 +       # E3..V3 = prev+93
        [88] * 8 +        # W3..AD3 = prev+88
        [90, 90, 91] +    # AE3, AF3, AG3
        [93] * 9          # AH3..AP3
    )
    q_cols = list(range(Q_START + 1, Q_END + 1))  # D..AP = cols 4..42
    prev_col = gcl(Q_START)
    for i, (col, off) in enumerate(zip(q_cols, offsets)):
        ltr = gcl(col)
        _c(3, col, f"={prev_col}3+{off}", fmt=_DATE_FMT)
        prev_col = ltr

    # ── Row 4: Year labels (quarterly cols + annual cols) ──────────────────────
    # Quarterly: Year 1 spans C4:F4, Year 2 G4:J4, ...
    for y in range(10):
        start_col = Q_START + y * 4
        _c(4, start_col, f"Year {y+1}", font=_BOLD, align=_CENTER)
        ws.merge_cells(f"{gcl(start_col)}4:{gcl(start_col+3)}4")
    # Annual section labels
    for y in range(10):
        _c(4, A_START + y, f"Year {y+1}", font=_BOLD, align=_CENTER)
    _c(4, BB_COL, "")
    _c(4, BC_COL, "Check", font=_BOLD, align=_CENTER)
    # Total label in B4
    _c(4, 2, "Total", font=_BOLD, align=_CENTER)

    # ── Row 5: Quarter labels ──────────────────────────────────────────────────
    for y in range(10):
        for q, label in enumerate(["Q1", "Q2", "Q3", "Q4"]):
            _c(5, Q_START + y * 4 + q, label, align=_CENTER)

    # ── Row 6: "Collections from" section header ───────────────────────────────
    _c(6, 1, "Collections from", font=_BOLD)

    # ── Rows 7-9: Broadcaster rows ─────────────────────────────────────────────
    for row, b_row in [(7, 4), (8, 5), (9, 6)]:
        _c(row, 1, f"=Sales!B{b_row}")
        _c(row, 2, f"=Sales!E{b_row}", fmt=_NUM)
        for col in range(Q_START, Q_END + 1):
            ltr = gcl(col)
            _c(row, col, f"=$B{row}*{ltr}$73", fmt=_NUM)
        # Annual
        for i, a_col in enumerate(range(A_START, A_END + 1)):
            q1 = gcl(Q_START + i * 4)
            q4 = gcl(Q_START + i * 4 + 3)
            _c(row, a_col, f"=SUM({q1}{row}:{q4}{row})", fmt=_NUM)
        _c(row, BC_COL, f"=B{row}-SUM({gcl(A_START)}{row}:{gcl(A_END)}{row})", fmt=_NUM)
        _c(row, BD_COL, f"=SUM({gcl(A_START)}{row}:{gcl(A_END)}{row})", fmt=_NUM)

    # ── Row 12: Tax Credit Estimate ────────────────────────────────────────────
    _c(12, 1, "Tax Credit Estimate", font=_BOLD)
    _c(12, 2, "=Breakdown!C18", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(12, col, f"=$B12*{ltr}$77", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(12, a_col, f"=SUM({q1}12:{q4}12)", fmt=_NUM)
    _c(12, BC_COL, f"=B12-SUM({gcl(A_START)}12:{gcl(A_END)}12)", fmt=_NUM)
    _c(12, BD_COL, f"=SUM({gcl(A_START)}12:{gcl(A_END)}12)", fmt=_NUM)

    # ── Row 16: International sales ────────────────────────────────────────────
    _c(16, 1, "International sales", font=_BOLD)
    _c(16, 2, "=Sales!D12", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(16, col,
           f"=IF($B$51=$A$83,$B16*{ltr}83,IF($B$51=$A$82,$B16*{ltr}82,IF($B$51=$A$81,$B16*{ltr}81,\"\")))",
           fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(16, a_col, f"=SUM({q1}16:{q4}16)", fmt=_NUM)
    _c(16, BC_COL, f"=B16-SUM({gcl(A_START)}16:{gcl(A_END)}16)", fmt=_NUM)
    _c(16, BD_COL, f"=SUM({gcl(A_START)}16:{gcl(A_END)}16)", fmt=_NUM)
    _c(16, BE_COL, "=+BD16/B16", fmt=_PCT)

    # ── Row 18: Sub-total inflows ──────────────────────────────────────────────
    _c(18, 1, "Sub-total inflows", font=_BOLD)
    _c(18, 2, "=SUM(B7:B16)", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(18, col, f"=SUM({ltr}7:{ltr}16)", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        start = gcl(A_START - 1)  # AQ col (blank) — use range from data rows
        q1 = gcl(Q_START + i * 4)  # not used here; use AQ..BA summing rows 6-17
        # The user formula: =SUM(AR6:AR17) etc — sum annual col rows 6-17
        a_ltr = gcl(a_col)
        _c(18, a_col, f"=SUM({a_ltr}6:{a_ltr}17)", fmt=_NUM)
    _c(18, BC_COL, f"=B18-SUM({gcl(A_START)}18:{gcl(A_END)}18)", fmt=_NUM)
    _c(18, BD_COL, f"=SUM({gcl(A_START)}18:{gcl(A_END)}18)", fmt=_NUM)

    # ── Row 20: Outflows header ────────────────────────────────────────────────
    _c(20, 1, "Outflows", font=_BOLD)

    # ── Row 21: Production costs / Advance ────────────────────────────────────
    _c(21, 1, "Production costs / Advance", font=_BOLD)
    _c(21, 2, "=-CTO!C9", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(21, col, f"=$B21*{ltr}$69", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(21, a_col, f"=SUM({q1}21:{q4}21)", fmt=_NUM)
    _c(21, BC_COL, f"=B21-SUM({gcl(A_START)}21:{gcl(A_END)}21)", fmt=_NUM)
    _c(21, BD_COL, f"=SUM({gcl(A_START)}21:{gcl(A_END)}21)", fmt=_NUM)

    # ── Row 22: Distribution expenses ─────────────────────────────────────────
    _c(22, 1, "Distribution expenses", font=_BOLD)
    _c(22, 2, "=-CTO!C16", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(22, col, f"=$B$22*{ltr}16/$B$16", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(22, a_col, f"=SUM({q1}22:{q4}22)", fmt=_NUM)
    _c(22, BC_COL, f"=B22-SUM({gcl(A_START)}22:{gcl(A_END)}22)", fmt=_NUM)
    _c(22, BD_COL, f"=SUM({gcl(A_START)}22:{gcl(A_END)}22)", fmt=_NUM)

    # ── Row 23: Back-end share ─────────────────────────────────────────────────
    _c(23, 1, "Back-end share", font=_BOLD)
    _c(23, 2, "=-CTO!C17", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(23, col, f"={ltr}46", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(23, a_col, f"=SUM({q1}23:{q4}23)", fmt=_NUM)
    _c(23, BC_COL, f"=B23-SUM({gcl(A_START)}23:{gcl(A_END)}23)", fmt=_NUM)
    _c(23, BD_COL, f"=SUM({gcl(A_START)}23:{gcl(A_END)}23)", fmt=_NUM)

    # ── Row 25: Sub-total outflows ─────────────────────────────────────────────
    _c(25, 1, "Sub-total out-flows", font=_BOLD)
    _c(25, 2, "=SUM(B21:B23)", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(25, col, f"=SUM({ltr}21:{ltr}23)", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        al = gcl(a_col)
        _c(25, a_col, f"=SUM({al}21:{al}23)", fmt=_NUM)
    _c(25, BC_COL, f"=B25-SUM({gcl(A_START)}25:{gcl(A_END)}25)", fmt=_NUM)
    _c(25, BD_COL, f"=SUM({gcl(A_START)}25:{gcl(A_END)}25)", fmt=_NUM)

    # ── Row 27: Net cash movement ──────────────────────────────────────────────
    _c(27, 1, "Net cash movement", font=_BOLD)
    _c(27, 2, "=B18+B25", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(27, col, f"={ltr}18+{ltr}25", fmt=_NUM)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        al = gcl(a_col)
        _c(27, a_col, f"={al}18+{al}25", fmt=_NUM)
    _c(27, BC_COL, f"=B27-SUM({gcl(A_START)}27:{gcl(A_END)}27)", fmt=_NUM)
    _c(27, BD_COL, f"=SUM({gcl(A_START)}27:{gcl(A_END)}27)", fmt=_NUM)

    # ── Row 29: Internals ─────────────────────────────────────────────────────
    _c(29, 1, "Internals")
    _c(29, 2, "=Breakdown!C17", fmt=_NUM)

    # ── Row 30: Commission ────────────────────────────────────────────────────
    _c(30, 1, "Commission")
    _c(30, 2, "=CTO!G15", fmt=_NUM)
    # Quarterly values — all zeros except specific quarters
    q_vals_30 = {
        11: 3913.37930621445,
        15: 3135.13759818328,
        19: 852.559327905654,
        23: 820.682097630452,
        27: 1578.58970655969,
        31: 304.2152709151,
        35: 1124.81070648314,
        39: 2061.03039631042,
    }
    for col in range(Q_START, Q_END + 1):
        _c(30, col, q_vals_30.get(col, 0), fmt=_NUM)
    # Annual summary
    a_vals_30 = {
        44: 0, 45: 0,
        46: 3913.37930621445,
        47: 3135.13759818328,
        48: 852.559327905654,
        49: 820.682097630452,
        50: 1578.58970655969,
        51: 304.2152709151,
        52: 1124.81070648314,
        53: 2061.03039631042,
    }
    for a_col in range(A_START, A_END + 1):
        _c(30, a_col, a_vals_30.get(a_col, 0), fmt=_NUM)

    # ── Row 31: Back end ──────────────────────────────────────────────────────
    _c(31, 1, "Back end")
    _c(31, 2, "=-CTO!C17", fmt=_NUM)

    # ── Row 32: Sum ───────────────────────────────────────────────────────────
    _c(32, 2, "=SUM(B29:B31)", fmt=_NUM)

    # ── Row 33: IRR calc ──────────────────────────────────────────────────────
    _c(33, 1, "IRR calc", font=_BOLD)
    _c(33, 2, f"=IRR({gcl(A_START)}27:{gcl(A_END)}27)", fmt="0.00%")

    # ── Row 34 ────────────────────────────────────────────────────────────────
    _c(34, 2, "=B30/B18", fmt=_PCT)

    # ── Row 37: Back-end calcs header ─────────────────────────────────────────
    _c(37, 1, "Back-end calcs", font=_BOLD)

    # ── Row 38: Cash collections ───────────────────────────────────────────────
    _c(38, 1, "Cash collections")
    # C38 = C18, D38 = C38+D18, ...
    _c(38, Q_START, f"={gcl(Q_START)}18", fmt=_NUM)
    for col in range(Q_START + 1, Q_END + 1):
        prev = gcl(col - 1)
        cur = gcl(col)
        _c(38, col, f"={prev}38+{cur}18", fmt=_NUM)
    # Annual: AR38 = AQ38+AR14, but AQ is blank spacer — user formula: =AQ38+AR14
    # AQ col = 43; user has =AQ38+AR14 for AR38, etc.
    # Since AQ is blank (0), AR38=AR14, AS38=AR38+AS14, etc.
    _c(38, A_START, f"={gcl(43)}38+{gcl(A_START)}14", fmt=_NUM)
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        prev_a = gcl(a_col - 1)
        cur_a = gcl(a_col)
        _c(38, a_col, f"={prev_a}38+{cur_a}14", fmt=_NUM)

    # ── Row 39: Production costs / advances ────────────────────────────────────
    _c(39, 1, "Production costs / advances")
    # C39=C21
    _c(39, Q_START, f"={gcl(Q_START)}21", fmt=_NUM)
    # D39=(D21*(SUM($AR$21:$AZ$21)/$B$21))+C39 (cols 4..10 use AZ range cap)
    for col in range(Q_START + 1, Q_END + 1):
        cur = gcl(col)
        prev = gcl(col - 1)
        if col <= 10:
            _c(39, col, f"=({cur}21*(SUM($AR$21:$AZ$21)/$B$21))+{prev}39", fmt=_NUM)
        else:
            _c(39, col, f"=({cur}21*(SUM($AR$21:$BA$21)/$B$21))+{prev}39", fmt=_NUM)
    # Annual: AR39=AR17, AS39=AR39+AS17, ...
    _c(39, A_START, f"={gcl(A_START)}17", fmt=_NUM)
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        prev_a = gcl(a_col - 1)
        cur_a = gcl(a_col)
        _c(39, a_col, f"={prev_a}39+{cur_a}17", fmt=_NUM)

    # ── Row 40: Allow for recoupment ───────────────────────────────────────────
    _c(40, 1, "Allow for recoupment of internals to calculate back-end")
    _c(40, 2, "=-B29", fmt=_NUM)
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(40, col, f"=$B40*{ltr}69", fmt=_NUM)
    # Annual: AR40=SUM(C40:F40), AS40=SUM(G40:J40)+AR40, etc.
    _c(40, A_START, f"=SUM({gcl(Q_START)}40:{gcl(Q_START+3)}40)", fmt=_NUM)
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        prev_a = gcl(a_col - 1)
        q1 = gcl(Q_START + (i + 1) * 4 - 3)  # start of next year's quarterly block
        q4 = gcl(Q_START + (i + 1) * 4)
        _c(40, a_col, f"=SUM({q1}40:{q4}40)+{prev_a}40", fmt=_NUM)

    # ── Row 41: Distribution expenses (running) ────────────────────────────────
    _c(41, 1, "Distribution expenses")
    _c(41, Q_START, f"={gcl(Q_START)}22", fmt=_NUM)
    for col in range(Q_START + 1, Q_END + 1):
        prev = gcl(col - 1)
        cur = gcl(col)
        _c(41, col, f"={prev}41+{cur}22", fmt=_NUM)
    # Annual: AQ41+AR18, AR41+AS18, ...
    _c(41, A_START, f"={gcl(43)}41+{gcl(A_START)}18", fmt=_NUM)
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        prev_a = gcl(a_col - 1)
        cur_a = gcl(a_col)
        _c(41, a_col, f"={prev_a}41+{cur_a}18", fmt=_NUM)

    # ── Row 42: Recoupment of advance ─────────────────────────────────────────
    _c(42, 1, "Recoupment of advance")
    _c(42, Q_START, f"=-B61", fmt=_NUM)
    for col in range(Q_START + 1, Q_END + 1):
        ltr = gcl(col)
        _c(42, col, f"={ltr}42", fmt=_NUM)
    # Annual: C42, AR42, AS42, ...
    _c(42, A_START, f"={gcl(Q_START)}42", fmt=_NUM)
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        prev_a = gcl(a_col - 1)
        _c(42, a_col, f"={prev_a}42", fmt=_NUM)

    # ── Row 43: Net cumulative cash ex back-end ────────────────────────────────
    _c(43, 1, "Net cumulative cash ex back-end")
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(43, col, f"=SUM({ltr}38:{ltr}42)", fmt=_NUM)
    for a_col in range(A_START, A_END + 1):
        al = gcl(a_col)
        _c(43, a_col, f"=SUM({al}38:{al}42)", fmt=_NUM)

    # ── Row 45: Back-end ──────────────────────────────────────────────────────
    _c(45, 1, "Back-end")
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(45, col, f"=IF({ltr}43>0,-Sales!$C$20*{ltr}43,0)", fmt=_NUM)
    for a_col in range(A_START, A_END + 1):
        al = gcl(a_col)
        _c(45, a_col, f"=IF({al}43>0,-Sales!$C$20*{al}43,0)", fmt=_NUM)

    # ── Row 46: Movement on back-end ──────────────────────────────────────────
    _c(46, 1, "Movement on back-end")
    # C46 blank (no formula in original), D46=D45-C45, ...
    for col in range(Q_START + 1, Q_END + 1):
        cur = gcl(col)
        prev = gcl(col - 1)
        _c(46, col, f"={cur}45-{prev}45", fmt=_NUM)
    # Annual: AS46=AS45-AR45, etc.
    for i, a_col in enumerate(range(A_START + 1, A_END + 1)):
        cur_a = gcl(a_col)
        prev_a = gcl(a_col - 1)
        _c(46, a_col, f"={cur_a}45-{prev_a}45", fmt=_NUM)

    # ── Row 49: Workings header ────────────────────────────────────────────────
    _c(49, 1, "Workings for IRR / Cashflow", font=_BOLD)

    # ── Row 50: Sales forecast per hour ───────────────────────────────────────
    _c(50, 1, "Sales forecast per hour")
    _c(50, 2, "=Sales!D13", fmt=_NUM)

    # ── Row 51: International sales category ──────────────────────────────────
    _c(51, 1, "International sales category")
    _c(51, 2, '=IF(AND(B50>=C55,B50<=D55),B55,IF(AND(B50>=C56,B50<=D56),B56,IF(AND(B50>=C57,B50<=D57),B57,"")))')

    # ── Rows 54-57: Genre table ────────────────────────────────────────────────
    for col, label in [(1, "Genre"), (2, "Category"), (3, "Range from"), (4, "Range to")]:
        _c(54, col, label, font=_BOLD)
    genre_rows = [
        ("Fact Ent", "High value", 110000, 4000000),
        ("Fact Ent", "Medium value", 67500, 110000),
        ("Fact Ent", "Low value", 10000, 67500),
    ]
    for i, (genre, cat, lo, hi) in enumerate(genre_rows):
        row = 55 + i
        _c(row, 1, genre)
        _c(row, 2, cat)
        _c(row, 3, lo, fmt=_NUM)
        _c(row, 4, hi, fmt=_NUM)

    # ── Row 59: Back-end on ultimates header ──────────────────────────────────
    _c(59, 1, "Back-end on ultimates", font=_BOLD)

    # ── Rows 60-65 ────────────────────────────────────────────────────────────
    _c(60, 1, "International sales")
    _c(60, 2, "=Sales!D12", fmt=_NUM)
    _c(61, 1, "Advance")
    _c(61, 2, "=FS!D21", fmt=_NUM)
    _c(62, 1, "Commish")
    _c(62, 2, "=CTO!G15", fmt=_NUM)
    _c(63, 1, "Dist expenses")
    _c(63, 2, "=CTO!C16", fmt=_NUM)
    _c(64, 1, "Net profit")
    _c(64, 2, "=IF((B60-B61-B62-B63)>1,B60-B61-B62-B63,0)", fmt=_NUM)
    _c(65, 1, "Share of net profit")
    _c(65, 2, "=B64*Sales!C20", fmt=_NUM)

    # ── Row 67: OUTFLOW - HARD COSTS header ───────────────────────────────────
    _c(67, 1, "OUTFLOW - HARD COSTS", font=_BOLD)
    # Year header references (every 4th quarterly col start)
    for y in range(10):
        start_q = Q_START + y * 4
        _c(67, start_q, f"={gcl(start_q)}4")
    # Annual section headers
    for y in range(10):
        _c(67, A_START + y, f"Year {y+1}", font=_BOLD, align=_CENTER)

    # ── Row 68: Quarter references ─────────────────────────────────────────────
    for col in range(Q_START, Q_END + 1):
        ltr = gcl(col)
        _c(68, col, f"={ltr}5")

    # ── Row 69: Outflow schedule values ───────────────────────────────────────
    outflow_vals = [
        0.0705957707638597,
        0.369415571024316,
        0.369124228289782,
        0.153239672680844,
        0.0322707824524556,
        0.00539494900702482,
        -0.0000409742182823332,
    ]
    for i, v in enumerate(outflow_vals):
        _c(69, Q_START + i, v, fmt=_PCT)
    # Annual sums
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(69, a_col, f"=SUM({q1}69:{q4}69)", fmt=_PCT)

    # ── Row 71: INFLOW - PRESALE header ───────────────────────────────────────
    _c(71, 1, "INFLOW - PRESALE", font=_BOLD)
    for y in range(4):
        _c(71, Q_START + y * 4, f"Year {y+1}", align=_CENTER)

    # ── Row 72: Quarter headers ────────────────────────────────────────────────
    for col in range(Q_START, Q_START + 16):
        ltr = gcl(col)
        _c(72, col, f"={ltr}5")

    # ── Row 73: Presale schedule ───────────────────────────────────────────────
    presale_vals = {
        3: 0, 4: 0.2666, 5: 0.1769, 6: 0.1576,
        7: 0.2394, 8: 0.0484, 9: 0.0611, 10: 0,
        11: 0.05, 12: 0,
    }
    for col in range(Q_START, Q_END + 1):
        v = presale_vals.get(col, 0)
        _c(73, col, v if v != 0 else 0, fmt=_PCT0)
    # Annual sums (only first 4 years shown in user data, but add all)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(73, a_col, f"=SUM({q1}73:{q4}73)", fmt=_PCT0)

    # ── Row 75: INFLOW - TAX CREDITS header ───────────────────────────────────
    _c(75, 1, "INFLOW - TAX CREDITS", font=_BOLD)
    for y in range(4):
        _c(75, Q_START + y * 4, f"Year {y+1}", align=_CENTER)

    # ── Row 76: Quarter headers ────────────────────────────────────────────────
    for col in range(Q_START, Q_START + 16):
        ltr = gcl(col)
        _c(76, col, f"={ltr}5")

    # ── Row 77: Tax credit schedule ────────────────────────────────────────────
    tc_vals = {7: 0.641583495172534, 11: 0.358416504827466}
    for col in range(Q_START, Q_END + 1):
        v = tc_vals.get(col, 0)
        _c(77, col, v if v != 0 else 0, fmt=_PCT)
    for i, a_col in enumerate(range(A_START, A_END + 1)):
        q1 = gcl(Q_START + i * 4)
        q4 = gcl(Q_START + i * 4 + 3)
        _c(77, a_col, f"=SUM({q1}77:{q4}77)", fmt=_PCT)

    # ── Row 79: INFLOW - INT'L SALES header ───────────────────────────────────
    _c(79, 1, "INFLOW - INT'L SALES", font=_BOLD)
    for y in range(4):
        _c(79, Q_START + y * 4, f"Year {y+1}", align=_CENTER)

    # ── Row 80: Quarter headers ────────────────────────────────────────────────
    for col in range(Q_START, Q_START + 16):
        ltr = gcl(col)
        _c(80, col, f"={ltr}5")

    # ── Rows 81-83: Int'l sales schedules ─────────────────────────────────────
    intl_data = {
        81: {
            "label": "High value",
            "q": {11: 0.408598204250028, 15: 0.14315770201277, 19: 0.100023357342737,
                  23: 0.0717846050033827, 27: 0.10559609365345, 31: 0.0662132130324036,
                  35: 0.0236877142264881, 39: 0.028836087294499},
            "a": {44: 0.408598204250028, 45: 0.14315770201277, 46: 0.100023357342737,
                  47: 0.0717846050033827, 48: 0.10559609365345, 49: 0.0662132130324036,
                  50: 0.0236877142264881, 51: 0.028836087294499, 52: 0, 53: 0},
        },
        82: {
            "label": "Medium value",
            "q": {11: 0.265054928604243, 15: 0.167242762035813, 19: 0.0852937183926511,
                  23: 0.050337429480966, 27: 0.109834524624246, 31: 0.126156938733889,
                  35: 0.0743179057268728, 39: 0.0517182281808347},
            "a": {44: 0.265054928604243, 45: 0.167242762035813, 46: 0.0852937183926511,
                  47: 0.050337429480966, 48: 0.109834524624246, 49: 0.126156938733889,
                  50: 0.0743179057268728, 51: 0.0517182281808347, 52: 0, 53: 0},
        },
        83: {
            "label": "Low value",
            "q": {11: 0.26089195374763, 15: 0.209009173212219, 19: 0.0568372885270436,
                  23: 0.0547121398420301, 27: 0.105239313770646, 31: 0.0202810180610066,
                  35: 0.0749873804322091, 39: 0.137402026420694},
            "a": {44: 0.26089195374763, 45: 0.209009173212219, 46: 0.0568372885270436,
                  47: 0.0547121398420301, 48: 0.105239313770646, 49: 0.0202810180610066,
                  50: 0.0749873804322091, 51: 0.137402026420694, 52: 0, 53: 0},
        },
    }
    for row, data in intl_data.items():
        _c(row, 1, data["label"])
        for col in range(Q_START, Q_END + 1):
            v = data["q"].get(col, 0)
            _c(row, col, v, fmt=_PCT)
        for a_col in range(A_START, A_END + 1):
            v = data["a"].get(a_col, 0)
            _c(row, a_col, v, fmt=_PCT)

    # ── Post-data styling ──────────────────────────────────────────────────────

    # Row 1: title grey (already filled), make font bold+white stand out
    for c in range(1, BE_COL + 1):
        ws.cell(row=1, column=c).fill = _GREY_FILL
    ws.cell(row=1, column=1).font = Font(bold=True, size=12)

    # C2: yellow "fill in Q1" reminder
    ws.cell(row=2, column=3).fill = _YELLOW_FILL

    # Rows 3-5: light blue column-header band across quarterly + annual cols
    _fill_range(3, 3, Q_START, Q_END, _HDR_FILL)
    _fill_range(3, 3, A_START, A_END, _HDR_FILL)
    _fill_range(4, 5, Q_START, Q_END, _HDR_FILL)
    _fill_range(4, 5, A_START, A_END, _HDR_FILL)
    # Bottom border on row 5 (separates headers from data)
    _border_bottom(5, 1, BE_COL)

    # Row 6 "Collections from" — bottom border
    _border_bottom(6, 1, BD_COL)

    # Row 9 — bottom border (last broadcaster)
    _border_bottom(9, 1, BD_COL)

    # Row 12 Tax Credit — top + bottom border
    _border_top(12, 1, BD_COL)
    _border_bottom(12, 1, BD_COL)

    # Row 16 International sales — bottom border
    _border_bottom(16, 1, BE_COL)

    # Row 18 Sub-total inflows — grey fill + top+bottom border
    _fill_range(18, 18, 1, BD_COL, _GREY_FILL)
    _border_top(18, 1, BD_COL)
    _border_bottom(18, 1, BD_COL)

    # Row 20 Outflows header — bottom border
    _border_bottom(20, 1, BD_COL)

    # Row 25 Sub-total outflows — grey fill + top+bottom border
    _fill_range(25, 25, 1, BD_COL, _GREY_FILL)
    _border_top(25, 1, BD_COL)
    _border_bottom(25, 1, BD_COL)

    # Row 27 Net cash movement — grey fill + top + double bottom border
    _fill_range(27, 27, 1, BD_COL, _GREY_FILL)
    _border_top_double_bottom(27, 1, BD_COL)

    # Rows 29-32 internals block — light indent, bottom border on row 32
    _border_bottom(32, 1, 2)

    # Row 33 IRR calc — yellow highlight
    ws.cell(row=33, column=1).font = _BOLD
    ws.cell(row=33, column=2).fill = _YELLOW_FILL
    ws.cell(row=33, column=2).font = Font(bold=True, size=11)
    _border_top_double_bottom(33, 1, 2)

    # Row 37 Back-end calcs section header — grey fill + bottom border
    _fill_range(37, 37, 1, BD_COL, _GREY_FILL)
    ws.cell(row=37, column=1).font = _BOLD
    _border_bottom(37, 1, BD_COL)

    # Row 43 Net cumulative cash — bottom border
    _border_bottom(43, 1, BD_COL)

    # Row 45 Back-end — bottom border
    _border_bottom(45, 1, BD_COL)

    # Row 46 Movement on back-end — bottom border
    _border_bottom(46, 1, BD_COL)

    # Row 49 Workings section header — grey fill + bottom border
    _fill_range(49, 49, 1, BD_COL, _GREY_FILL)
    ws.cell(row=49, column=1).font = _BOLD
    _border_bottom(49, 1, BD_COL)

    # Genre table rows 54-57 — borders + header fill
    _fill_range(54, 54, 1, 4, _HDR_FILL)
    for r in range(54, 58):
        for c in range(1, 5):
            ws.cell(row=r, column=c).border = Border(
                top=_T, bottom=_T, left=_T, right=_T)

    # Row 59 Back-end on ultimates — grey fill + bottom border
    _fill_range(59, 59, 1, 4, _GREY_FILL)
    ws.cell(row=59, column=1).font = _BOLD
    _border_bottom(59, 1, 4)
    _border_bottom(65, 1, 2)

    # OUTFLOW/INFLOW section headers (rows 67, 71, 75, 79) — blue header fill
    for hdr_row in [67, 71, 75, 79]:
        _fill_range(hdr_row, hdr_row, 1, BD_COL, _SUBHDR_FILL)
        ws.cell(row=hdr_row, column=1).font = _BOLD
        _border_bottom(hdr_row, 1, BD_COL)

    # Q-label rows under each schedule (68, 72, 76, 80) — light header fill
    for q_row in [68, 72, 76, 80]:
        _fill_range(q_row, q_row, Q_START, Q_END, _HDR_FILL)
        _border_bottom(q_row, 1, BD_COL)

    # Int'l sales rows 81-83 — teal fill on non-zero % cells
    for intl_row in [81, 82, 83]:
        _border_bottom(intl_row, 1, BD_COL)

    # ── Freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "C6"

    # ── Tab color ─────────────────────────────────────────────────────────────
    ws.sheet_properties.tabColor = "A8FFC1"


def write_tax_credit_excel(
    budget: ParsedBudget,
    title: str,
    overrides: dict | None = None,
    global_bible: dict | None = None,
    num_episodes: int | None = None,
    duration_minutes: int | None = None,
) -> BytesIO:
    """Build a tax credit filing workbook and return as BytesIO.

    ``overrides``     — maps account_code → BreakoutOverride (project-specific).
    ``global_bible``  — maps account_code → (non_prov, pl, fl, psl, sp, fsl) tuple,
                        superseding BREAKOUT_BIBLE defaults before project overrides
                        are applied.
    """
    # Build effective bible: hardcoded defaults → global customisations
    effective_bible = dict(BREAKOUT_BIBLE)
    if global_bible:
        effective_bible.update(global_bible)

    wb = Workbook()

    # Remove the default empty sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    ws_topsheet = wb.create_sheet("Topsheet")
    _write_topsheet(ws_topsheet, budget, title)

    ws_lines = wb.create_sheet("Budget Lines")
    _write_budget_lines(ws_lines, budget)

    ws_detail = wb.create_sheet("Detail Budget")
    _write_detail_budget(ws_detail, budget)

    ws_breakout = wb.create_sheet("Breakout Budget")
    _write_breakout_budget(ws_breakout, budget, overrides or {}, effective_bible)
    ws_breakout.sheet_properties.tabColor = "B4FFF8"

    ws_breakdown = wb.create_sheet("Breakdown")
    _write_breakdown_sheet(ws_breakdown, title, num_episodes=num_episodes)
    ws_breakdown.sheet_properties.tabColor = "B4FFF8"

    ws_form6 = wb.create_sheet("Form6")
    _write_form6_sheet(ws_form6)
    ws_form6.sheet_properties.tabColor = "FFD4D2"

    ws_ofttc = wb.create_sheet("Ontario - OFTTC")
    _write_ofttc_sheet(ws_ofttc, title)
    ws_ofttc.sheet_properties.tabColor = "FFFEC8"

    ws_opstc = wb.create_sheet("Ontario - OPSTC")
    _write_opstc_sheet(ws_opstc, title)
    ws_opstc.sheet_properties.tabColor = "FFFEC8"

    ws_sodec = wb.create_sheet("Sodec")
    _write_sodec_sheet(ws_sodec)
    ws_sodec.sheet_properties.tabColor = "FFFEC8"

    ws_fs = wb.create_sheet("FS")
    _write_fs_sheet(ws_fs, title, num_episodes=num_episodes, duration_minutes=duration_minutes)
    ws_fs.sheet_properties.tabColor = "A8FFC1"

    ws_sales = wb.create_sheet("Sales")
    _write_sales_sheet(ws_sales)
    ws_sales.sheet_properties.tabColor = "A8FFC1"

    ws_cto = wb.create_sheet("CTO")
    _write_cto_sheet(ws_cto)
    ws_cto.sheet_properties.tabColor = "A8FFC1"

    ws_irr = wb.create_sheet("IRR")
    _write_irr_sheet(ws_irr)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
