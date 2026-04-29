"""Generate formatted Excel cashflow workbook from CashflowOutput."""

from collections import Counter
from datetime import date, timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter

from app.models.budget import ParsedBudget
from app.models.cashflow import CashflowOutput
from app.models.production import ProductionParameters

# Style constants
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=14)
SUBTITLE_FONT = Font(bold=True, size=10, color="666666")
CURRENCY_FORMAT = '#,##0'
CURRENCY_FORMAT_TOTAL = '#,##0'
ACCOUNTING_NO_DECIMAL_FORMAT = '_-* #,##0_-;-* #,##0_-;_-* "-"??_-;_-@_-'
BASE_ALIGNMENT = Alignment(horizontal="left", vertical="center")
SUBTLE_TOTAL_FILL = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
NO_BORDER = Border()

# Phase colors
PHASE_COLORS = {
    "PREP": PatternFill(start_color="B3D9FF", end_color="B3D9FF", fill_type="solid"),       # Light blue
    "SHOOT": PatternFill(start_color="B3FFB3", end_color="B3FFB3", fill_type="solid"),      # Light green
    "WRAP": PatternFill(start_color="FFFFB3", end_color="FFFFB3", fill_type="solid"),       # Light yellow
    "POST": PatternFill(start_color="FFD9B3", end_color="FFD9B3", fill_type="solid"),       # Light orange
    "HIATUS": PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid"),     # Light gray
    "DELIVERY": PatternFill(start_color="FFB3B3", end_color="FFB3B3", fill_type="solid"),   # Light red
}
TOTAL_FILL = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")       # Steel blue
NONZERO_FILL = PatternFill(start_color="F2F7F2", end_color="F2F7F2", fill_type="solid")     # Very light green
INFLOW_HEADER_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Light green
INFLOW_TOTAL_FILL = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")   # Medium green
CASH_POS_FILL = PatternFill(start_color="D9D2EA", end_color="D9D2EA", fill_type="solid")       # Light purple
INTEREST_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")       # Light salmon/red
FINANCING_FILL = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")      # Soft peach
FINANCING_TOTAL_FILL = PatternFill(start_color="F4B183", end_color="F4B183", fill_type="solid") # Warm orange

HARD_COSTS_FILL = PatternFill(start_color="C6D9F0", end_color="C6D9F0", fill_type="solid")  # Medium blue

DATA_START_ROW = 6
CODE_COL = 1
DESC_COL = 2
TOTAL_COL = 3
FIRST_WEEK_COL = 4

# Top-level summary accounts for the "Summary CF" sheet.
# Each entry is (code, description). SUMIF uses a wildcard ("0100*") so any
# sub-codes present in the detailed budget (e.g. 010001, 010002) are captured.
SUMMARY_ACCOUNTS: list[tuple[str, str]] = [
    ("0100", "STORY RIGHTS/ACQUISITIONS"),
    ("0200", "SCENARIO"),
    ("0300", "DEVELOPMENT COSTS"),
    ("0400", "PRODUCER"),
    ("0500", "DIRECTOR"),
    ("0600", "STARS"),
    ("1000", "CAST"),
    ("1100", "EXTRAS"),
    ("1200", "PRODUCTION STAFF"),
    ("1300", "DESIGN LABOUR"),
    ("1400", "CONSTRUCTION LABOUR"),
    ("1500", "SET DRESSING LABOUR"),
    ("1600", "PROPERTY LABOUR"),
    ("1700", "SPECIAL EFFECTS LABOUR"),
    ("1800", "WRANGLING LABOUR"),
    ("1900", "WARDROBE LABOUR"),
    ("2000", "MAKEUP/HAIR LABOUR"),
    ("2100", "VIDEO TECHNICAL CREW"),
    ("2200", "CAMERA LABOUR"),
    ("2300", "ELECTRICAL LABOUR"),
    ("2400", "GRIP LABOUR"),
    ("2500", "PRODUCTION SOUND LABOUR"),
    ("2600", "HEALTH AND SAFETY LABOUR"),
    ("2700", "FRINGE BENEFITS"),
    ("2800", "PRODUCTION OFFICE EXPENSES"),
    ("2900", "STUDIO/BACKLOT EXPENSES"),
    ("3000", "LOCATION OFFICE EXPENSES"),
    ("3100", "SITE EXPENSES"),
    ("3200", "UNIT EXPENSES"),
    ("3300", "TRAVEL & LIVING EXPENSES"),
    ("3400", "TRANSPORTATION"),
    ("3500", "CONSTRUCTION MATERIALS"),
    ("3600", "ART SUPPLIES"),
    ("3700", "SET DRESSING"),
    ("3800", "PROPS"),
    ("3900", "SPECIAL EFFECTS"),
    ("4000", "HEALTH AND SAFETY PREVENTION"),
    ("4100", "WARDROBE SUPPLIES"),
    ("4200", "MAKEUP/HAIR SUPPLIES"),
    ("4300", "VIDEO STUDIO FACILITIES"),
    ("4400", "VIDEO REMOTE TECHNICAL FACILITIES"),
    ("4500", "CAMERA EQUIPMENT"),
    ("4600", "ELECTRICAL EQUIPMENT"),
    ("4700", "GRIP EQUIPMENT"),
    ("4800", "SOUND EQUIPMENT"),
    ("4900", "SECOND UNIT"),
    ("5000", "VIDEOTAPE STOCK"),
    ("5100", "PRODUCTION LABORATORY"),
    ("6000", "EDITORIAL LABOUR"),
    ("6100", "EDITORIAL EQUIPMENT"),
    ("6200", "VIDEO POST PRODUCTION (PICTURE)"),
    ("6300", "VIDEO POST PRODUCTION (SOUND)"),
    ("6400", "POST PRODUCTION LABORATORY"),
    ("6500", "FILM POST PRODUCTION SOUND"),
    ("6600", "MUSIC"),
    ("6700", "TITLES/OPTICALS/STOCK FOOTAGE"),
    ("6800", "VERSIONING"),
    ("6900", "AMORTIZATIONS (SERIES)"),
    ("7000", "UNIT PUBLICITY"),
    ("7100", "GENERAL EXPENSES"),
    ("7200", "INDIRECT COSTS"),
    ("8000", "CONTINGENCY"),
    ("8100", "COMPLETION GUARANTEE"),
    ("8200", "COST OF ISSUE"),
]


def _get_outflow_component_codes(
    budget: ParsedBudget | None, cashflow_rows: list
) -> tuple[set[str], set[str]]:
    """Return (financing_codes, internal_oh_codes) for the outflow component rows.

    financing_codes  — rows whose account code normalizes to 7220 (interim financing)
    internal_oh_codes — rows whose Account Details entries carry the group "Internal OH"
    """
    financing_codes: set[str] = set()
    internal_oh_codes: set[str] = set()

    for row in cashflow_rows:
        if row.code.replace(".", "").replace(" ", "").strip().zfill(4) == "7220":
            financing_codes.add(row.code)

    if budget is not None and budget.detail_rows:
        internal_oh_parents: set[str] = set()
        for detail in budget.detail_rows:
            if detail.groups and "internal oh" in detail.groups.lower():
                clean = detail.account.replace(".", "").replace(" ", "").strip()
                if len(clean) >= 4 and clean[:4].isdigit():
                    internal_oh_parents.add(clean[:4])
                elif len(clean) >= 2 and clean[:2].isdigit():
                    internal_oh_parents.add(clean[:2].zfill(2) + "00")

        for row in cashflow_rows:
            if row.code.replace(".", "").replace(" ", "").strip().zfill(4) in internal_oh_parents:
                internal_oh_codes.add(row.code)

    return financing_codes, internal_oh_codes


def _get_summary_code(code: str) -> str:
    """Map a detail account code to its parent summary account code.

    Detail codes share the first 2 digits with their parent summary account:
      0408, 0415 → 0400  |  1312, 1320 → 1300  |  0501 → 0500

    We zero-pad to 4 digits first so numeric codes without leading zeros
    (e.g. '408' stored as the number 408) are handled correctly.
    Returns an empty string if no matching summary account is found.
    """
    padded = code.strip().zfill(4)
    prefix = padded[:2]
    for sc, _ in SUMMARY_ACCOUNTS:
        if sc[:2] == prefix:
            return sc
    return ""


def _get_phase_fill(label: str) -> PatternFill:
    """Return the fill color for a given phase label."""
    upper = label.upper()
    for key, fill in PHASE_COLORS.items():
        if key in upper:
            return fill
    return PatternFill()  # No fill


def _apply_outside_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    """Apply a thin outside border around a rectangular range."""
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=Side(style="thin") if col == min_col else Side(style=None),
                right=Side(style="thin") if col == max_col else Side(style=None),
                top=Side(style="thin") if row == min_row else Side(style=None),
                bottom=Side(style="thin") if row == max_row else Side(style=None),
            )


def _apply_requested_cashflow_formatting(
    ws,
    *,
    max_col: int,
    totals_rows: set[int],
    outflow_min_row: int,
    outflow_max_row: int,
    inflow_min_row: int,
    inflow_max_row: int,
    cash_pos_row: int,
    interest_cost_row: int,
    financing_min_row: int,
    financing_max_row: int,
    interest_rate_row: int,
    keep_paycycle_colors: bool,
) -> None:
    """Normalize sheet formatting to requested style for Cashflow outputs."""
    for row in range(1, ws.max_row + 1):
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name="Calibri", size=10, bold=cell.font.bold)
            cell.alignment = BASE_ALIGNMENT
            cell.border = NO_BORDER
            cell.fill = PatternFill()

            if col == TOTAL_COL and row >= 5:
                cell.font = Font(name="Calibri", size=10, bold=True)

            if row in totals_rows:
                # Financing summary rows only span DESC+TOTAL cols — no week data
                if financing_min_row <= row <= financing_max_row and col > TOTAL_COL:
                    pass
                else:
                    cell.fill = SUBTLE_TOTAL_FILL
                    cell.font = Font(name="Calibri", size=10, bold=True)

            if (
                row >= DATA_START_ROW
                and TOTAL_COL <= col <= max_col
                and row != interest_rate_row
                and cell.number_format != "DD-MMM-YYYY"
                and cell.number_format != "0.00%"
            ):
                cell.number_format = ACCOUNTING_NO_DECIMAL_FORMAT

    if keep_paycycle_colors:
        payroll_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ap_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        for col in range(FIRST_WEEK_COL, max_col + 1):
            pay_cell = ws.cell(row=3, column=col)
            if pay_cell.value == "Payroll":
                pay_cell.fill = payroll_fill
            elif pay_cell.value == "AP":
                pay_cell.fill = ap_fill

    _apply_outside_border(ws, outflow_min_row, outflow_max_row, CODE_COL, max_col)
    _apply_outside_border(ws, inflow_min_row, inflow_max_row, CODE_COL, max_col)
    _apply_outside_border(ws, cash_pos_row, cash_pos_row, CODE_COL, max_col)
    _apply_outside_border(ws, interest_cost_row, interest_cost_row, CODE_COL, max_col)
    _apply_outside_border(ws, financing_min_row, financing_max_row, CODE_COL, TOTAL_COL)


def _write_main_sheet(wb: Workbook, output: CashflowOutput, params: ProductionParameters, budget: ParsedBudget | None = None):
    """Write the main Cashflow sheet."""
    ws = wb.active
    ws.title = "Cashflow"

    num_weeks = len(output.weeks)
    last_week_col = FIRST_WEEK_COL + num_weeks - 1
    last_week_col_letter = get_column_letter(last_week_col)
    # Pre-tax-credit column: one week before the tax credit date (last week + 358 days).
    # Lets the reader judge realistic interest cost just before the TC receipt.
    pre_tc_col = FIRST_WEEK_COL + num_weeks
    pre_tc_col_letter = get_column_letter(pre_tc_col)
    # Tax-credit collection column: last week commencing + 365 days.
    tax_credit_col = FIRST_WEEK_COL + num_weeks + 1
    tc_col_letter = get_column_letter(tax_credit_col)
    # Helper column: summary code sits one further right.
    summary_code_col = FIRST_WEEK_COL + num_weeks + 2

    # Row 1: Title
    ws.cell(row=1, column=1, value=f"{output.title} - Cashflow Forecast").font = TITLE_FONT

    # Row 2: Metadata
    series_number = getattr(params, "series_number", None)
    series_text = f"Series {series_number}" if series_number else ""
    ep_text = f"{params.episode_count} Episodes"
    meta = " | ".join(filter(None, [series_text, ep_text]))
    ws.cell(row=2, column=1, value=meta).font = SUBTITLE_FONT

    # Row 3: Payroll / AP indicator (only if payroll cycle is configured)
    has_payroll = any(w.is_payroll_week is not None for w in output.weeks)
    if has_payroll:
        ws.cell(row=3, column=DESC_COL, value="Pay Cycle").font = Font(bold=True, size=8, color="666666")
        payroll_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ap_fill = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        for i, week in enumerate(output.weeks):
            col = FIRST_WEEK_COL + i
            if week.is_payroll_week is True:
                cell = ws.cell(row=3, column=col, value="Payroll")
                cell.fill = payroll_fill
            elif week.is_payroll_week is False:
                cell = ws.cell(row=3, column=col, value="AP")
                cell.fill = ap_fill
            else:
                cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True, size=7)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # Row 4: Phase labels
    for i, week in enumerate(output.weeks):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=4, column=col, value=week.phase_label)
        cell.font = Font(bold=True, size=8)
        cell.fill = _get_phase_fill(week.phase_label)
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        cell.border = THIN_BORDER

    # Row 5: Week headers (week commencing dates)
    ws.cell(row=5, column=CODE_COL, value="Code").font = HEADER_FONT
    ws.cell(row=5, column=DESC_COL, value="Description").font = HEADER_FONT
    ws.cell(row=5, column=TOTAL_COL, value="Total").font = HEADER_FONT
    ws.cell(row=5, column=CODE_COL).border = THIN_BORDER
    ws.cell(row=5, column=DESC_COL).border = THIN_BORDER
    ws.cell(row=5, column=TOTAL_COL).border = THIN_BORDER

    for i, week in enumerate(output.weeks):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=5, column=col, value=week.week_commencing)
        cell.font = Font(bold=True, size=8)
        cell.number_format = "DD-MMM-YYYY"
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # Pre-tax-credit column: date = last week commencing + 358 days (one week before TC)
    pre_tc_date = output.weeks[-1].week_commencing + timedelta(days=358)
    pre_tc_phase_cell = ws.cell(row=4, column=pre_tc_col, value="PRE TAX CREDIT")
    pre_tc_phase_cell.font = Font(bold=True, size=8)
    pre_tc_phase_cell.fill = PatternFill(start_color="EAF4D3", end_color="EAF4D3", fill_type="solid")
    pre_tc_phase_cell.alignment = Alignment(horizontal="center", text_rotation=90)
    pre_tc_phase_cell.border = THIN_BORDER
    pre_tc_date_cell = ws.cell(row=5, column=pre_tc_col, value=pre_tc_date)
    pre_tc_date_cell.font = Font(bold=True, size=8)
    pre_tc_date_cell.number_format = "DD-MMM-YYYY"
    pre_tc_date_cell.alignment = Alignment(horizontal="center")
    pre_tc_date_cell.border = THIN_BORDER

    # Tax-credit column: date = last week commencing + 365 days
    tc_date = output.weeks[-1].week_commencing + timedelta(days=365)
    tc_phase_cell = ws.cell(row=4, column=tax_credit_col, value="TAX CREDIT")
    tc_phase_cell.font = Font(bold=True, size=8)
    tc_phase_cell.fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
    tc_phase_cell.alignment = Alignment(horizontal="center", text_rotation=90)
    tc_phase_cell.border = THIN_BORDER
    tc_date_cell = ws.cell(row=5, column=tax_credit_col, value=tc_date)
    tc_date_cell.font = Font(bold=True, size=8)
    tc_date_cell.number_format = "DD-MMM-YYYY"
    tc_date_cell.alignment = Alignment(horizontal="center")
    tc_date_cell.border = THIN_BORDER

    # Row 5: summary code helper column header
    sc_hdr = ws.cell(row=5, column=summary_code_col, value="Smry Code")
    sc_hdr.font = Font(bold=True, size=8, color="999999")
    sc_hdr.alignment = Alignment(horizontal="center")
    sc_hdr.border = THIN_BORDER

    # Data rows
    first_data_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_data_col_letter = get_column_letter(last_week_col)
    for row_idx, row_data in enumerate(output.rows):
        excel_row = DATA_START_ROW + row_idx

        ws.cell(row=excel_row, column=CODE_COL, value=row_data.code).border = THIN_BORDER
        ws.cell(row=excel_row, column=DESC_COL, value=row_data.description).border = THIN_BORDER

        # Total column with SUM formula
        total_cell = ws.cell(
            row=excel_row,
            column=TOTAL_COL,
            value=f"=SUM({first_data_col_letter}{excel_row}:{last_data_col_letter}{excel_row})",
        )
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font = Font(bold=True)
        total_cell.border = THIN_BORDER

        # Weekly amounts
        for col_offset, amount in enumerate(row_data.weekly_amounts):
            col = FIRST_WEEK_COL + col_offset
            cell = ws.cell(row=excel_row, column=col, value=round(amount, 2) if amount else 0)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            if amount and amount > 0:
                cell.fill = NONZERO_FILL

        # Tax-credit column — blank/zero placeholder for each detail row
        tc_cell = ws.cell(row=excel_row, column=tax_credit_col, value=0)
        tc_cell.number_format = CURRENCY_FORMAT
        tc_cell.border = THIN_BORDER

        # Summary code — Python-computed so it's always clean text regardless
        # of how the original budget stored numeric codes
        sc_val = _get_summary_code(row_data.code)
        sc_cell = ws.cell(row=excel_row, column=summary_code_col, value=sc_val)
        sc_cell.font = Font(size=8, color="999999")
        sc_cell.border = THIN_BORDER

    # Weekly totals row
    totals_row = DATA_START_ROW + len(output.rows)
    ws.cell(row=totals_row, column=DESC_COL, value="WEEKLY TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=totals_row, column=DESC_COL).border = THIN_BORDER

    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=totals_row,
            column=col,
            value=f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{totals_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.fill = TOTAL_FILL
        cell.border = THIN_BORDER

    # Grand total for totals row
    grand_total_cell = ws.cell(
        row=totals_row,
        column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{totals_row}:{last_data_col_letter}{totals_row})",
    )
    grand_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    grand_total_cell.font = Font(bold=True)
    grand_total_cell.fill = TOTAL_FILL
    grand_total_cell.border = THIN_BORDER

    # Cumulative totals row
    cum_row = totals_row + 1
    ws.cell(row=cum_row, column=DESC_COL, value="CUMULATIVE TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=cum_row, column=DESC_COL).border = THIN_BORDER

    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        if i == 0:
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=cum_row, column=col, value=f"={col_letter}{totals_row}"
            )
        else:
            prev_col_letter = get_column_letter(col - 1)
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=cum_row,
                column=col,
                value=f"={prev_col_letter}{cum_row}+{col_letter}{totals_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True, italic=True)
        cell.border = THIN_BORDER

    # Cash Inflows section (2 blank rows below CUMULATIVE TOTAL)
    inflow_header_row = cum_row + 3  # cum_row+1 and cum_row+2 are blank
    ws.cell(row=inflow_header_row, column=DESC_COL, value="CASH INFLOWS").font = Font(bold=True, size=11)
    ws.cell(row=inflow_header_row, column=DESC_COL).fill = INFLOW_HEADER_FILL
    ws.cell(row=inflow_header_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=inflow_header_row, column=col)
        cell.fill = INFLOW_HEADER_FILL
        cell.border = THIN_BORDER
    for extra_col in (pre_tc_col, tax_credit_col):
        c = ws.cell(row=inflow_header_row, column=extra_col)
        c.fill = INFLOW_HEADER_FILL
        c.border = THIN_BORDER

    # Inflow data rows
    inflow_data_start = inflow_header_row + 1
    for row_idx, inflow_row in enumerate(output.cash_inflows):
        excel_row = inflow_data_start + row_idx
        ws.cell(row=excel_row, column=DESC_COL, value=inflow_row.label).border = THIN_BORDER
        total_cell = ws.cell(
            row=excel_row,
            column=TOTAL_COL,
            value=f"=SUM({first_data_col_letter}{excel_row}:{tc_col_letter}{excel_row})",
        )
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font = Font(bold=True)
        total_cell.border = THIN_BORDER
        for col_offset, amount in enumerate(inflow_row.weekly_amounts):
            col = FIRST_WEEK_COL + col_offset
            cell = ws.cell(row=excel_row, column=col, value=round(amount, 2) if amount else 0)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
        # Pre-TC and TC inflow placeholders — user enters amounts here
        for extra_col in (pre_tc_col, tax_credit_col):
            ec = ws.cell(row=excel_row, column=extra_col, value=0)
            ec.number_format = CURRENCY_FORMAT
            ec.border = THIN_BORDER

    # Weekly inflow total row
    inflow_total_row = inflow_data_start + len(output.cash_inflows)
    ws.cell(row=inflow_total_row, column=DESC_COL, value="WEEKLY INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=inflow_total_row, column=DESC_COL).border = THIN_BORDER
    ws.cell(row=inflow_total_row, column=DESC_COL).fill = INFLOW_TOTAL_FILL
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=inflow_total_row,
            column=col,
            value=f"=SUM({col_letter}{inflow_data_start}:{col_letter}{inflow_total_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.fill = INFLOW_TOTAL_FILL
        cell.border = THIN_BORDER
    # Pre-TC and TC weekly inflow totals
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(
            row=inflow_total_row, column=extra_col,
            value=f"=SUM({extra_letter}{inflow_data_start}:{extra_letter}{inflow_total_row - 1})",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font = Font(bold=True)
        ec.fill = INFLOW_TOTAL_FILL
        ec.border = THIN_BORDER
    first_data_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_data_col_letter = get_column_letter(last_week_col)
    inflow_grand_cell = ws.cell(
        row=inflow_total_row,
        column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{inflow_total_row}:{tc_col_letter}{inflow_total_row})",
    )
    inflow_grand_cell.number_format = CURRENCY_FORMAT_TOTAL
    inflow_grand_cell.font = Font(bold=True)
    inflow_grand_cell.fill = INFLOW_TOTAL_FILL
    inflow_grand_cell.border = THIN_BORDER

    # Cumulative inflow total row
    inflow_cum_row = inflow_total_row + 1
    ws.cell(row=inflow_cum_row, column=DESC_COL, value="CUMULATIVE INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=inflow_cum_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        if i == 0:
            col_letter = get_column_letter(col)
            cell = ws.cell(row=inflow_cum_row, column=col, value=f"={col_letter}{inflow_total_row}")
        else:
            prev_col_letter = get_column_letter(col - 1)
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=inflow_cum_row,
                column=col,
                value=f"={prev_col_letter}{inflow_cum_row}+{col_letter}{inflow_total_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True, italic=True)
        cell.border = THIN_BORDER
    # Pre-TC cumulative inflow = last week's running total + pre_tc weekly inflow
    pre_tc_cum_cell = ws.cell(
        row=inflow_cum_row, column=pre_tc_col,
        value=f"={last_week_col_letter}{inflow_cum_row}+{pre_tc_col_letter}{inflow_total_row}",
    )
    pre_tc_cum_cell.number_format = CURRENCY_FORMAT_TOTAL
    pre_tc_cum_cell.font = Font(bold=True, italic=True)
    pre_tc_cum_cell.border = THIN_BORDER
    # TC cumulative inflow = pre_tc running total + tc weekly inflow
    tc_cum_inflow = ws.cell(
        row=inflow_cum_row, column=tax_credit_col,
        value=f"={pre_tc_col_letter}{inflow_cum_row}+{tc_col_letter}{inflow_total_row}",
    )
    tc_cum_inflow.number_format = CURRENCY_FORMAT_TOTAL
    tc_cum_inflow.font = Font(bold=True, italic=True)
    tc_cum_inflow.border = THIN_BORDER

    # Cumulative cash position (2 rows below cumulative inflow total)
    # = cumulative inflows − cumulative outflows for each week
    cash_pos_row = inflow_cum_row + 2
    ws.cell(row=cash_pos_row, column=DESC_COL, value="CUMULATIVE CASH POSITION").font = Font(bold=True, size=11)
    ws.cell(row=cash_pos_row, column=DESC_COL).fill = CASH_POS_FILL
    ws.cell(row=cash_pos_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=cash_pos_row,
            column=col,
            value=f"={col_letter}{inflow_cum_row}-{col_letter}{cum_row}",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.fill = CASH_POS_FILL
        cell.border = THIN_BORDER
    # Pre-TC and TC cash positions: cumulative inflow − last week cumulative outflow
    # (no new outflows after the schedule ends, so outflows stay at last week's level)
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(
            row=cash_pos_row, column=extra_col,
            value=f"={extra_letter}{inflow_cum_row}-{last_week_col_letter}{cum_row}",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font = Font(bold=True)
        ec.fill = CASH_POS_FILL
        ec.border = THIN_BORDER

    # Interest cost (2 rows below cumulative cash position)
    # = ABS(cash_position) × rate × (days in period) / 365  — only when position is negative
    # The rate lives in TOTAL_COL of interest_rate_row (directly below), so it's easy to adjust.
    interest_cost_row = cash_pos_row + 2
    # Interest rate row sits immediately below the interest cost row.
    # Defining it here so the interest cost formulas can reference it.
    interest_rate_row = interest_cost_row + 1
    rate_cell_ref = f"$C${interest_rate_row}"  # absolute ref to the rate value

    ws.cell(row=interest_cost_row, column=DESC_COL, value="INTEREST COST").font = Font(bold=True, size=11)
    ws.cell(row=interest_cost_row, column=DESC_COL).fill = INTEREST_FILL
    ws.cell(row=interest_cost_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        # Days in this period = next week date − this week date.
        # For the last week, mirror the previous interval (same logic, reversed).
        if num_weeks == 1:
            days_formula = "7"
        elif i < num_weeks - 1:
            next_col_letter = get_column_letter(col + 1)
            days_formula = f"({next_col_letter}5-{col_letter}5)"
        else:
            prev_col_letter = get_column_letter(col - 1)
            days_formula = f"({col_letter}5-{prev_col_letter}5)"
        cell = ws.cell(
            row=interest_cost_row,
            column=col,
            value=f"=IF({col_letter}{cash_pos_row}<0,ABS({col_letter}{cash_pos_row})*{rate_cell_ref}*{days_formula}/365,0)",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.fill = INTEREST_FILL
        cell.border = THIN_BORDER
    # Pre-TC interest cost: days = pre_tc_date − last_week_date (≈ 358)
    pre_tc_interest = ws.cell(
        row=interest_cost_row, column=pre_tc_col,
        value=f"=IF({pre_tc_col_letter}{cash_pos_row}<0,ABS({pre_tc_col_letter}{cash_pos_row})*{rate_cell_ref}*({pre_tc_col_letter}5-{last_week_col_letter}5)/365,0)",
    )
    pre_tc_interest.number_format = CURRENCY_FORMAT_TOTAL
    pre_tc_interest.font = Font(bold=True)
    pre_tc_interest.fill = INTEREST_FILL
    pre_tc_interest.border = THIN_BORDER
    # TC interest cost: days = tc_date − pre_tc_date (= 7)
    tc_interest = ws.cell(
        row=interest_cost_row, column=tax_credit_col,
        value=f"=IF({tc_col_letter}{cash_pos_row}<0,ABS({tc_col_letter}{cash_pos_row})*{rate_cell_ref}*({tc_col_letter}5-{pre_tc_col_letter}5)/365,0)",
    )
    tc_interest.number_format = CURRENCY_FORMAT_TOTAL
    tc_interest.font = Font(bold=True)
    tc_interest.fill = INTEREST_FILL
    tc_interest.border = THIN_BORDER
    # Grand total: sum of all weekly + pre-TC + TC interest costs
    interest_grand_cell = ws.cell(
        row=interest_cost_row,
        column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{interest_cost_row}:{tc_col_letter}{interest_cost_row})",
    )
    interest_grand_cell.number_format = CURRENCY_FORMAT_TOTAL
    interest_grand_cell.font = Font(bold=True)
    interest_grand_cell.fill = INTEREST_FILL
    interest_grand_cell.border = THIN_BORDER

    # Interest rate row — label + editable rate value on the Cashflow sheet
    ws.cell(row=interest_rate_row, column=DESC_COL, value="Annual Interest Rate").font = Font(
        bold=True, size=10, italic=True
    )
    ws.cell(row=interest_rate_row, column=DESC_COL).border = THIN_BORDER
    rate_value_cell = ws.cell(row=interest_rate_row, column=TOTAL_COL, value=0.065)
    rate_value_cell.number_format = "0.00%"
    rate_value_cell.font = Font(bold=True)
    rate_value_cell.border = THIN_BORDER

    # Financing cost summary (2 rows below Annual Interest Rate)
    # Rows: Interest Cost | Setup Fee | Legal Cost | Total
    fin_interest_row = interest_rate_row + 2
    fin_setup_row    = fin_interest_row + 1
    fin_legal_row    = fin_setup_row + 1
    fin_total_row    = fin_legal_row + 1

    # Interest Cost — links to the grand total from the INTEREST COST row
    ws.cell(row=fin_interest_row, column=DESC_COL, value="Interest Cost").font = Font(bold=True)
    ws.cell(row=fin_interest_row, column=DESC_COL).fill = FINANCING_FILL
    ws.cell(row=fin_interest_row, column=DESC_COL).border = THIN_BORDER
    fin_interest_cell = ws.cell(
        row=fin_interest_row, column=TOTAL_COL,
        value=f"=C{interest_cost_row}",
    )
    fin_interest_cell.number_format = CURRENCY_FORMAT_TOTAL
    fin_interest_cell.font = Font(bold=True)
    fin_interest_cell.fill = FINANCING_FILL
    fin_interest_cell.border = THIN_BORDER

    # Setup Fee — 1.5% of the peak (most negative) cumulative cash position
    ws.cell(row=fin_setup_row, column=DESC_COL, value="Setup Fee (1.5% of peak loan)").font = Font(bold=True)
    ws.cell(row=fin_setup_row, column=DESC_COL).fill = FINANCING_FILL
    ws.cell(row=fin_setup_row, column=DESC_COL).border = THIN_BORDER
    cash_pos_range = f"{first_data_col_letter}{cash_pos_row}:{last_data_col_letter}{cash_pos_row}"
    fin_setup_cell = ws.cell(
        row=fin_setup_row, column=TOTAL_COL,
        value=f"=IF(MIN({cash_pos_range})<0,-MIN({cash_pos_range})*0.015,0)",
    )
    fin_setup_cell.number_format = CURRENCY_FORMAT_TOTAL
    fin_setup_cell.font = Font(bold=True)
    fin_setup_cell.fill = FINANCING_FILL
    fin_setup_cell.border = THIN_BORDER

    # Legal Cost — fixed amount
    ws.cell(row=fin_legal_row, column=DESC_COL, value="Legal Cost").font = Font(bold=True)
    ws.cell(row=fin_legal_row, column=DESC_COL).fill = FINANCING_FILL
    ws.cell(row=fin_legal_row, column=DESC_COL).border = THIN_BORDER
    fin_legal_cell = ws.cell(row=fin_legal_row, column=TOTAL_COL, value=5000)
    fin_legal_cell.number_format = CURRENCY_FORMAT_TOTAL
    fin_legal_cell.font = Font(bold=True)
    fin_legal_cell.fill = FINANCING_FILL
    fin_legal_cell.border = THIN_BORDER

    # Total — sum of the three items above
    ws.cell(row=fin_total_row, column=DESC_COL, value="TOTAL FINANCING COST").font = Font(bold=True, size=11)
    ws.cell(row=fin_total_row, column=DESC_COL).fill = FINANCING_TOTAL_FILL
    ws.cell(row=fin_total_row, column=DESC_COL).border = THIN_BORDER
    fin_total_cell = ws.cell(
        row=fin_total_row, column=TOTAL_COL,
        value=f"=SUM(C{fin_interest_row}:C{fin_legal_row})",
    )
    fin_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    fin_total_cell.font = Font(bold=True)
    fin_total_cell.fill = FINANCING_TOTAL_FILL
    fin_total_cell.border = THIN_BORDER

    # Outflow components + hard costs (starting 2 rows below TOTAL FINANCING COST)
    internals_outflow_row = fin_total_row + 2
    financing_outflow_row = fin_total_row + 3
    hard_costs_row        = fin_total_row + 4

    financing_codes, internal_oh_codes = _get_outflow_component_codes(budget, output.rows)

    internals_weekly = [0.0] * num_weeks
    internals_total  = 0.0
    fin_out_weekly   = [0.0] * num_weeks
    fin_out_total    = 0.0
    for row_data in output.rows:
        if row_data.code in internal_oh_codes:
            internals_total += row_data.total
            for j, amount in enumerate(row_data.weekly_amounts):
                internals_weekly[j] += amount
        if row_data.code in financing_codes:
            fin_out_total += row_data.total
            for j, amount in enumerate(row_data.weekly_amounts):
                fin_out_weekly[j] += amount

    # INTERNALS OUTFLOW row
    ws.cell(row=internals_outflow_row, column=DESC_COL, value="INTERNALS OUTFLOW").font = Font(bold=True, size=11)
    ws.cell(row=internals_outflow_row, column=DESC_COL).border = THIN_BORDER
    int_total_cell = ws.cell(row=internals_outflow_row, column=TOTAL_COL, value=round(internals_total, 2))
    int_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    int_total_cell.font = Font(bold=True)
    int_total_cell.border = THIN_BORDER
    for i, amount in enumerate(internals_weekly):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=internals_outflow_row, column=col, value=round(amount, 2) if amount else 0)
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    # FINANCING OUTFLOW row
    ws.cell(row=financing_outflow_row, column=DESC_COL, value="FINANCING OUTFLOW").font = Font(bold=True, size=11)
    ws.cell(row=financing_outflow_row, column=DESC_COL).border = THIN_BORDER
    fin_out_total_cell = ws.cell(row=financing_outflow_row, column=TOTAL_COL, value=round(fin_out_total, 2))
    fin_out_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    fin_out_total_cell.font = Font(bold=True)
    fin_out_total_cell.border = THIN_BORDER
    for i, amount in enumerate(fin_out_weekly):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=financing_outflow_row, column=col, value=round(amount, 2) if amount else 0)
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    # HARD COSTS OUTFLOW row — formula: WEEKLY TOTAL minus internals and financing
    ws.cell(row=hard_costs_row, column=DESC_COL, value="HARD COSTS OUTFLOW").font = Font(bold=True, size=11)
    ws.cell(row=hard_costs_row, column=DESC_COL).border = THIN_BORDER
    hc_total_cell = ws.cell(
        row=hard_costs_row, column=TOTAL_COL,
        value=f"=C{totals_row}-C{internals_outflow_row}-C{financing_outflow_row}",
    )
    hc_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    hc_total_cell.font = Font(bold=True)
    hc_total_cell.border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=hard_costs_row, column=col,
            value=f"={col_letter}{totals_row}-{col_letter}{internals_outflow_row}-{col_letter}{financing_outflow_row}",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font = Font(bold=True)
        cell.border = THIN_BORDER

    _apply_requested_cashflow_formatting(
        ws,
        max_col=tax_credit_col,
        totals_rows={totals_row, cum_row, inflow_total_row, inflow_cum_row, fin_total_row},
        outflow_min_row=5,
        outflow_max_row=cum_row,
        inflow_min_row=inflow_header_row,
        inflow_max_row=inflow_cum_row,
        cash_pos_row=cash_pos_row,
        interest_cost_row=interest_cost_row,
        financing_min_row=fin_interest_row,
        financing_max_row=fin_total_row,
        interest_rate_row=interest_rate_row,
        keep_paycycle_colors=has_payroll,
    )

    # Apply fills and borders after formatting (formatting resets all fills)
    for component_row in (internals_outflow_row, financing_outflow_row):
        for col in range(CODE_COL, last_week_col + 1):
            ws.cell(row=component_row, column=col).fill = SUBTLE_TOTAL_FILL
        _apply_outside_border(ws, component_row, component_row, CODE_COL, last_week_col)
    for col in range(CODE_COL, last_week_col + 1):
        ws.cell(row=hard_costs_row, column=col).fill = HARD_COSTS_FILL
    _apply_outside_border(ws, hard_costs_row, hard_costs_row, CODE_COL, last_week_col)

    # Column widths
    ws.column_dimensions[get_column_letter(CODE_COL)].width = 10
    ws.column_dimensions[get_column_letter(DESC_COL)].width = 35
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 15
    for i in range(num_weeks):
        ws.column_dimensions[get_column_letter(FIRST_WEEK_COL + i)].width = 12
    ws.column_dimensions[get_column_letter(pre_tc_col)].width = 14
    ws.column_dimensions[get_column_letter(tax_credit_col)].width = 14
    ws.column_dimensions[get_column_letter(summary_code_col)].width = 9

    # Freeze panes: fix code/desc/total columns and header rows
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=FIRST_WEEK_COL)


def _write_summary_cf_sheet(wb: Workbook, output: CashflowOutput, params: ProductionParameters) -> None:
    """Write the 'Summary CF' sheet.

    Each outflow row uses SUMIF against the detailed Cashflow sheet so any
    sub-codes (e.g. 010001, 010002) roll up into their parent (0100).
    Inflow values are linked directly; cumulative cash position, interest cost
    and financing summary are all formula-driven and consistent with the detail.
    """
    DETAIL = "Cashflow"
    ws = wb.create_sheet("Summary CF")

    num_weeks = len(output.weeks)
    last_week_col = FIRST_WEEK_COL + num_weeks - 1
    last_week_col_letter = get_column_letter(last_week_col)
    first_data_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_data_col_letter = get_column_letter(last_week_col)
    pre_tc_col = FIRST_WEEK_COL + num_weeks
    pre_tc_col_letter = get_column_letter(pre_tc_col)
    tax_credit_col = FIRST_WEEK_COL + num_weeks + 1
    tc_col_letter = get_column_letter(tax_credit_col)
    # Summary code column in the Cashflow sheet (matches _write_main_sheet).
    # Sits three columns after the last week column: +1 = pre-TC, +2 = TC, +3 = smry code.
    detail_summary_code_col_letter = get_column_letter(FIRST_WEEK_COL + num_weeks + 2)

    # Mirror the detail sheet row positions (must stay in sync with _write_main_sheet)
    detail_totals_row       = DATA_START_ROW + len(output.rows)
    detail_cum_row          = detail_totals_row + 1
    detail_inflow_hdr_row   = detail_cum_row + 3
    detail_inflow_data_start = detail_inflow_hdr_row + 1
    detail_inflow_total_row = detail_inflow_data_start + len(output.cash_inflows)
    detail_inflow_cum_row   = detail_inflow_total_row + 1
    detail_cash_pos_row     = detail_inflow_cum_row + 2
    detail_interest_cost_row = detail_cash_pos_row + 2
    detail_interest_rate_row = detail_interest_cost_row + 1
    detail_fin_legal_row    = detail_interest_rate_row + 4  # interest_rate+2+1+1

    # ── Rows 1-2: title / metadata ───────────────────────────────────────────
    ws.cell(row=1, column=1, value=f"{output.title} - Summary Cashflow Forecast").font = TITLE_FONT
    series_number = getattr(params, "series_number", None)
    series_text = f"Series {series_number}" if series_number else ""
    ep_text = f"{params.episode_count} Episodes"
    meta = " | ".join(filter(None, [series_text, ep_text]))
    ws.cell(row=2, column=1, value=meta).font = SUBTITLE_FONT

    # ── Row 3: pay cycle indicator ───────────────────────────────────────────
    has_payroll = any(w.is_payroll_week is not None for w in output.weeks)
    if has_payroll:
        ws.cell(row=3, column=DESC_COL, value="Pay Cycle").font = Font(bold=True, size=8, color="666666")
        payroll_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
        ap_fill      = PatternFill(start_color="FFF3E0", end_color="FFF3E0", fill_type="solid")
        for i, week in enumerate(output.weeks):
            col = FIRST_WEEK_COL + i
            if week.is_payroll_week is True:
                cell = ws.cell(row=3, column=col, value="Payroll")
                cell.fill = payroll_fill
            elif week.is_payroll_week is False:
                cell = ws.cell(row=3, column=col, value="AP")
                cell.fill = ap_fill
            else:
                cell = ws.cell(row=3, column=col)
            cell.font = Font(bold=True, size=7)
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER

    # ── Row 4: phase labels ──────────────────────────────────────────────────
    for i, week in enumerate(output.weeks):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=4, column=col, value=week.phase_label)
        cell.font = Font(bold=True, size=8)
        cell.fill = _get_phase_fill(week.phase_label)
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        cell.border = THIN_BORDER
    # Pre-TC and TC phase labels — linked from Cashflow sheet row 4
    for extra_col, extra_letter, fill_color in (
        (pre_tc_col, pre_tc_col_letter, "EAF4D3"),
        (tax_credit_col, tc_col_letter, "D9EAD3"),
    ):
        ec = ws.cell(row=4, column=extra_col, value=f"={DETAIL}!{extra_letter}4")
        ec.font = Font(bold=True, size=8)
        ec.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        ec.alignment = Alignment(horizontal="center", text_rotation=90)
        ec.border = THIN_BORDER

    # ── Row 5: column headers — dates linked from Cashflow sheet ─────────────
    ws.cell(row=5, column=CODE_COL,   value="Code").font        = HEADER_FONT
    ws.cell(row=5, column=DESC_COL,   value="Description").font = HEADER_FONT
    ws.cell(row=5, column=TOTAL_COL,  value="Total").font       = HEADER_FONT
    ws.cell(row=5, column=CODE_COL).border  = THIN_BORDER
    ws.cell(row=5, column=DESC_COL).border  = THIN_BORDER
    ws.cell(row=5, column=TOTAL_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(row=5, column=col, value=f"={DETAIL}!{col_letter}5")
        cell.font = Font(bold=True, size=8)
        cell.number_format = "DD-MMM-YYYY"
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    # Pre-TC and TC dates — linked from Cashflow sheet row 5
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(row=5, column=extra_col, value=f"={DETAIL}!{extra_letter}5")
        ec.font = Font(bold=True, size=8)
        ec.number_format = "DD-MMM-YYYY"
        ec.alignment = Alignment(horizontal="center")
        ec.border = THIN_BORDER

    # ── Summary account rows ─────────────────────────────────────────────────
    for row_idx, (code, description) in enumerate(SUMMARY_ACCOUNTS):
        excel_row = DATA_START_ROW + row_idx
        ws.cell(row=excel_row, column=CODE_COL,  value=code).border        = THIN_BORDER
        ws.cell(row=excel_row, column=DESC_COL,  value=description).border = THIN_BORDER

        total_cell = ws.cell(
            row=excel_row, column=TOTAL_COL,
            value=f"=SUM({first_data_col_letter}{excel_row}:{last_data_col_letter}{excel_row})",
        )
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font   = Font(bold=True)
        total_cell.border = THIN_BORDER

        for i in range(num_weeks):
            col = FIRST_WEEK_COL + i
            col_letter = get_column_letter(col)
            # SUMIF against the pre-computed summary code column (exact match).
            # Using the helper column avoids all issues with numeric codes /
            # leading-zero stripping that break wildcard matching on column A.
            cell = ws.cell(
                row=excel_row, column=col,
                value=f'=SUMIF({DETAIL}!${detail_summary_code_col_letter}:${detail_summary_code_col_letter},"{code}",{DETAIL}!{col_letter}:{col_letter})',
            )
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER

    # ── Weekly totals row ────────────────────────────────────────────────────
    sum_totals_row = DATA_START_ROW + len(SUMMARY_ACCOUNTS)
    ws.cell(row=sum_totals_row, column=DESC_COL, value="WEEKLY TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_totals_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_totals_row, column=col,
            value=f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{sum_totals_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font  = Font(bold=True)
        cell.fill  = TOTAL_FILL
        cell.border = THIN_BORDER
    grand_total_cell = ws.cell(
        row=sum_totals_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_totals_row}:{last_data_col_letter}{sum_totals_row})",
    )
    grand_total_cell.number_format = CURRENCY_FORMAT_TOTAL
    grand_total_cell.font  = Font(bold=True)
    grand_total_cell.fill  = TOTAL_FILL
    grand_total_cell.border = THIN_BORDER

    # ── Cumulative totals row ────────────────────────────────────────────────
    sum_cum_row = sum_totals_row + 1
    ws.cell(row=sum_cum_row, column=DESC_COL, value="CUMULATIVE TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_cum_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        if i == 0:
            cell = ws.cell(row=sum_cum_row, column=col, value=f"={col_letter}{sum_totals_row}")
        else:
            prev_col = get_column_letter(col - 1)
            cell = ws.cell(
                row=sum_cum_row, column=col,
                value=f"={prev_col}{sum_cum_row}+{col_letter}{sum_totals_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font  = Font(bold=True, italic=True)
        cell.border = THIN_BORDER

    # ── Cash Inflows section (2 blank rows below cumulative) ─────────────────
    sum_inflow_hdr_row   = sum_cum_row + 3
    sum_inflow_data_start = sum_inflow_hdr_row + 1
    sum_inflow_total_row = sum_inflow_data_start + len(output.cash_inflows)
    sum_inflow_cum_row   = sum_inflow_total_row + 1

    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL, value="CASH INFLOWS").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL).fill   = INFLOW_HEADER_FILL
    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        cell = ws.cell(row=sum_inflow_hdr_row, column=col)
        cell.fill   = INFLOW_HEADER_FILL
        cell.border = THIN_BORDER
    for extra_col in (pre_tc_col, tax_credit_col):
        c = ws.cell(row=sum_inflow_hdr_row, column=extra_col)
        c.fill   = INFLOW_HEADER_FILL
        c.border = THIN_BORDER

    for row_idx, inflow_row in enumerate(output.cash_inflows):
        excel_row        = sum_inflow_data_start + row_idx
        detail_inflow_row = detail_inflow_data_start + row_idx
        ws.cell(row=excel_row, column=DESC_COL, value=inflow_row.label).border = THIN_BORDER
        total_cell = ws.cell(
            row=excel_row, column=TOTAL_COL,
            value=f"={DETAIL}!C{detail_inflow_row}",
        )
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font   = Font(bold=True)
        total_cell.border = THIN_BORDER
        for i in range(num_weeks):
            col = FIRST_WEEK_COL + i
            col_letter = get_column_letter(col)
            cell = ws.cell(
                row=excel_row, column=col,
                value=f"={DETAIL}!{col_letter}{detail_inflow_row}",
            )
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
        # Pre-TC and TC inflows — linked from Cashflow sheet
        for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
            ec = ws.cell(
                row=excel_row, column=extra_col,
                value=f"={DETAIL}!{extra_letter}{detail_inflow_row}",
            )
            ec.number_format = CURRENCY_FORMAT
            ec.border = THIN_BORDER

    # Weekly inflow total
    ws.cell(row=sum_inflow_total_row, column=DESC_COL, value="WEEKLY INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_total_row, column=DESC_COL).fill   = INFLOW_TOTAL_FILL
    ws.cell(row=sum_inflow_total_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_inflow_total_row, column=col,
            value=f"=SUM({col_letter}{sum_inflow_data_start}:{col_letter}{sum_inflow_total_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = INFLOW_TOTAL_FILL
        cell.border = THIN_BORDER
    # Pre-TC and TC weekly inflow totals
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(
            row=sum_inflow_total_row, column=extra_col,
            value=f"=SUM({extra_letter}{sum_inflow_data_start}:{extra_letter}{sum_inflow_total_row - 1})",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True)
        ec.fill   = INFLOW_TOTAL_FILL
        ec.border = THIN_BORDER
    inflow_grand = ws.cell(
        row=sum_inflow_total_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_inflow_total_row}:{tc_col_letter}{sum_inflow_total_row})",
    )
    inflow_grand.number_format = CURRENCY_FORMAT_TOTAL
    inflow_grand.font   = Font(bold=True)
    inflow_grand.fill   = INFLOW_TOTAL_FILL
    inflow_grand.border = THIN_BORDER

    # Cumulative inflow total
    ws.cell(row=sum_inflow_cum_row, column=DESC_COL, value="CUMULATIVE INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_cum_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        if i == 0:
            cell = ws.cell(row=sum_inflow_cum_row, column=col, value=f"={col_letter}{sum_inflow_total_row}")
        else:
            prev_col = get_column_letter(col - 1)
            cell = ws.cell(
                row=sum_inflow_cum_row, column=col,
                value=f"={prev_col}{sum_inflow_cum_row}+{col_letter}{sum_inflow_total_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True, italic=True)
        cell.border = THIN_BORDER
    # Pre-TC cumulative = last week running + pre_tc weekly
    pre_tc_cum = ws.cell(
        row=sum_inflow_cum_row, column=pre_tc_col,
        value=f"={last_week_col_letter}{sum_inflow_cum_row}+{pre_tc_col_letter}{sum_inflow_total_row}",
    )
    pre_tc_cum.number_format = CURRENCY_FORMAT_TOTAL
    pre_tc_cum.font   = Font(bold=True, italic=True)
    pre_tc_cum.border = THIN_BORDER
    # TC cumulative = pre_tc running + tc weekly
    tc_cum_inflow = ws.cell(
        row=sum_inflow_cum_row, column=tax_credit_col,
        value=f"={pre_tc_col_letter}{sum_inflow_cum_row}+{tc_col_letter}{sum_inflow_total_row}",
    )
    tc_cum_inflow.number_format = CURRENCY_FORMAT_TOTAL
    tc_cum_inflow.font   = Font(bold=True, italic=True)
    tc_cum_inflow.border = THIN_BORDER

    # ── Cumulative cash position (2 rows below cumulative inflow total) ───────
    sum_cash_pos_row = sum_inflow_cum_row + 2
    ws.cell(row=sum_cash_pos_row, column=DESC_COL, value="CUMULATIVE CASH POSITION").font = Font(bold=True, size=11)
    ws.cell(row=sum_cash_pos_row, column=DESC_COL).fill   = CASH_POS_FILL
    ws.cell(row=sum_cash_pos_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_cash_pos_row, column=col,
            value=f"={col_letter}{sum_inflow_cum_row}-{col_letter}{sum_cum_row}",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = CASH_POS_FILL
        cell.border = THIN_BORDER
    # Pre-TC and TC cash positions: cumulative inflow − last week cumulative outflow
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(
            row=sum_cash_pos_row, column=extra_col,
            value=f"={extra_letter}{sum_inflow_cum_row}-{last_week_col_letter}{sum_cum_row}",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True)
        ec.fill   = CASH_POS_FILL
        ec.border = THIN_BORDER

    # ── Interest cost (2 rows below cumulative cash position) ────────────────
    # Re-computed using this sheet's cash position; rate linked from Cashflow sheet.
    sum_interest_cost_row = sum_cash_pos_row + 2
    sum_interest_rate_row = sum_interest_cost_row + 1
    rate_ref = f"{DETAIL}!$C${detail_interest_rate_row}"

    ws.cell(row=sum_interest_cost_row, column=DESC_COL, value="INTEREST COST").font = Font(bold=True, size=11)
    ws.cell(row=sum_interest_cost_row, column=DESC_COL).fill   = INTEREST_FILL
    ws.cell(row=sum_interest_cost_row, column=DESC_COL).border = THIN_BORDER
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        if num_weeks == 1:
            days_formula = "7"
        elif i < num_weeks - 1:
            days_formula = f"({get_column_letter(col + 1)}5-{col_letter}5)"
        else:
            days_formula = f"({col_letter}5-{get_column_letter(col - 1)}5)"
        cell = ws.cell(
            row=sum_interest_cost_row, column=col,
            value=f"=IF({col_letter}{sum_cash_pos_row}<0,ABS({col_letter}{sum_cash_pos_row})*{rate_ref}*{days_formula}/365,0)",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = INTEREST_FILL
        cell.border = THIN_BORDER
    # Pre-TC interest: days = pre_tc_date − last_week_date (≈ 358)
    pre_tc_interest = ws.cell(
        row=sum_interest_cost_row, column=pre_tc_col,
        value=f"=IF({pre_tc_col_letter}{sum_cash_pos_row}<0,ABS({pre_tc_col_letter}{sum_cash_pos_row})*{rate_ref}*({pre_tc_col_letter}5-{last_week_col_letter}5)/365,0)",
    )
    pre_tc_interest.number_format = CURRENCY_FORMAT_TOTAL
    pre_tc_interest.font   = Font(bold=True)
    pre_tc_interest.fill   = INTEREST_FILL
    pre_tc_interest.border = THIN_BORDER
    # TC interest: days = tc_date − pre_tc_date (= 7)
    tc_interest = ws.cell(
        row=sum_interest_cost_row, column=tax_credit_col,
        value=f"=IF({tc_col_letter}{sum_cash_pos_row}<0,ABS({tc_col_letter}{sum_cash_pos_row})*{rate_ref}*({tc_col_letter}5-{pre_tc_col_letter}5)/365,0)",
    )
    tc_interest.number_format = CURRENCY_FORMAT_TOTAL
    tc_interest.font   = Font(bold=True)
    tc_interest.fill   = INTEREST_FILL
    tc_interest.border = THIN_BORDER
    interest_grand = ws.cell(
        row=sum_interest_cost_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_interest_cost_row}:{tc_col_letter}{sum_interest_cost_row})",
    )
    interest_grand.number_format = CURRENCY_FORMAT_TOTAL
    interest_grand.font   = Font(bold=True)
    interest_grand.fill   = INTEREST_FILL
    interest_grand.border = THIN_BORDER

    # Annual Interest Rate — linked from Cashflow sheet (one source of truth)
    ws.cell(row=sum_interest_rate_row, column=DESC_COL, value="Annual Interest Rate").font = Font(
        bold=True, size=10, italic=True
    )
    ws.cell(row=sum_interest_rate_row, column=DESC_COL).border = THIN_BORDER
    rate_link = ws.cell(
        row=sum_interest_rate_row, column=TOTAL_COL,
        value=f"={DETAIL}!$C${detail_interest_rate_row}",
    )
    rate_link.number_format = "0.00%"
    rate_link.font   = Font(bold=True)
    rate_link.border = THIN_BORDER

    # ── Financing cost summary (2 rows below interest rate) ──────────────────
    sum_fin_interest_row = sum_interest_rate_row + 2
    sum_fin_setup_row    = sum_fin_interest_row + 1
    sum_fin_legal_row    = sum_fin_setup_row + 1
    sum_fin_total_row    = sum_fin_legal_row + 1
    cash_pos_range = f"{first_data_col_letter}{sum_cash_pos_row}:{last_data_col_letter}{sum_cash_pos_row}"

    ws.cell(row=sum_fin_interest_row, column=DESC_COL, value="Interest Cost").font = Font(bold=True)
    ws.cell(row=sum_fin_interest_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_interest_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(row=sum_fin_interest_row, column=TOTAL_COL, value=f"=C{sum_interest_cost_row}")
    c.number_format = CURRENCY_FORMAT_TOTAL
    c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_setup_row, column=DESC_COL, value="Setup Fee (1.5% of peak loan)").font = Font(bold=True)
    ws.cell(row=sum_fin_setup_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_setup_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(
        row=sum_fin_setup_row, column=TOTAL_COL,
        value=f"=IF(MIN({cash_pos_range})<0,-MIN({cash_pos_range})*0.015,0)",
    )
    c.number_format = CURRENCY_FORMAT_TOTAL
    c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_legal_row, column=DESC_COL, value="Legal Cost").font = Font(bold=True)
    ws.cell(row=sum_fin_legal_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_legal_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(
        row=sum_fin_legal_row, column=TOTAL_COL,
        value=f"={DETAIL}!C{detail_fin_legal_row}",   # linked so Legal Cost stays in sync
    )
    c.number_format = CURRENCY_FORMAT_TOTAL
    c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_total_row, column=DESC_COL, value="TOTAL FINANCING COST").font = Font(bold=True, size=11)
    ws.cell(row=sum_fin_total_row, column=DESC_COL).fill   = FINANCING_TOTAL_FILL
    ws.cell(row=sum_fin_total_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(
        row=sum_fin_total_row, column=TOTAL_COL,
        value=f"=SUM(C{sum_fin_interest_row}:C{sum_fin_legal_row})",
    )
    c.number_format = CURRENCY_FORMAT_TOTAL
    c.font = Font(bold=True); c.fill = FINANCING_TOTAL_FILL; c.border = THIN_BORDER

    # Outflow components + hard costs — row offsets mirror the main Cashflow sheet
    detail_fin_total_row         = detail_fin_legal_row + 1
    detail_internals_outflow_row = detail_fin_total_row + 2
    detail_financing_outflow_row = detail_fin_total_row + 3
    detail_hard_costs_row        = detail_fin_total_row + 4
    sum_internals_outflow_row    = sum_fin_total_row + 2
    sum_financing_outflow_row    = sum_fin_total_row + 3
    sum_hard_costs_row           = sum_fin_total_row + 4

    def _link_row(label: str, sum_row: int, detail_row: int, bold_size: int = 11) -> None:
        ws.cell(row=sum_row, column=DESC_COL, value=label).font = Font(bold=True, size=bold_size)
        ws.cell(row=sum_row, column=DESC_COL).border = THIN_BORDER
        tc = ws.cell(row=sum_row, column=TOTAL_COL, value=f"={DETAIL}!C{detail_row}")
        tc.number_format = CURRENCY_FORMAT_TOTAL
        tc.font = Font(bold=True)
        tc.border = THIN_BORDER
        for i in range(num_weeks):
            col = FIRST_WEEK_COL + i
            col_letter = get_column_letter(col)
            c = ws.cell(row=sum_row, column=col, value=f"={DETAIL}!{col_letter}{detail_row}")
            c.number_format = CURRENCY_FORMAT_TOTAL
            c.font = Font(bold=True)
            c.border = THIN_BORDER

    _link_row("INTERNALS OUTFLOW", sum_internals_outflow_row, detail_internals_outflow_row)
    _link_row("FINANCING OUTFLOW", sum_financing_outflow_row, detail_financing_outflow_row)
    _link_row("HARD COSTS OUTFLOW", sum_hard_costs_row, detail_hard_costs_row)

    _apply_requested_cashflow_formatting(
        ws,
        max_col=tax_credit_col,
        totals_rows={sum_totals_row, sum_cum_row, sum_inflow_total_row, sum_inflow_cum_row, sum_fin_total_row},
        outflow_min_row=5,
        outflow_max_row=sum_cum_row,
        inflow_min_row=sum_inflow_hdr_row,
        inflow_max_row=sum_inflow_cum_row,
        cash_pos_row=sum_cash_pos_row,
        interest_cost_row=sum_interest_cost_row,
        financing_min_row=sum_fin_interest_row,
        financing_max_row=sum_fin_total_row,
        interest_rate_row=sum_interest_rate_row,
        keep_paycycle_colors=has_payroll,
    )

    # Apply fills and borders after formatting (formatting resets all fills)
    for component_row in (sum_internals_outflow_row, sum_financing_outflow_row):
        for col in range(CODE_COL, last_week_col + 1):
            ws.cell(row=component_row, column=col).fill = SUBTLE_TOTAL_FILL
        _apply_outside_border(ws, component_row, component_row, CODE_COL, last_week_col)
    for col in range(CODE_COL, last_week_col + 1):
        ws.cell(row=sum_hard_costs_row, column=col).fill = HARD_COSTS_FILL
    _apply_outside_border(ws, sum_hard_costs_row, sum_hard_costs_row, CODE_COL, last_week_col)

    # ── Column widths and freeze panes ───────────────────────────────────────
    ws.column_dimensions[get_column_letter(CODE_COL)].width  = 10
    ws.column_dimensions[get_column_letter(DESC_COL)].width  = 35
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 15
    for i in range(num_weeks):
        ws.column_dimensions[get_column_letter(FIRST_WEEK_COL + i)].width = 12
    ws.column_dimensions[get_column_letter(pre_tc_col)].width = 14
    ws.column_dimensions[get_column_letter(tax_credit_col)].width = 14
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=FIRST_WEEK_COL)


def _write_summary_sheet(wb: Workbook, output: CashflowOutput, params: ProductionParameters):
    """Write the Summary sheet with phase and monthly totals."""
    ws = wb.create_sheet("Summary")

    ws.cell(row=1, column=1, value="Cashflow Summary").font = TITLE_FONT

    # Phase totals
    ws.cell(row=3, column=1, value="Phase Totals").font = HEADER_FONT
    ws.cell(row=4, column=1, value="Phase").font = Font(bold=True)
    ws.cell(row=4, column=2, value="Total").font = Font(bold=True)

    phase_totals: dict[str, float] = {}
    for i, week in enumerate(output.weeks):
        label = week.phase_label
        # Simplify label for grouping
        if "SHOOT" in label.upper():
            group = "PRODUCTION"
        elif "WRAP" in label.upper():
            group = "WRAP"
        elif "POST" in label.upper():
            group = "POST-PRODUCTION"
        elif "PREP" in label.upper():
            group = "PRE-PRODUCTION"
        elif "HIATUS" in label.upper():
            group = "HIATUS"
        else:
            group = label
        phase_totals[group] = phase_totals.get(group, 0) + output.weekly_totals[i]

    row = 5
    for phase, total in phase_totals.items():
        ws.cell(row=row, column=1, value=phase)
        cell = ws.cell(row=row, column=2, value=round(total, 2))
        cell.number_format = CURRENCY_FORMAT
        row += 1

    # Monthly totals
    row += 2
    ws.cell(row=row, column=1, value="Monthly Totals").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Month").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Total").font = Font(bold=True)
    row += 1

    monthly: dict[str, float] = {}
    for i, week in enumerate(output.weeks):
        month_key = week.week_commencing.strftime("%Y-%m")
        monthly[month_key] = monthly.get(month_key, 0) + output.weekly_totals[i]

    for month, total in monthly.items():
        ws.cell(row=row, column=1, value=month)
        cell = ws.cell(row=row, column=2, value=round(total, 2))
        cell.number_format = CURRENCY_FORMAT
        row += 1

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18


def _write_parameters_sheet(wb: Workbook, params: ProductionParameters) -> None:
    """Write the Parameters sheet documenting the production schedule."""
    ws = wb.create_sheet("Parameters")

    ws.cell(row=1, column=1, value="Production Parameters").font = TITLE_FONT

    row = 3
    fields = [
        ("Title", params.title),
        ("Series Number", getattr(params, "series_number", None) or "N/A"),
        ("Episode Count", params.episode_count),
        ("Prep Start", params.prep_start.isoformat()),
        ("PP Start", params.pp_start.isoformat()),
        ("PP End", params.pp_end.isoformat()),
        ("Edit Start", params.edit_start.isoformat()),
        ("Final Delivery", params.final_delivery_date.isoformat()),
        ("First Payroll Week", params.first_payroll_week.isoformat() if params.first_payroll_week else "N/A"),
    ]

    for label, value in fields:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    # Shooting blocks
    row += 1
    ws.cell(row=row, column=1, value="Shooting Blocks").font = HEADER_FONT
    row += 1
    ws.cell(row=row, column=1, value="Block").font = Font(bold=True)
    ws.cell(row=row, column=2, value="Type").font = Font(bold=True)
    ws.cell(row=row, column=3, value="Episodes").font = Font(bold=True)
    ws.cell(row=row, column=4, value="Start").font = Font(bold=True)
    ws.cell(row=row, column=5, value="End").font = Font(bold=True)
    row += 1

    for block in params.shooting_blocks:
        ws.cell(row=row, column=1, value=block.block_number)
        ws.cell(row=row, column=2, value=block.block_type or "Shoot")
        ws.cell(row=row, column=3, value=block.shoot_start.isoformat())
        ws.cell(row=row, column=4, value=block.shoot_end.isoformat())
        row += 1

    # Episode deliveries
    row += 1
    ws.cell(row=row, column=1, value="Episode Deliveries").font = HEADER_FONT
    row += 1
    headers = ["Episode", "Rough Cut", "Picture Lock", "Online", "Mix", "Delivery"]
    for i, h in enumerate(headers):
        ws.cell(row=row, column=i + 1, value=h).font = Font(bold=True)
    row += 1

    for ep in params.episode_deliveries:
        ws.cell(row=row, column=1, value=ep.episode_number)
        ws.cell(row=row, column=2, value=ep.rough_cut_date.isoformat() if ep.rough_cut_date else "")
        ws.cell(row=row, column=3, value=ep.picture_lock_date.isoformat() if ep.picture_lock_date else "")
        ws.cell(row=row, column=4, value=ep.online_date.isoformat() if ep.online_date else "")
        ws.cell(row=row, column=5, value=ep.mix_date.isoformat() if ep.mix_date else "")
        ws.cell(row=row, column=6, value=ep.delivery_date.isoformat())
        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 15
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 15


def _write_vertical_cf_sheet(wb: Workbook, output: CashflowOutput) -> None:
    """Write the 'Vertical CF' sheet — one row per week, formula-linked to Cashflow."""
    DETAIL = "Cashflow"
    ws = wb.create_sheet("Vertical CF")

    num_weeks = len(output.weeks)
    # Weekly totals row in the Cashflow sheet (mirrors _write_main_sheet)
    totals_row = DATA_START_ROW + len(output.rows)

    BLACK_FILL = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    WHITE_BOLD = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")

    # Dollar-sign accounting format: "$ 22,928" / "$ (22,928)" / "$ -"
    DOLLAR_FMT = '_($* #,##0_);_($* (#,##0);_($* "-"??_);_(@_)'

    # ── Rows 1-2: title / subtitle ────────────────────────────────────────────
    ws.cell(row=1, column=1, value=output.title).font = TITLE_FONT
    ws.cell(row=2, column=1, value="PRODUCTION CASHFLOW").font = Font(
        name="Calibri", size=11, bold=True
    )

    # ── Row 4: "Tax credits:" label on black, spanning Ontario + Federal cols ─
    for col in range(1, 7):
        ws.cell(row=4, column=col).fill = BLACK_FILL
    tax_cell = ws.cell(row=4, column=4, value="Tax credits:")
    tax_cell.font = WHITE_BOLD
    tax_cell.fill = BLACK_FILL
    tax_cell.alignment = CENTER
    ws.merge_cells(start_row=4, start_column=4, end_row=4, end_column=5)

    # ── Row 5: column headers on black ───────────────────────────────────────
    headers = [
        ("Week", RIGHT),
        ("Date", CENTER),
        ("Weekly Cash Out", CENTER),
        ("Ontario", CENTER),
        ("Federal", CENTER),
        ("Distribution\nAdvance", CENTER),
    ]
    for col, (text, align) in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=text)
        cell.font = WHITE_BOLD
        cell.fill = BLACK_FILL
        cell.alignment = Alignment(
            horizontal=align.horizontal, vertical="center", wrap_text=True
        )
    ws.row_dimensions[4].height = 18
    ws.row_dimensions[5].height = 28

    # ── Data rows: one per week ───────────────────────────────────────────────
    for i in range(num_weeks):
        excel_row = 6 + i
        col_letter = get_column_letter(FIRST_WEEK_COL + i)

        # Week number
        cell = ws.cell(row=excel_row, column=1, value=i + 1)
        cell.alignment = RIGHT

        # Date — linked from Cashflow row 5
        cell = ws.cell(row=excel_row, column=2, value=f"={DETAIL}!{col_letter}5")
        cell.number_format = "DD-MMM-YY"
        cell.alignment = CENTER

        # Weekly Cash Out — linked from Cashflow weekly totals row
        cell = ws.cell(
            row=excel_row, column=3,
            value=f"={DETAIL}!{col_letter}{totals_row}",
        )
        cell.number_format = DOLLAR_FMT

        # Ontario tax credit — always blank (zero not displayed)
        ws.cell(row=excel_row, column=4)

        # Federal tax credit — always blank
        ws.cell(row=excel_row, column=5)

        # Distribution Advance = Tax Credits − Weekly Cash Out = 0 − C
        cell = ws.cell(
            row=excel_row, column=6,
            value=f"=D{excel_row}+E{excel_row}-C{excel_row}",
        )
        cell.number_format = DOLLAR_FMT

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22


def _write_monthly_cf_sheet(wb: Workbook, output: CashflowOutput, params: ProductionParameters) -> None:
    """Write the 'Monthly CF' sheet.

    Aggregates weekly data from 'Summary CF' into calendar-month columns using
    SUM formulas.  Row structure mirrors Summary CF; column structure replaces
    individual week columns with one column per calendar month.
    Pre-TC and TC event columns are preserved at the end, linked directly from
    'Summary CF'.
    """
    SUMMARY = "'Summary CF'"
    ws = wb.create_sheet("Monthly CF")

    num_weeks = len(output.weeks)

    # ── Build month groups (ordered, contiguous) ──────────────────────────────
    # month_order: list of (year, month) in sequence
    # month_sc_cols: (year, month) -> (first_sc_col, last_sc_col) in Summary CF
    month_order: list[tuple[int, int]] = []
    month_sc_cols: dict[tuple[int, int], tuple[int, int]] = {}

    for i, week in enumerate(output.weeks):
        wc = week.week_commencing
        key = (wc.year, wc.month)
        sc_col = FIRST_WEEK_COL + i
        if key not in month_sc_cols:
            month_sc_cols[key] = (sc_col, sc_col)
            month_order.append(key)
        else:
            month_sc_cols[key] = (month_sc_cols[key][0], sc_col)

    num_months = len(month_order)
    last_month_col = FIRST_WEEK_COL + num_months - 1
    last_month_col_letter = get_column_letter(last_month_col)
    first_data_col_letter = get_column_letter(FIRST_WEEK_COL)

    # Pre-TC and TC columns on Monthly CF (immediately after the month columns)
    pre_tc_col = FIRST_WEEK_COL + num_months
    pre_tc_col_letter = get_column_letter(pre_tc_col)
    tax_credit_col = FIRST_WEEK_COL + num_months + 1
    tc_col_letter = get_column_letter(tax_credit_col)

    # Corresponding columns in Summary CF
    sc_pre_tc_letter = get_column_letter(FIRST_WEEK_COL + num_weeks)
    sc_tc_letter = get_column_letter(FIRST_WEEK_COL + num_weeks + 1)

    # Row positions in Summary CF (kept in sync with _write_summary_cf_sheet)
    sc_totals_row        = DATA_START_ROW + len(SUMMARY_ACCOUNTS)
    sc_cum_row           = sc_totals_row + 1
    sc_inflow_hdr_row    = sc_cum_row + 3
    sc_inflow_data_start = sc_inflow_hdr_row + 1
    sc_inflow_total_row  = sc_inflow_data_start + len(output.cash_inflows)
    sc_inflow_cum_row    = sc_inflow_total_row + 1
    sc_cash_pos_row      = sc_inflow_cum_row + 2
    sc_interest_cost_row = sc_cash_pos_row + 2
    sc_interest_rate_row = sc_interest_cost_row + 1
    sc_fin_interest_row  = sc_interest_rate_row + 2
    sc_fin_legal_row     = sc_fin_interest_row + 2

    # Helper: formula that sums the weeks of one month from Summary CF for a given row
    def sc_sum(year: int, month: int, row: int) -> str:
        first_col, last_col = month_sc_cols[(year, month)]
        first_letter = get_column_letter(first_col)
        last_letter = get_column_letter(last_col)
        if first_col == last_col:
            return f"={SUMMARY}!{first_letter}{row}"
        return f"=SUM({SUMMARY}!{first_letter}{row}:{last_letter}{row})"

    # ── Rows 1-2: title / metadata ────────────────────────────────────────────
    ws.cell(row=1, column=1, value=f"{output.title} - Monthly Cashflow Forecast").font = TITLE_FONT
    series_number = getattr(params, "series_number", None)
    series_text = f"Series {series_number}" if series_number else ""
    ep_text = f"{params.episode_count} Episodes"
    meta = " | ".join(filter(None, [series_text, ep_text]))
    ws.cell(row=2, column=1, value=meta).font = SUBTITLE_FONT

    # ── Row 4: dominant phase label per month ─────────────────────────────────
    for m_idx, (year, month) in enumerate(month_order):
        col = FIRST_WEEK_COL + m_idx
        first_sc, last_sc = month_sc_cols[(year, month)]
        phase_labels = [
            output.weeks[i - FIRST_WEEK_COL].phase_label
            for i in range(first_sc, last_sc + 1)
        ]
        dominant = Counter(phase_labels).most_common(1)[0][0]
        cell = ws.cell(row=4, column=col, value=dominant)
        cell.font = Font(bold=True, size=8)
        cell.fill = _get_phase_fill(dominant)
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
        ec = ws.cell(row=4, column=extra_col, value=f"={SUMMARY}!{sc_letter}4")
        ec.font = Font(bold=True, size=8)
        ec.alignment = Alignment(horizontal="center")
        ec.border = THIN_BORDER

    # ── Row 5: month labels / column headers ──────────────────────────────────
    ws.cell(row=5, column=CODE_COL,  value="Code").font        = HEADER_FONT
    ws.cell(row=5, column=DESC_COL,  value="Description").font = HEADER_FONT
    ws.cell(row=5, column=TOTAL_COL, value="Total").font       = HEADER_FONT
    ws.cell(row=5, column=CODE_COL).border  = THIN_BORDER
    ws.cell(row=5, column=DESC_COL).border  = THIN_BORDER
    ws.cell(row=5, column=TOTAL_COL).border = THIN_BORDER
    for m_idx, (year, month) in enumerate(month_order):
        col = FIRST_WEEK_COL + m_idx
        cell = ws.cell(row=5, column=col, value=date(year, month, 1))
        cell.font = Font(bold=True, size=9)
        cell.number_format = "MMM-YYYY"
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER
    for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
        ec = ws.cell(row=5, column=extra_col, value=f"={SUMMARY}!{sc_letter}5")
        ec.font = Font(bold=True, size=8)
        ec.number_format = "DD-MMM-YYYY"
        ec.alignment = Alignment(horizontal="center")
        ec.border = THIN_BORDER

    # ── Summary account rows ──────────────────────────────────────────────────
    for row_idx, (code, description) in enumerate(SUMMARY_ACCOUNTS):
        excel_row = DATA_START_ROW + row_idx
        ws.cell(row=excel_row, column=CODE_COL,  value=code).border        = THIN_BORDER
        ws.cell(row=excel_row, column=DESC_COL,  value=description).border = THIN_BORDER
        total_cell = ws.cell(row=excel_row, column=TOTAL_COL, value=f"={SUMMARY}!C{excel_row}")
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font   = Font(bold=True)
        total_cell.border = THIN_BORDER
        for m_idx, (year, month) in enumerate(month_order):
            col = FIRST_WEEK_COL + m_idx
            cell = ws.cell(row=excel_row, column=col, value=sc_sum(year, month, excel_row))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
        for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
            ec = ws.cell(row=excel_row, column=extra_col, value=f"={SUMMARY}!{sc_letter}{excel_row}")
            ec.number_format = CURRENCY_FORMAT
            ec.border = THIN_BORDER

    # ── Monthly totals row ────────────────────────────────────────────────────
    sum_totals_row = DATA_START_ROW + len(SUMMARY_ACCOUNTS)
    ws.cell(row=sum_totals_row, column=DESC_COL, value="MONTHLY TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_totals_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        col = FIRST_WEEK_COL + m_idx
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_totals_row, column=col,
            value=f"=SUM({col_letter}{DATA_START_ROW}:{col_letter}{sum_totals_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font  = Font(bold=True)
        cell.fill  = TOTAL_FILL
        cell.border = THIN_BORDER
    grand_total = ws.cell(
        row=sum_totals_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_totals_row}:{last_month_col_letter}{sum_totals_row})",
    )
    grand_total.number_format = CURRENCY_FORMAT_TOTAL
    grand_total.font  = Font(bold=True)
    grand_total.fill  = TOTAL_FILL
    grand_total.border = THIN_BORDER
    for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
        ec = ws.cell(row=sum_totals_row, column=extra_col, value=f"={SUMMARY}!{sc_letter}{sc_totals_row}")
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font  = Font(bold=True)
        ec.fill  = TOTAL_FILL
        ec.border = THIN_BORDER

    # ── Cumulative totals row (running sum of monthly totals) ─────────────────
    sum_cum_row = sum_totals_row + 1
    ws.cell(row=sum_cum_row, column=DESC_COL, value="CUMULATIVE TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_cum_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        col = FIRST_WEEK_COL + m_idx
        col_letter = get_column_letter(col)
        if m_idx == 0:
            cell = ws.cell(row=sum_cum_row, column=col, value=f"={col_letter}{sum_totals_row}")
        else:
            prev = get_column_letter(col - 1)
            cell = ws.cell(
                row=sum_cum_row, column=col,
                value=f"={prev}{sum_cum_row}+{col_letter}{sum_totals_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font  = Font(bold=True, italic=True)
        cell.border = THIN_BORDER
    # Pre-TC / TC carry the last month's cumulative outflow (no new outflows)
    for extra_col in (pre_tc_col, tax_credit_col):
        ec = ws.cell(row=sum_cum_row, column=extra_col, value=f"={last_month_col_letter}{sum_cum_row}")
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True, italic=True)
        ec.border = THIN_BORDER

    # ── Cash Inflows section ──────────────────────────────────────────────────
    sum_inflow_hdr_row    = sum_cum_row + 3
    sum_inflow_data_start = sum_inflow_hdr_row + 1
    sum_inflow_total_row  = sum_inflow_data_start + len(output.cash_inflows)
    sum_inflow_cum_row    = sum_inflow_total_row + 1

    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL, value="CASH INFLOWS").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL).fill   = INFLOW_HEADER_FILL
    ws.cell(row=sum_inflow_hdr_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        cell = ws.cell(row=sum_inflow_hdr_row, column=FIRST_WEEK_COL + m_idx)
        cell.fill   = INFLOW_HEADER_FILL
        cell.border = THIN_BORDER
    for extra_col in (pre_tc_col, tax_credit_col):
        c = ws.cell(row=sum_inflow_hdr_row, column=extra_col)
        c.fill   = INFLOW_HEADER_FILL
        c.border = THIN_BORDER

    for row_idx, inflow_row in enumerate(output.cash_inflows):
        excel_row    = sum_inflow_data_start + row_idx
        sc_inflow_row = sc_inflow_data_start + row_idx
        ws.cell(row=excel_row, column=DESC_COL, value=inflow_row.label).border = THIN_BORDER
        total_cell = ws.cell(row=excel_row, column=TOTAL_COL, value=f"={SUMMARY}!C{sc_inflow_row}")
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font   = Font(bold=True)
        total_cell.border = THIN_BORDER
        for m_idx, (year, month) in enumerate(month_order):
            col = FIRST_WEEK_COL + m_idx
            cell = ws.cell(row=excel_row, column=col, value=sc_sum(year, month, sc_inflow_row))
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
        for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
            ec = ws.cell(row=excel_row, column=extra_col, value=f"={SUMMARY}!{sc_letter}{sc_inflow_row}")
            ec.number_format = CURRENCY_FORMAT
            ec.border = THIN_BORDER

    # Monthly inflow total
    ws.cell(row=sum_inflow_total_row, column=DESC_COL, value="MONTHLY INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_total_row, column=DESC_COL).fill   = INFLOW_TOTAL_FILL
    ws.cell(row=sum_inflow_total_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        col = FIRST_WEEK_COL + m_idx
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_inflow_total_row, column=col,
            value=f"=SUM({col_letter}{sum_inflow_data_start}:{col_letter}{sum_inflow_total_row - 1})",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = INFLOW_TOTAL_FILL
        cell.border = THIN_BORDER
    for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
        extra_letter = get_column_letter(extra_col)
        ec = ws.cell(
            row=sum_inflow_total_row, column=extra_col,
            value=f"=SUM({extra_letter}{sum_inflow_data_start}:{extra_letter}{sum_inflow_total_row - 1})",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True)
        ec.fill   = INFLOW_TOTAL_FILL
        ec.border = THIN_BORDER
    inflow_grand = ws.cell(
        row=sum_inflow_total_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_inflow_total_row}:{tc_col_letter}{sum_inflow_total_row})",
    )
    inflow_grand.number_format = CURRENCY_FORMAT_TOTAL
    inflow_grand.font   = Font(bold=True)
    inflow_grand.fill   = INFLOW_TOTAL_FILL
    inflow_grand.border = THIN_BORDER

    # Cumulative inflow total (running sum)
    ws.cell(row=sum_inflow_cum_row, column=DESC_COL, value="CUMULATIVE INFLOW TOTAL").font = Font(bold=True, size=11)
    ws.cell(row=sum_inflow_cum_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        col = FIRST_WEEK_COL + m_idx
        col_letter = get_column_letter(col)
        if m_idx == 0:
            cell = ws.cell(row=sum_inflow_cum_row, column=col, value=f"={col_letter}{sum_inflow_total_row}")
        else:
            prev = get_column_letter(col - 1)
            cell = ws.cell(
                row=sum_inflow_cum_row, column=col,
                value=f"={prev}{sum_inflow_cum_row}+{col_letter}{sum_inflow_total_row}",
            )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True, italic=True)
        cell.border = THIN_BORDER
    # Pre-TC cumulative inflow = last month cum + pre_tc monthly inflow
    pre_tc_cum = ws.cell(
        row=sum_inflow_cum_row, column=pre_tc_col,
        value=f"={last_month_col_letter}{sum_inflow_cum_row}+{pre_tc_col_letter}{sum_inflow_total_row}",
    )
    pre_tc_cum.number_format = CURRENCY_FORMAT_TOTAL
    pre_tc_cum.font   = Font(bold=True, italic=True)
    pre_tc_cum.border = THIN_BORDER
    # TC cumulative inflow = pre_tc cum + tc monthly inflow
    tc_cum = ws.cell(
        row=sum_inflow_cum_row, column=tax_credit_col,
        value=f"={pre_tc_col_letter}{sum_inflow_cum_row}+{tc_col_letter}{sum_inflow_total_row}",
    )
    tc_cum.number_format = CURRENCY_FORMAT_TOTAL
    tc_cum.font   = Font(bold=True, italic=True)
    tc_cum.border = THIN_BORDER

    # ── Cumulative cash position ───────────────────────────────────────────────
    sum_cash_pos_row = sum_inflow_cum_row + 2
    ws.cell(row=sum_cash_pos_row, column=DESC_COL, value="CUMULATIVE CASH POSITION").font = Font(bold=True, size=11)
    ws.cell(row=sum_cash_pos_row, column=DESC_COL).fill   = CASH_POS_FILL
    ws.cell(row=sum_cash_pos_row, column=DESC_COL).border = THIN_BORDER
    for m_idx in range(num_months):
        col = FIRST_WEEK_COL + m_idx
        col_letter = get_column_letter(col)
        cell = ws.cell(
            row=sum_cash_pos_row, column=col,
            value=f"={col_letter}{sum_inflow_cum_row}-{col_letter}{sum_cum_row}",
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = CASH_POS_FILL
        cell.border = THIN_BORDER
    for extra_col, extra_letter in ((pre_tc_col, pre_tc_col_letter), (tax_credit_col, tc_col_letter)):
        ec = ws.cell(
            row=sum_cash_pos_row, column=extra_col,
            value=f"={extra_letter}{sum_inflow_cum_row}-{last_month_col_letter}{sum_cum_row}",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True)
        ec.fill   = CASH_POS_FILL
        ec.border = THIN_BORDER

    # ── Interest cost (sum of weekly interest from Summary CF) ────────────────
    sum_interest_cost_row = sum_cash_pos_row + 2
    sum_interest_rate_row = sum_interest_cost_row + 1

    ws.cell(row=sum_interest_cost_row, column=DESC_COL, value="INTEREST COST").font = Font(bold=True, size=11)
    ws.cell(row=sum_interest_cost_row, column=DESC_COL).fill   = INTEREST_FILL
    ws.cell(row=sum_interest_cost_row, column=DESC_COL).border = THIN_BORDER
    for m_idx, (year, month) in enumerate(month_order):
        col = FIRST_WEEK_COL + m_idx
        cell = ws.cell(
            row=sum_interest_cost_row, column=col,
            value=sc_sum(year, month, sc_interest_cost_row),
        )
        cell.number_format = CURRENCY_FORMAT_TOTAL
        cell.font   = Font(bold=True)
        cell.fill   = INTEREST_FILL
        cell.border = THIN_BORDER
    for extra_col, sc_letter in ((pre_tc_col, sc_pre_tc_letter), (tax_credit_col, sc_tc_letter)):
        ec = ws.cell(
            row=sum_interest_cost_row, column=extra_col,
            value=f"={SUMMARY}!{sc_letter}{sc_interest_cost_row}",
        )
        ec.number_format = CURRENCY_FORMAT_TOTAL
        ec.font   = Font(bold=True)
        ec.fill   = INTEREST_FILL
        ec.border = THIN_BORDER
    interest_grand = ws.cell(
        row=sum_interest_cost_row, column=TOTAL_COL,
        value=f"=SUM({first_data_col_letter}{sum_interest_cost_row}:{tc_col_letter}{sum_interest_cost_row})",
    )
    interest_grand.number_format = CURRENCY_FORMAT_TOTAL
    interest_grand.font   = Font(bold=True)
    interest_grand.fill   = INTEREST_FILL
    interest_grand.border = THIN_BORDER

    ws.cell(row=sum_interest_rate_row, column=DESC_COL, value="Annual Interest Rate").font = Font(
        bold=True, size=10, italic=True,
    )
    ws.cell(row=sum_interest_rate_row, column=DESC_COL).border = THIN_BORDER
    rate_link = ws.cell(
        row=sum_interest_rate_row, column=TOTAL_COL,
        value=f"={SUMMARY}!$C${sc_interest_rate_row}",
    )
    rate_link.number_format = "0.00%"
    rate_link.font   = Font(bold=True)
    rate_link.border = THIN_BORDER

    # ── Financing cost summary (totals linked from Summary CF) ────────────────
    sum_fin_interest_row = sum_interest_rate_row + 2
    sum_fin_setup_row    = sum_fin_interest_row + 1
    sum_fin_legal_row    = sum_fin_setup_row + 1
    sum_fin_total_row    = sum_fin_legal_row + 1
    cash_pos_range = f"{first_data_col_letter}{sum_cash_pos_row}:{last_month_col_letter}{sum_cash_pos_row}"

    ws.cell(row=sum_fin_interest_row, column=DESC_COL, value="Interest Cost").font = Font(bold=True)
    ws.cell(row=sum_fin_interest_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_interest_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(row=sum_fin_interest_row, column=TOTAL_COL, value=f"=C{sum_interest_cost_row}")
    c.number_format = CURRENCY_FORMAT_TOTAL; c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_setup_row, column=DESC_COL, value="Setup Fee (1.5% of peak loan)").font = Font(bold=True)
    ws.cell(row=sum_fin_setup_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_setup_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(
        row=sum_fin_setup_row, column=TOTAL_COL,
        value=f"=IF(MIN({cash_pos_range})<0,-MIN({cash_pos_range})*0.015,0)",
    )
    c.number_format = CURRENCY_FORMAT_TOTAL; c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_legal_row, column=DESC_COL, value="Legal Cost").font = Font(bold=True)
    ws.cell(row=sum_fin_legal_row, column=DESC_COL).fill   = FINANCING_FILL
    ws.cell(row=sum_fin_legal_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(row=sum_fin_legal_row, column=TOTAL_COL, value=f"={SUMMARY}!C{sc_fin_legal_row}")
    c.number_format = CURRENCY_FORMAT_TOTAL; c.font = Font(bold=True); c.fill = FINANCING_FILL; c.border = THIN_BORDER

    ws.cell(row=sum_fin_total_row, column=DESC_COL, value="TOTAL FINANCING COST").font = Font(bold=True, size=11)
    ws.cell(row=sum_fin_total_row, column=DESC_COL).fill   = FINANCING_TOTAL_FILL
    ws.cell(row=sum_fin_total_row, column=DESC_COL).border = THIN_BORDER
    c = ws.cell(
        row=sum_fin_total_row, column=TOTAL_COL,
        value=f"=SUM(C{sum_fin_interest_row}:C{sum_fin_legal_row})",
    )
    c.number_format = CURRENCY_FORMAT_TOTAL; c.font = Font(bold=True); c.fill = FINANCING_TOTAL_FILL; c.border = THIN_BORDER

    _apply_requested_cashflow_formatting(
        ws,
        max_col=tax_credit_col,
        totals_rows={sum_totals_row, sum_cum_row, sum_inflow_total_row, sum_inflow_cum_row, sum_fin_total_row},
        outflow_min_row=5,
        outflow_max_row=sum_cum_row,
        inflow_min_row=sum_inflow_hdr_row,
        inflow_max_row=sum_inflow_cum_row,
        cash_pos_row=sum_cash_pos_row,
        interest_cost_row=sum_interest_cost_row,
        financing_min_row=sum_fin_interest_row,
        financing_max_row=sum_fin_total_row,
        interest_rate_row=sum_interest_rate_row,
        keep_paycycle_colors=False,
    )

    # ── Column widths and freeze panes ────────────────────────────────────────
    ws.column_dimensions[get_column_letter(CODE_COL)].width  = 10
    ws.column_dimensions[get_column_letter(DESC_COL)].width  = 35
    ws.column_dimensions[get_column_letter(TOTAL_COL)].width = 15
    for m_idx in range(num_months):
        ws.column_dimensions[get_column_letter(FIRST_WEEK_COL + m_idx)].width = 14
    ws.column_dimensions[get_column_letter(pre_tc_col)].width  = 14
    ws.column_dimensions[get_column_letter(tax_credit_col)].width = 14
    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=FIRST_WEEK_COL)


def write_cashflow_excel(output: CashflowOutput, params: ProductionParameters, budget: ParsedBudget | None = None) -> BytesIO:
    """Generate a complete cashflow Excel workbook.

    Returns a BytesIO buffer containing the .xlsx file.
    """
    wb = Workbook()

    _write_main_sheet(wb, output, params, budget=budget)
    _write_summary_cf_sheet(wb, output, params)
    _write_monthly_cf_sheet(wb, output, params)
    _write_vertical_cf_sheet(wb, output)
    _write_summary_sheet(wb, output, params)
    _write_parameters_sheet(wb, params)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
