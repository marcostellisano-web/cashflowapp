"""Parse production parameters from an uploaded Excel file.

Expected format — a single-sheet workbook with vertical key-value layout:

  Row 2: instruction text ("** please enter information in the YELLOW cells")
  Row 4: Title          | <value>
  Row 5: Episode Count  | <value>
  Row 8: Prep Start     | <date>
  Row 9: PP Start       | <date>
  Row 10: PP End        | <date>
  Row 11: Edit Start    | <date>
  Row 12: Final Delivery| <date>
  Row 13: First Payroll Week | <date>

  Row 15: "Shoot Blocks" section header
  Row 16: Block | Type | Shoot Start | Shoot End
  Rows 17+: data

  Row 29: "Episode Deliveries" section header
  Row 30: Episode | Rough Cut | Picture Lock | Online | Mix | Delivery
  Rows 31+: data

The parser also supports the old 3-sheet format for backwards compatibility.
"""

from datetime import date, datetime
from io import BytesIO
from typing import BinaryIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.models.production import EpisodeDelivery, ProductionParameters, ShootingBlock


# ---------------------------------------------------------------------------
# Date / value parsing helpers
# ---------------------------------------------------------------------------

def _parse_date(val: object) -> date | None:
    """Coerce a cell value to a date, or None."""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%d-%b-%Y", "%Y/%m/%d",
                     "%d-%b", "%b-%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
        raise ValueError(f"Cannot parse date: {val}")
    raise ValueError(f"Unexpected date value type: {type(val)}")


def _require_date(val: object, field_name: str) -> date:
    result = _parse_date(val)
    if result is None:
        raise ValueError(f"Missing required date: {field_name}")
    return result


# ---------------------------------------------------------------------------
# Single-sheet parser (your template format)
# ---------------------------------------------------------------------------

def _parse_single_sheet(ws) -> ProductionParameters:
    """Parse the single-sheet vertical key-value layout."""
    # Build a lookup: scan column A for labels, grab column B values
    kv: dict[str, object] = {}
    for row in ws.iter_rows(min_row=1, max_col=2, values_only=False):
        label_cell = row[0]
        value_cell = row[1] if len(row) > 1 else None
        label = str(label_cell.value or "").strip().lower()
        if label and value_cell is not None:
            kv[label] = value_cell.value

    def _kv_get(*keys: str) -> object:
        """Fuzzy match: find first key where any of the given substrings appear."""
        for stored_key, val in kv.items():
            for k in keys:
                if k in stored_key:
                    return val
        return None

    title = str(_kv_get("title") or "Untitled")
    episode_count = int(_kv_get("episode count", "episode_count", "episodes") or 6)
    prep_start = _require_date(_kv_get("prep start", "prep_start", "prep"), "Prep Start")
    pp_start = _require_date(_kv_get("pp start", "pp_start"), "PP Start")
    pp_end = _require_date(_kv_get("pp end", "pp_end"), "PP End")
    edit_start = _require_date(_kv_get("edit start", "edit_start", "edit"), "Edit Start")
    final_delivery = _require_date(_kv_get("final delivery", "final_delivery", "final"), "Final Delivery")
    first_payroll = _parse_date(_kv_get("payroll", "first payroll"))

    # Find "Shoot Blocks" and "Episode Deliveries" section headers by scanning col A
    blocks_header_row = None
    deliveries_header_row = None

    for row in ws.iter_rows(min_row=1, max_col=1, values_only=False):
        cell = row[0]
        text = str(cell.value or "").strip().lower()
        if "shoot block" in text or "shooting block" in text:
            blocks_header_row = cell.row
        elif "episode deliver" in text or "deliveries" == text:
            deliveries_header_row = cell.row

    # Parse shooting blocks
    blocks: list[ShootingBlock] = []
    if blocks_header_row is not None:
        # Header row is the row after the section title
        header_row = blocks_header_row + 1
        b_headers = [
            str(ws.cell(row=header_row, column=c).value or "").strip().lower()
            for c in range(1, ws.max_column + 1)
        ]

        def _bcol(row_vals: list, name: str) -> object:
            for i, h in enumerate(b_headers):
                if name in h:
                    return row_vals[i] if i < len(row_vals) else None
            return None

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_vals = list(row)
            if not any(row_vals):
                continue
            # Stop if we hit the next section
            first_val = str(row_vals[0] or "").strip().lower()
            if "episode" in first_val and "deliver" in first_val:
                break
            if "deliver" in first_val and not first_val.replace(".", "").isdigit():
                break

            block_num = _bcol(row_vals, "block")
            if block_num is None or not str(block_num).strip():
                continue
            try:
                block_num = int(block_num)
            except (ValueError, TypeError):
                continue

            blocks.append(
                ShootingBlock(
                    block_number=block_num,
                    block_type=str(_bcol(row_vals, "type") or "Shoot"),
                    shoot_start=_require_date(
                        _bcol(row_vals, "shoot start") or _bcol(row_vals, "start"),
                        f"Block {block_num} Shoot Start",
                    ),
                    shoot_end=_require_date(
                        _bcol(row_vals, "shoot end") or _bcol(row_vals, "end"),
                        f"Block {block_num} Shoot End",
                    ),
                )
            )

    # Parse episode deliveries
    deliveries: list[EpisodeDelivery] = []
    if deliveries_header_row is not None:
        header_row = deliveries_header_row + 1
        d_headers = [
            str(ws.cell(row=header_row, column=c).value or "").strip().lower()
            for c in range(1, ws.max_column + 1)
        ]

        def _dcol(row_vals: list, name: str) -> object:
            for i, h in enumerate(d_headers):
                if name in h:
                    return row_vals[i] if i < len(row_vals) else None
            return None

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            row_vals = list(row)
            if not any(row_vals):
                continue
            ep_num = _dcol(row_vals, "episode")
            if ep_num is None or not str(ep_num).strip():
                continue
            try:
                ep_num = int(ep_num)
            except (ValueError, TypeError):
                continue

            deliveries.append(
                EpisodeDelivery(
                    episode_number=ep_num,
                    rough_cut_date=_parse_date(_dcol(row_vals, "rough") or _dcol(row_vals, "cut")),
                    picture_lock_date=_parse_date(
                        _dcol(row_vals, "picture") or _dcol(row_vals, "lock")
                    ),
                    online_date=_parse_date(_dcol(row_vals, "online")),
                    mix_date=_parse_date(_dcol(row_vals, "mix")),
                    delivery_date=_require_date(
                        _dcol(row_vals, "delivery"),
                        f"Ep {ep_num} Delivery",
                    ),
                )
            )

    return ProductionParameters(
        title=title,
        episode_count=episode_count,
        prep_start=prep_start,
        pp_start=pp_start,
        pp_end=pp_end,
        edit_start=edit_start,
        shooting_blocks=blocks,
        episode_deliveries=deliveries,
        final_delivery_date=final_delivery,
        first_payroll_week=first_payroll,
        hiatus_periods=[],
    )


# ---------------------------------------------------------------------------
# Multi-sheet parser (legacy 3-sheet format)
# ---------------------------------------------------------------------------

def _parse_multi_sheet(wb) -> ProductionParameters:
    """Parse the legacy 3-sheet format."""
    # --- Sheet: Info ---
    info_sheet = None
    for name in ("Info", "info", "INFO", "Production Info"):
        if name in wb.sheetnames:
            info_sheet = wb[name]
            break
    if info_sheet is None:
        info_sheet = wb.worksheets[0]

    headers = [str(c.value or "").strip().lower() for c in info_sheet[1]]
    values = [c.value for c in info_sheet[2]]

    def _col(name: str) -> object:
        name_lower = name.lower()
        for i, h in enumerate(headers):
            if name_lower in h:
                return values[i] if i < len(values) else None
        return None

    title = str(_col("title") or "Untitled")
    episode_count = int(_col("episode") or _col("count") or 6)
    prep_start = _require_date(_col("prep"), "Prep Start")
    pp_start = _require_date(_col("pp start") or _col("pp_start"), "PP Start")
    pp_end = _require_date(_col("pp end") or _col("pp_end"), "PP End")
    edit_start = _require_date(_col("edit"), "Edit Start")
    final_delivery = _require_date(_col("final") or _col("delivery"), "Final Delivery")
    first_payroll = _parse_date(_col("payroll"))

    # --- Sheet: Shooting Blocks ---
    blocks_sheet = None
    for name in ("Shooting Blocks", "Blocks", "blocks", "shooting blocks", "SHOOTING BLOCKS"):
        if name in wb.sheetnames:
            blocks_sheet = wb[name]
            break

    blocks: list[ShootingBlock] = []
    if blocks_sheet is not None:
        b_headers = [str(c.value or "").strip().lower() for c in blocks_sheet[1]]

        def _bcol(row_vals: list, name: str) -> object:
            for i, h in enumerate(b_headers):
                if name in h:
                    return row_vals[i] if i < len(row_vals) else None
            return None

        for row in blocks_sheet.iter_rows(min_row=2, values_only=True):
            row_vals = list(row)
            if not any(row_vals):
                continue
            block_num = _bcol(row_vals, "block")
            if block_num is None:
                block_num = len(blocks) + 1
            blocks.append(
                ShootingBlock(
                    block_number=int(block_num),
                    block_type=str(_bcol(row_vals, "type") or "Shoot"),
                    shoot_start=_require_date(_bcol(row_vals, "start"), f"Block {block_num} Shoot Start"),
                    shoot_end=_require_date(_bcol(row_vals, "end"), f"Block {block_num} Shoot End"),
                )
            )

    # --- Sheet: Episode Deliveries ---
    del_sheet = None
    for name in ("Episode Deliveries", "Deliveries", "deliveries", "episode deliveries", "EPISODE DELIVERIES"):
        if name in wb.sheetnames:
            del_sheet = wb[name]
            break

    deliveries: list[EpisodeDelivery] = []
    if del_sheet is not None:
        d_headers = [str(c.value or "").strip().lower() for c in del_sheet[1]]

        def _dcol(row_vals: list, name: str) -> object:
            for i, h in enumerate(d_headers):
                if name in h:
                    return row_vals[i] if i < len(row_vals) else None
            return None

        for row in del_sheet.iter_rows(min_row=2, values_only=True):
            row_vals = list(row)
            if not any(row_vals):
                continue
            ep_num = _dcol(row_vals, "episode")
            if ep_num is None:
                ep_num = len(deliveries) + 1
            deliveries.append(
                EpisodeDelivery(
                    episode_number=int(ep_num),
                    rough_cut_date=_parse_date(_dcol(row_vals, "rough") or _dcol(row_vals, "cut")),
                    picture_lock_date=_parse_date(_dcol(row_vals, "picture") or _dcol(row_vals, "lock")),
                    online_date=_parse_date(_dcol(row_vals, "online")),
                    mix_date=_parse_date(_dcol(row_vals, "mix")),
                    delivery_date=_require_date(_dcol(row_vals, "delivery"), f"Ep {ep_num} Delivery"),
                )
            )

    return ProductionParameters(
        title=title,
        episode_count=episode_count,
        prep_start=prep_start,
        pp_start=pp_start,
        pp_end=pp_end,
        edit_start=edit_start,
        shooting_blocks=blocks,
        episode_deliveries=deliveries,
        final_delivery_date=final_delivery,
        first_payroll_week=first_payroll,
        hiatus_periods=[],
    )


# ---------------------------------------------------------------------------
# Entry point — auto-detects format
# ---------------------------------------------------------------------------

def parse_parameters_excel(file: BinaryIO) -> ProductionParameters:
    """Parse an Excel workbook into ProductionParameters.

    Auto-detects format:
    - Single sheet with section headers → vertical key-value layout
    - Multiple sheets (Info, Shooting Blocks, Episode Deliveries) → legacy format
    """
    wb = openpyxl.load_workbook(file, data_only=True)

    try:
        if len(wb.sheetnames) >= 3:
            # Check if it looks like the legacy 3-sheet format
            has_blocks_sheet = any(
                name.lower() in ("shooting blocks", "blocks")
                for name in wb.sheetnames
            )
            if has_blocks_sheet:
                return _parse_multi_sheet(wb)

        # Default: single-sheet format
        ws = wb.worksheets[0]
        return _parse_single_sheet(ws)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Template generator — matches your single-sheet layout
# ---------------------------------------------------------------------------

_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
_GREEN = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
_CYAN = PatternFill(start_color="00B0F0", end_color="00B0F0", fill_type="solid")
_GRAY_HEADER = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
_BOLD = Font(bold=True)
_BOLD_WHITE = Font(bold=True, color="FFFFFF")


def _set_cell(ws, row, col, value, fill=None, font=None, border=True):
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if border:
        cell.border = _THIN_BORDER
    return cell


def generate_parameters_template() -> BytesIO:
    """Generate the single-sheet Excel template for production parameters."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parameters"

    # Column widths
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    # Row 2: instruction
    ws.cell(row=2, column=1, value="** please enter information in the YELLOW cells").font = Font(
        bold=True, italic=True, color="FF0000"
    )

    # Rows 4-5: Title and Episode Count
    _set_cell(ws, 4, 1, "Title", font=_BOLD)
    _set_cell(ws, 4, 2, "Project Title", fill=_YELLOW)

    _set_cell(ws, 5, 1, "Episode Count", font=_BOLD)
    _set_cell(ws, 5, 2, 1, fill=_YELLOW)

    # Rows 8-13: Key dates
    date_rows = [
        (8, "Prep Start"),
        (9, "PP Start"),
        (10, "PP End"),
        (11, "Edit Start"),
        (12, "Final Delivery"),
        (13, "First Payroll Week"),
    ]
    for row_num, label in date_rows:
        _set_cell(ws, row_num, 1, label, font=_BOLD)
        _set_cell(ws, row_num, 2, None, fill=_YELLOW)
        ws.cell(row=row_num, column=2).number_format = "DD-MMM-YYYY"

    # Row 15: Shoot Blocks section header
    for col in range(1, 5):
        _set_cell(ws, 15, col, "Shoot Blocks" if col == 1 else None, fill=_GREEN, font=_BOLD_WHITE)
    ws.merge_cells(start_row=15, start_column=1, end_row=15, end_column=4)

    # Row 16: Shoot Blocks column headers
    for col, header in enumerate(["Block", "Type", "Shoot Start", "Shoot End"], start=1):
        _set_cell(ws, 16, col, header, fill=_GRAY_HEADER, font=_BOLD)

    # Rows 17-26: 10 blank block rows (yellow)
    for row_num in range(17, 27):
        _set_cell(ws, row_num, 1, row_num - 16 if row_num == 17 else None, fill=_YELLOW)
        _set_cell(ws, row_num, 2, "Shoot" if row_num == 17 else None, fill=_YELLOW)
        _set_cell(ws, row_num, 3, None, fill=_YELLOW)
        ws.cell(row=row_num, column=3).number_format = "DD-MMM-YYYY"
        _set_cell(ws, row_num, 4, None, fill=_YELLOW)
        ws.cell(row=row_num, column=4).number_format = "DD-MMM-YYYY"

    # Row 29: Episode Deliveries section header
    for col in range(1, 7):
        _set_cell(ws, 29, col, "Episode Deliveries" if col == 1 else None, fill=_CYAN, font=_BOLD_WHITE)
    ws.merge_cells(start_row=29, start_column=1, end_row=29, end_column=6)

    # Row 30: Episode Deliveries column headers
    for col, header in enumerate(["Episode", "Rough Cut", "Picture Lock", "Online", "Mix", "Delivery"], start=1):
        _set_cell(ws, 30, col, header, fill=_GRAY_HEADER, font=_BOLD)

    # Rows 31-40: 10 blank delivery rows (yellow)
    for row_num in range(31, 41):
        _set_cell(ws, row_num, 1, row_num - 30 if row_num == 31 else None, fill=_YELLOW)
        for col in range(2, 7):
            _set_cell(ws, row_num, col, None, fill=_YELLOW)
            ws.cell(row=row_num, column=col).number_format = "DD-MMM-YYYY"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb.close()
    return buf
