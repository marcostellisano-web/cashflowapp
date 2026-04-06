"""Parse a breakout-bible Excel file into a list of entry dicts.

Expected format (matches the download produced by write_bible_excel):
  Row 1-3 : title / legend / spacer  (skipped)
  Row 4   : headers — Account | OUT | Prov Labour % | Fed Labour % |
                      Prov Svc Labour % | Svc Property % | Fed Svc Labour %
  Row 5+  : data rows

The parser is lenient:
  - It scans the first 20 rows to locate the header row.
  - Column detection is case-insensitive and fuzzy (substring match for pct cols).
  - Percentage values may be stored as 0–1 floats (openpyxl data_only) or as
    0–100 numbers; both are normalised to 0–1.
  - Blank rows and rows with no account code are skipped.
"""

from __future__ import annotations

from openpyxl import load_workbook

# Maps lower-cased header text → field name
_HEADER_MAP: dict[str, str] = {
    "account":            "account_code",
    "out":                "is_non_prov",
    "prov labour %":      "prov_labour_pct",
    "fed labour %":       "fed_labour_pct",
    "prov svc labour %":  "prov_svc_labour_pct",
    "svc property %":     "svc_property_pct",
    "fed svc labour %":   "fed_svc_labour_pct",
    "description":        "description",
}


def _parse_pct(value) -> float:
    """Normalise a percentage value to the 0–1 range."""
    if value is None:
        return 0.0
    try:
        v = float(value)
    except (TypeError, ValueError):
        return 0.0
    # openpyxl returns 0%-formatted cells as floats in 0–1 range already.
    # If the user typed 65 instead of 0.65 we handle that too.
    if abs(v) > 1.5:
        return v / 100.0
    return v


def parse_bible_excel(file) -> list[dict]:
    """Parse *file* (file-like object) and return a list of entry dicts.

    Each dict has keys:
        account_code, description, is_non_prov,
        prov_labour_pct, fed_labour_pct, prov_svc_labour_pct,
        svc_property_pct, fed_svc_labour_pct
    """
    wb = load_workbook(file, data_only=True)
    ws = wb.active

    # ── locate header row ────────────────────────────────────────────────────
    header_row_idx: int | None = None
    col_map: dict[int, str] = {}  # column index (1-based) → field name

    for row in ws.iter_rows(max_row=20):
        for cell in row:
            if str(cell.value or "").strip().lower() == "account":
                header_row_idx = cell.row
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError(
            "Could not find a header row containing 'Account'. "
            "Please use the downloaded Breakout Bible Excel as a template."
        )

    # Map column numbers to field names
    for cell in ws[header_row_idx]:
        key = str(cell.value or "").strip().lower()
        if key in _HEADER_MAP:
            col_map[cell.column] = _HEADER_MAP[key]

    if "account_code" not in col_map.values():
        raise ValueError("Header row found but no 'Account' column could be mapped.")

    # ── parse data rows ──────────────────────────────────────────────────────
    entries: list[dict] = []

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_data: dict[str, object] = {}
        for cell in row:
            field = col_map.get(cell.column)
            if field:
                row_data[field] = cell.value

        account_code = str(row_data.get("account_code") or "").strip()
        if not account_code:
            continue  # blank row

        # OUT column: written as the text "OUT" when True, empty otherwise
        out_raw = row_data.get("is_non_prov")
        if isinstance(out_raw, str):
            is_non_prov = out_raw.strip().upper() == "OUT"
        elif isinstance(out_raw, bool):
            is_non_prov = out_raw
        else:
            is_non_prov = False

        entries.append({
            "account_code":        account_code,
            "description":         str(row_data.get("description") or "").strip(),
            "is_non_prov":         is_non_prov,
            "prov_labour_pct":     _parse_pct(row_data.get("prov_labour_pct")),
            "fed_labour_pct":      _parse_pct(row_data.get("fed_labour_pct")),
            "prov_svc_labour_pct": _parse_pct(row_data.get("prov_svc_labour_pct")),
            "svc_property_pct":    _parse_pct(row_data.get("svc_property_pct")),
            "fed_svc_labour_pct":  _parse_pct(row_data.get("fed_svc_labour_pct")),
        })

    if not entries:
        raise ValueError("No account rows found in the uploaded file.")

    return entries
