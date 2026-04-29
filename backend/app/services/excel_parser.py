"""Parse Movie Magic Budgeting Excel exports into structured budget data."""

from io import BytesIO
from typing import BinaryIO

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from app.models.budget import BudgetCategory, BudgetDetailRow, BudgetLineItem, ParsedBudget

# Keywords used to detect header rows (case-insensitive)
CODE_HEADERS = {
    "code", "account", "acct", "account no", "acct no", "account number",
    "no", "no.", "account #", "acct #", "#",
}
DESC_HEADERS = {
    "description", "desc", "detail", "category", "account name", "name",
    "categories",
}
TOTAL_HEADERS = {"total", "budget", "amount", "budgeted", "estimate", "total budget"}

# Keywords that indicate a subtotal/total row to skip
SKIP_KEYWORDS = {"total", "subtotal", "sub-total", "sub total", "grand total"}

# Sheet names to look for (in order of priority)
PREFERRED_SHEETS = ["categories", "top sheet", "budget summary", "budget", "summary"]

# Sheet name variants that contain the pre-aggregated CAVCO topsheet totals
TOPSHEET_NAMES = {"topsheet", "top sheet", "top-sheet", "budget topsheet"}


DETAIL_SHEET_NAMES = {"account details", "account detail"}

_DETAIL_HEADERS = {
    "account": {"account", "acct", "code", "account #", "account no", "account number"},
    "description": {"description", "desc", "detail"},
    "amount": {"amount", "no.", "no", "number"},
    "unit": {"unit"},
    "x": {"x", "*"},
    "unit2": {"unit 2", "unit2", "units"},
    "currency": {"currency", "cur"},
    "rate": {"rate", "rate/amt", "amt", "rate/amount"},
    "unit3": {"unit 3", "unit3"},
    "unit4": {"unit 4", "unit4"},
    "subtotal": {"subtotal", "sub total", "sub-total", "total"},
    "agg": {"agg", "agg%", "agg %", "aggregate", "fringe%", "fringe %", "fringe rate"},
    "groups": {"groups", "group", "group code", "grp", "grp."},
}


def _find_budget_sheet(wb: openpyxl.Workbook) -> Worksheet:
    """Find the most likely budget sheet in the workbook."""
    sheet_names_lower = {name.lower(): name for name in wb.sheetnames}
    for preferred in PREFERRED_SHEETS:
        if preferred in sheet_names_lower:
            return wb[sheet_names_lower[preferred]]
    # Fallback: partial match (e.g. "Categories (Detail)" or "Budget Categories")
    for sheet_lower, sheet_real in sheet_names_lower.items():
        if "categories" in sheet_lower:
            return wb[sheet_real]
    # Fallback to the first sheet
    return wb.active or wb.worksheets[0]


def _detect_headers(ws: Worksheet, max_scan_rows: int = 25) -> tuple[int, dict[str, int]]:
    """Scan the first N rows to find the header row and column mapping.

    Returns (header_row_index, {"code": col, "description": col, "total": col}).
    """
    for row_idx in range(1, min(max_scan_rows + 1, ws.max_row + 1)):
        cells = {}
        for col_idx in range(1, min(ws.max_column + 1, 50)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                cells[col_idx] = str(val).strip().lower()

        code_col = None
        desc_col = None
        total_col = None

        for col, text in cells.items():
            if text in CODE_HEADERS and code_col is None:
                code_col = col
            elif text in DESC_HEADERS and desc_col is None:
                desc_col = col
            elif text in TOTAL_HEADERS and total_col is None:
                total_col = col

        if code_col and desc_col and total_col:
            return row_idx, {"code": code_col, "description": desc_col, "total": total_col}

    # Fallback: try the Movie Magic "Categories" tab fixed layout.
    # Account # in column A (1), Description in column B (2), Total in column E (5).
    # Detect header row by looking for a numeric-looking value in column A.
    col_map = {"code": 1, "description": 2, "total": 5}
    header_row = _detect_categories_header_row(ws, col_map, max_scan_rows)
    if header_row is not None:
        return header_row, col_map

    raise ValueError(
        "Could not find header row with Code, Description, and Total columns. "
        "Please ensure your budget file has columns labeled with recognizable headers "
        "(e.g. 'Account #', 'Description', 'Total') or uses the Movie Magic Categories layout."
    )


def _detect_categories_header_row(
    ws: Worksheet, col_map: dict[str, int], max_scan_rows: int
) -> int | None:
    """Detect the header row for the Movie Magic Categories fixed layout.

    Looks for the first row where column A contains a budget-code-like value
    (numeric string like '1001') and column E contains a parseable amount.
    Returns the row *before* that as the header row, or 0 if data starts at row 1.
    """
    for row_idx in range(1, min(max_scan_rows + 1, (ws.max_row or 0) + 1)):
        code_val = ws.cell(row=row_idx, column=col_map["code"]).value
        total_val = ws.cell(row=row_idx, column=col_map["total"]).value

        if code_val is None:
            continue

        code_str = str(code_val).strip()

        # Check if it looks like a header row itself (text like "Account #")
        code_lower = code_str.lower()
        if code_lower in CODE_HEADERS or any(kw in code_lower for kw in ("account", "acct", "code")):
            return row_idx

        # Check if it looks like a budget code (numeric, 2-6 digits)
        if code_str.isdigit() and 2 <= len(code_str) <= 6 and _parse_amount(total_val) is not None:
            # Data starts here; header row is the row before, or 0 if row 1
            return max(row_idx - 1, 0)

    return None


def _classify_category(code: str) -> BudgetCategory:
    """Classify a budget code into a category based on Movie Magic prefix ranges."""
    if len(code) < 2:
        return BudgetCategory.OTHER
    try:
        prefix = int(code[:2])
    except ValueError:
        return BudgetCategory.OTHER

    if 10 <= prefix <= 19:
        return BudgetCategory.ABOVE_THE_LINE
    elif 20 <= prefix <= 39:
        return BudgetCategory.BELOW_THE_LINE_PRODUCTION
    elif 40 <= prefix <= 49:
        return BudgetCategory.BELOW_THE_LINE_POST
    else:
        return BudgetCategory.OTHER


def _is_subtotal_row(description: str) -> bool:
    """Check if a row is a subtotal/total row based on its description."""
    desc_lower = description.lower().strip()
    return any(kw in desc_lower for kw in SKIP_KEYWORDS)


def _parse_amount(value) -> float | None:
    """Try to parse a cell value as a float amount."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Remove currency symbols, commas, whitespace
        cleaned = value.replace("$", "").replace(",", "").replace(" ", "").strip()
        if not cleaned or cleaned == "-":
            return None
        try:
            return float(cleaned)
        except ValueError:
            return None
    return None


def _parse_topsheet_tab(wb: openpyxl.Workbook) -> dict[str, float]:
    """Find and parse the Topsheet tab, returning a {account_code: total} dict.

    The sheet is expected to have columns: Account, Category Description,
    [Fringe], [Original], Total, [Variance].  We identify the "Total" column
    by scanning the header row for the word "total" and use the first numeric
    4-digit code column as the account key.

    Returns an empty dict if no matching sheet or data is found.
    """
    sheet_names_lower = {name.lower().strip(): name for name in wb.sheetnames}
    ws = None
    for variant in TOPSHEET_NAMES:
        if variant in sheet_names_lower:
            ws = wb[sheet_names_lower[variant]]
            break
    if ws is None:
        return {}

    # Scan first 10 rows for a header row containing "total"
    total_col: int | None = None
    account_col: int | None = None
    header_row_idx: int | None = None

    for row_idx in range(1, min(11, (ws.max_row or 0) + 1)):
        row_texts = {}
        for col_idx in range(1, min(ws.max_column + 1, 20)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                row_texts[col_idx] = str(val).strip().lower()

        # Require at least "account" (or similar) and "total" in the same row
        has_account = any(t in CODE_HEADERS or "account" in t for t in row_texts.values())
        total_cols = [c for c, t in row_texts.items() if t == "total"]
        if has_account and total_cols:
            header_row_idx = row_idx
            total_col = total_cols[-1]  # pick the rightmost "total" if multiple
            # Account column: first column whose header is in CODE_HEADERS
            for c, t in row_texts.items():
                if t in CODE_HEADERS or "account" in t:
                    account_col = c
                    break
            break

    if header_row_idx is None or total_col is None or account_col is None:
        # Fallback: assume account in col 1, total in col 5
        header_row_idx = 1
        account_col = 1
        total_col = 5

    totals: dict[str, float] = {}
    for row_idx in range(header_row_idx + 1, (ws.max_row or 0) + 1):
        acct_val = ws.cell(row=row_idx, column=account_col).value
        total_val = ws.cell(row=row_idx, column=total_col).value

        if acct_val is None:
            continue
        acct_str = str(acct_val).strip().replace("'", "").replace("`", "")
        # Must be exactly 4 numeric digits (e.g. "0100", "2200")
        if not acct_str.isdigit() or len(acct_str) != 4:
            continue

        amount = _parse_amount(total_val)
        if amount is not None:
            totals[acct_str] = amount

    return totals




def _find_sheet_by_name_variants(wb: openpyxl.Workbook, variants: set[str]) -> Worksheet | None:
    sheet_names_lower = {name.lower().strip(): name for name in wb.sheetnames}
    for variant in variants:
        if variant in sheet_names_lower:
            return wb[sheet_names_lower[variant]]
    for lowered, real in sheet_names_lower.items():
        if any(v in lowered for v in variants):
            return wb[real]
    return None


def _detect_detail_headers(ws: Worksheet, max_scan_rows: int = 20) -> tuple[int, dict[str, int]] | None:
    for row_idx in range(1, min(max_scan_rows + 1, (ws.max_row or 0) + 1)):
        columns: dict[str, int] = {}
        for col_idx in range(1, min(ws.max_column + 1, 40)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is None:
                continue
            token = str(val).strip().lower()
            for key, variants in _DETAIL_HEADERS.items():
                if token in variants and key not in columns:
                    columns[key] = col_idx
        if columns.get("account") and columns.get("description") and columns.get("subtotal"):
            return row_idx, columns
    return None


def _parse_account_details_tab(wb: openpyxl.Workbook) -> list[BudgetDetailRow]:
    ws = _find_sheet_by_name_variants(wb, DETAIL_SHEET_NAMES)
    if ws is None:
        return []

    detected = _detect_detail_headers(ws)
    if detected is None:
        return []

    header_row, col_map = detected
    rows: list[BudgetDetailRow] = []
    last_account = ""
    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        account_val = ws.cell(row=row_idx, column=col_map["account"]).value
        desc_val = ws.cell(row=row_idx, column=col_map["description"]).value
        subtotal_val = ws.cell(row=row_idx, column=col_map["subtotal"]).value

        if account_val is None and desc_val is None and subtotal_val is None:
            continue

        account = str(account_val or "").strip()
        desc = str(desc_val or "").strip()
        subtotal = _parse_amount(subtotal_val) or 0.0

        if account:
            last_account = account

        if not account:
            # "Total Fringes" aggregate rows in Movie Magic often have no account
            # number. Inherit the parent account so they can be matched later.
            if subtotal > 0 and desc.strip().lower() == "total fringes" and last_account:
                account = last_account
            else:
                continue
        elif subtotal <= 0:
            continue

        def _txt(col_name: str) -> str | None:
            col = col_map.get(col_name)
            if not col:
                return None
            val = ws.cell(row=row_idx, column=col).value
            if val is None:
                return None
            text = str(val).strip()
            return text or None

        def _num(col_name: str) -> float | None:
            col = col_map.get(col_name)
            if not col:
                return None
            return _parse_amount(ws.cell(row=row_idx, column=col).value)

        rows.append(
            BudgetDetailRow(
                account=account,
                description=desc,
                amount=_num("amount"),
                unit=_txt("unit"),
                x=_txt("x"),
                unit2=_txt("unit2"),
                currency=_txt("currency"),
                rate=_num("rate"),
                unit3=_txt("unit3"),
                unit4=_txt("unit4"),
                subtotal=subtotal,
                agg=_num("agg"),
                groups=_txt("groups"),
            )
        )

    return rows


def parse_budget_excel(file: BinaryIO, filename: str = "uploaded.xlsx") -> ParsedBudget:
    """Parse a Movie Magic Budgeting Excel export into a ParsedBudget.

    Args:
        file: A file-like object containing the Excel data.
        filename: Original filename for reference.

    Returns:
        ParsedBudget with extracted line items and any warnings.
    """
    wb = openpyxl.load_workbook(file, data_only=True)
    ws = _find_budget_sheet(wb)

    header_row, col_map = _detect_headers(ws)

    line_items: list[BudgetLineItem] = []
    warnings: list[str] = []
    current_group: str | None = None

    for row_idx in range(header_row + 1, (ws.max_row or 0) + 1):
        code_val = ws.cell(row=row_idx, column=col_map["code"]).value
        desc_val = ws.cell(row=row_idx, column=col_map["description"]).value
        total_val = ws.cell(row=row_idx, column=col_map["total"]).value

        # Skip fully empty rows
        if code_val is None and desc_val is None and total_val is None:
            continue

        code_str = str(code_val).strip() if code_val is not None else ""
        desc_str = str(desc_val).strip() if desc_val is not None else ""

        # Detect section headers (have description but no code or no amount)
        amount = _parse_amount(total_val)
        if not code_str and desc_str and amount is None:
            current_group = desc_str
            continue

        # Skip subtotal rows
        if _is_subtotal_row(desc_str):
            continue

        # Skip rows with no code
        if not code_str:
            continue

        # Skip rows with zero or no amount
        if amount is None or amount == 0:
            if desc_str:
                warnings.append(f"Row {row_idx}: '{desc_str}' has no amount, skipped")
            continue

        category = _classify_category(code_str)

        line_items.append(
            BudgetLineItem(
                code=code_str,
                description=desc_str,
                total=amount,
                category=category,
                account_group=current_group,
            )
        )

    topsheet_totals = _parse_topsheet_tab(wb)
    detail_rows = _parse_account_details_tab(wb)
    wb.close()

    total_budget = sum(item.total for item in line_items)

    return ParsedBudget(
        line_items=line_items,
        total_budget=total_budget,
        source_filename=filename,
        warnings=warnings,
        topsheet_totals=topsheet_totals,
        detail_rows=detail_rows,
    )
