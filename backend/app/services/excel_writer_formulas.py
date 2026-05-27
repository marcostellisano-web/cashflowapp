"""Formula-based cashflow Excel workbook generator.

Produces the standard workbook plus a "Timing" sheet with one editable row
per timing category from the production bible (e.g. "Shoot Payroll",
"Internals", "Travel", "Pick Lock", "After Delivery", …).

Each Timing row contains 0/1 values across every week column, computed from
the real production schedule (payroll weeks, shoot blocks, picture-lock dates,
delivery dates, etc.).  Users can flip any 0 ↔ 1 to change when a cost is
distributed.

Cashflow weekly cells use:
    =IFERROR( $C{row} * Timing!{col}${prow} / SUM(Timing!$D${prow}:$X${prow}), 0 )
"""

from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.budget import ParsedBudget
from app.models.cashflow import CashflowOutput
from app.models.distribution import LineItemDistribution, PhaseAssignment
from app.models.production import ProductionParameters
from app.services.excel_writer import (
    CURRENCY_FORMAT,
    DATA_START_ROW,
    FIRST_WEEK_COL,
    HEADER_FONT,
    THIN_BORDER,
    TITLE_FONT,
    TOTAL_COL,
    _get_phase_fill,
    write_cashflow_excel,
)

# ── Bible: account code → timing title ───────────────────────────────────────
# First 4 digits of account code → Timing Title label.
# Codes with sub-accounts (e.g. "040501") are matched on the first 4 digits.

ACCOUNT_TIMING_MAP: dict[str, str] = {
    "0220": "PP to End",
    "0401": "Internals", "0405": "Internals", "0407": "Full Payroll",
    "0408": "Full Payroll", "0410": "Full Payroll",
    "0460": "Travel", "0461": "Travel", "0465": "Travel",
    "0466": "Travel", "0470": "Travel", "0490": "Travel",
    "0501": "Shoot Payroll", "0601": "Shoot Payroll", "0640": "Shoot Payroll",
    "1001": "Shoot Payroll", "1010": "Shoot Payroll",
    "1025": "Narrator",
    "1070": "Casting",
    "1095": "Narrator",
    "1101": "Shoot Payroll", "1110": "Shoot Payroll",
    "1201": "Internals", "1205": "Full Payroll",
    "1210": "Shoot Payroll", "1220": "Shoot Payroll", "1223": "Shoot Payroll",
    "1235": "Shoot Payroll", "1243": "Full Payroll",
    "1248": "Internals", "1250": "Internals", "1254": "Internals",
    "1270": "Shoot Payroll", "1301": "Shoot Payroll", "1310": "Shoot Payroll",
    "1312": "Shoot Payroll", "1320": "Shoot Payroll", "1401": "Shoot Payroll",
    "1495": "Shoot Payroll", "1501": "Shoot Payroll", "1510": "Shoot Payroll",
    "1520": "Shoot Payroll", "1530": "Shoot Payroll",
    "1601": "Shoot Payroll", "1610": "Shoot Payroll", "1616": "Shoot Payroll",
    "1905": "Shoot Payroll", "1910": "Shoot Payroll",
    "2001": "Shoot Payroll", "2010": "Shoot Payroll",
    "2201": "Shoot Payroll", "2205": "Shoot Payroll", "2210": "Shoot Payroll",
    "2212": "Shoot Payroll", "2220": "Shoot Payroll", "2250": "Shoot Payroll",
    "2260": "Travel", "2263": "Travel",
    "2270": "Still Photo",
    "2301": "Shoot Payroll", "2310": "Shoot Payroll", "2320": "Shoot Payroll",
    "2330": "Shoot Payroll", "2340": "Shoot Payroll", "2350": "Shoot Payroll",
    "2401": "Shoot Payroll", "2410": "Shoot Payroll",
    "2501": "Shoot Payroll",
    "2801": "Internals", "2805": "Internals", "2807": "Internals",
    "2810": "Internals", "2815": "Internals", "2820": "Internals",
    "2830": "Internals", "2835": "Internals", "2840": "Internals",
    "2845": "Internals",
    "2901": "Monthly (shoot months)",
    "2955": "Shoot Purchases",
    "3020": "Monthly (shoot months)", "3095": "Monthly (shoot months)",
    "3105": "Monthly (shoot months)",
    "3110": "Shoot Rentals",
    "3142": "Monthly (shoot months)", "3195": "Monthly (shoot months)",
    "3210": "Per Diem", "3215": "Per Diem",
    "3218": "Shoot Rentals",
    "3225": "Pre-Shoot", "3260": "Pre-Shoot",
    "3301": "Travel", "3310": "Travel",
    "3320": "Per Diem", "3330": "Per Diem", "3335": "Per Diem",
    "3340": "Monthly (shoot months)",
    "3395": "Legal",
    "3401": "Monthly (shoot months)", "3405": "Monthly (shoot months)",
    "3412": "Monthly (shoot months)", "3415": "Monthly (shoot months)",
    "3430": "Per Diem", "3440": "Per Diem", "3445": "Per Diem",
    "3447": "Per Diem",
    "3510": "Shoot Rentals", "3515": "Shoot Purchases",
    "3545": "Shoot Purchases",
    "3710": "Shoot Rentals", "3730": "Shoot Purchases",
    "3740": "Shoot Purchases",
    "3810": "Shoot Rentals", "3830": "Shoot Purchases",
    "3850": "Shoot Rentals",
    "3910": "Shoot Rentals", "3930": "Shoot Purchases",
    "4110": "Shoot Rentals", "4130": "Shoot Purchases",
    "4140": "Shoot Purchases", "4148": "Shoot Purchases",
    "4210": "Shoot Rentals", "4212": "Shoot Purchases",
    "4510": "Shoot Rentals", "4512": "Shoot Rentals",
    "4515": "Shoot Rentals", "4525": "Shoot Rentals",
    "4530": "Shoot Purchases",
    "4595": "Shoot Rentals",
    "4610": "Shoot Rentals", "4612": "Shoot Rentals",
    "4630": "Shoot Purchases",
    "4710": "Shoot Rentals", "4730": "Shoot Purchases",
    "4795": "Shoot Purchases",
    "4812": "Shoot Rentals", "4828": "Shoot Purchases",
    "4830": "Shoot Purchases",
    "5001": "Pre-Shoot",
    "6001": "Edit Internals",
    "6002": "PP to End",
    "6010": "Edit", "6012": "Edit",
    "6042": "Archive Research",
    "6070": "Pick Lock",
    "6101": "Edit Internals", "6110": "Edit Internals",
    "6130": "Edit Internals",
    "6215": "Edit Internals",
    "6221": "Online Editor",
    "6264": "Delivery copies",
    "6325": "Mix",
    "6610": "Composer",
    "6670": "After Delivery", "6695": "After Delivery",
    "6710": "Graphics", "6730": "Graphics", "6795": "Graphics",
    "6890": "Pick Lock", "6892": "Pick Lock",
    "7001": "Edit Internals",
    "7025": "Still Photo",
    "7040": "Pick Lock",
    "7050": "Edit Internals",
    "7101": "Insurance",
    "7110": "Legal",
    "7120": "After Delivery", "7125": "After Delivery",
    "7130": "Full AP",
    "7201": "Internals",
    "7210": "After Delivery",
    "7220": "Financing",
    "7295": "After Delivery",
}

# Phase-assignment fallback when code is not in the bible
_PHASE_FALLBACK: dict[PhaseAssignment, str] = {
    PhaseAssignment.PREP: "Internals",
    PhaseAssignment.PRODUCTION: "Shoot Payroll",
    PhaseAssignment.POST: "Edit",
    PhaseAssignment.DELIVERY: "After Delivery",
    PhaseAssignment.FULL_SPAN: "Full Payroll",
    PhaseAssignment.PREP_AND_PRODUCTION: "Full Payroll",
    PhaseAssignment.PRODUCTION_AND_POST: "Edit",
}

_TIMING_DESCRIPTIONS: dict[str, str] = {
    "Full Payroll":         "Payroll weeks — start of prep through final delivery",
    "Full AP":              "AP weeks — start of prep through final delivery",
    "Internals":            "Monthly mid-month AP weeks — prep through delivery",
    "Shoot Payroll":        "Payroll weeks during the shoot (starts 1-2 weeks in)",
    "Shoot Rentals":        "AP weeks during the shoot (starts 1-2 weeks in)",
    "Shoot Purchases":      "AP weeks during each shoot block",
    "Per Diem":             "AP weeks during each shoot block (per diems / catering)",
    "Monthly (shoot months)": "Last AP week of each calendar month during the shoot",
    "Travel":               "AP weeks ~2 weeks before each shoot block",
    "Still Photo":          "Payroll weeks ~2 weeks after each shoot block ends",
    "Pre-Shoot":            "AP week ~2 weeks before Principal Photography starts",
    "Casting":              "Two payroll weeks: 4 wks and 2 wks before PP start",
    "Legal":                "4 AP weeks spread evenly from prep start to final delivery",
    "PP to End":            "Payroll weeks from 2 wks before edit start through final delivery",
    "Archive Research":     "Payroll weeks from 2 wks before PP start through final delivery",
    "Narrator":             "Payroll weeks from ~4 wks after edit start through final delivery",
    "Edit":                 "Payroll weeks from edit start through final delivery",
    "Edit Internals":       "Monthly mid-month AP weeks during the edit period",
    "Pick Lock":            "AP weeks ~2-3 wks after each episode picture lock",
    "Online Editor":        "Payroll weeks around each episode online date",
    "Delivery copies":      "AP weeks ~2-3 wks before each episode delivery",
    "Mix":                  "AP weeks ~2-3 wks after each episode mix date",
    "Composer":             "Payroll weeks: edit midpoint and near final picture lock",
    "Graphics":             "Bi-weekly AP weeks from edit start through final delivery",
    "After Delivery":       "AP week ~4 wks after final delivery",
    "Financing":            "AP week nearest to end of fiscal year (September) per year",
    "Insurance":            "AP week ~2 wks after prep start",
}

# Fills for the Timing sheet
_ACTIVE_FILL = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
_INACTIVE_FILL = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")

_TIMING_SHEET = "Timing"


# ── Lookup helpers ────────────────────────────────────────────────────────────

def _get_timing_title(code: str, phase: PhaseAssignment) -> str:
    """Return the timing title for a budget code, falling back to phase."""
    padded = code.strip().zfill(6)[:4]   # normalise: zero-pad, take first 4 digits
    return ACCOUNT_TIMING_MAP.get(padded, _PHASE_FALLBACK.get(phase, "Shoot Payroll"))


# ── Pattern computation ───────────────────────────────────────────────────────

def _nearest_week_idx(
    weeks,
    target: date,
    predicate,
    window_days: int = 45,
) -> int | None:
    best_idx, best_diff = None, float("inf")
    for i, w in enumerate(weeks):
        if predicate(w):
            diff = abs((w.week_commencing - target).days)
            if diff < best_diff and diff <= window_days:
                best_diff = diff
                best_idx = i
    return best_idx


def _compute_timing_patterns(
    titles: list[str],
    output: CashflowOutput,
    params: ProductionParameters,
) -> dict[str, list[int]]:
    """Return {title: [0/1, …]} for each title, derived from the schedule."""
    weeks = output.weeks
    n = len(weeks)
    blocks = params.shooting_blocks
    eps = params.episode_deliveries

    has_payroll = any(w.is_payroll_week is not None for w in weeks)

    def payroll(w) -> bool:
        return w.is_payroll_week is True if has_payroll else True

    def ap(w) -> bool:
        return w.is_payroll_week is False if has_payroll else True

    def shoot(w) -> bool:
        return "SHOOT" in w.phase_label.upper()

    def in_range(w, start: date, end: date) -> bool:
        return start <= w.week_commencing <= end

    def zeros() -> list[int]:
        return [0] * n

    def mark_nearest(target: date, pred, window=45) -> list[int]:
        result = zeros()
        idx = _nearest_week_idx(weeks, target, pred, window)
        if idx is not None:
            result[idx] = 1
        return result

    def mark_multi(targets: list[date], pred, window=35) -> list[int]:
        result = zeros()
        for t in targets:
            idx = _nearest_week_idx(weeks, t, pred, window)
            if idx is not None:
                result[idx] = 1
        return result

    patterns: dict[str, list[int]] = {}

    for title in titles:
        if title == "Full Payroll":
            patterns[title] = [1 if payroll(w) else 0 for w in weeks]

        elif title == "Full AP":
            patterns[title] = [1 if ap(w) else 0 for w in weeks]

        elif title == "Internals":
            # Mid-month AP weeks (roughly the week containing the 12th–19th)
            patterns[title] = [
                1 if (ap(w) and 8 <= w.week_commencing.day <= 21) else 0
                for w in weeks
            ]

        elif title == "Shoot Payroll":
            patterns[title] = [1 if (payroll(w) and shoot(w)) else 0 for w in weeks]

        elif title == "Shoot Rentals":
            patterns[title] = [1 if (ap(w) and shoot(w)) else 0 for w in weeks]

        elif title in ("Shoot Purchases", "Per Diem"):
            patterns[title] = [1 if (ap(w) and shoot(w)) else 0 for w in weeks]

        elif title == "Monthly (shoot months)":
            # Last AP week of each calendar month during the shoot
            result = zeros()
            shoot_months: set[tuple[int, int]] = set()
            for w in weeks:
                if shoot(w):
                    shoot_months.add((w.week_commencing.year, w.week_commencing.month))
            for i, w in enumerate(weeks):
                ym = (w.week_commencing.year, w.week_commencing.month)
                if ym not in shoot_months or not ap(w) or not shoot(w):
                    continue
                # Is this the last AP+shoot week of the month?
                is_last = not any(
                    ap(weeks[j]) and shoot(weeks[j])
                    and (weeks[j].week_commencing.year, weeks[j].week_commencing.month) == ym
                    for j in range(i + 1, n)
                )
                if is_last:
                    result[i] = 1
            patterns[title] = result

        elif title == "Travel":
            # AP week ~2 weeks before each shoot block
            targets = [b.shoot_start - timedelta(weeks=2) for b in blocks]
            patterns[title] = mark_multi(targets, ap)

        elif title == "Still Photo":
            # Payroll week ~2 weeks after each shoot block ends
            targets = [b.shoot_end + timedelta(weeks=2) for b in blocks]
            patterns[title] = mark_multi(targets, payroll)

        elif title == "Pre-Shoot":
            # AP week ~2 weeks before PP starts
            patterns[title] = mark_nearest(params.pp_start - timedelta(weeks=2), ap)

        elif title == "Casting":
            # Two payroll weeks: 4 weeks and 2 weeks before PP start
            targets = [
                params.pp_start - timedelta(weeks=4),
                params.pp_start - timedelta(weeks=2),
            ]
            patterns[title] = mark_multi(targets, payroll)

        elif title == "Legal":
            # 4 AP weeks evenly spread from prep start to final delivery
            total = (params.final_delivery_date - params.prep_start).days
            targets = [
                params.prep_start + timedelta(days=int(total * (q + 0.5) / 4))
                for q in range(4)
            ]
            patterns[title] = mark_multi(targets, ap, window=60)

        elif title == "PP to End":
            start = params.edit_start - timedelta(weeks=2)
            patterns[title] = [
                1 if (payroll(w) and in_range(w, start, params.final_delivery_date))
                else 0
                for w in weeks
            ]

        elif title == "Archive Research":
            start = params.pp_start - timedelta(weeks=2)
            patterns[title] = [
                1 if (payroll(w) and in_range(w, start, params.final_delivery_date))
                else 0
                for w in weeks
            ]

        elif title == "Narrator":
            start = params.edit_start + timedelta(weeks=4)
            patterns[title] = [
                1 if (payroll(w) and in_range(w, start, params.final_delivery_date))
                else 0
                for w in weeks
            ]

        elif title == "Edit":
            patterns[title] = [
                1 if (payroll(w) and in_range(w, params.edit_start, params.final_delivery_date))
                else 0
                for w in weeks
            ]

        elif title == "Edit Internals":
            patterns[title] = [
                1 if (
                    ap(w)
                    and 8 <= w.week_commencing.day <= 21
                    and in_range(w, params.edit_start, params.final_delivery_date)
                )
                else 0
                for w in weeks
            ]

        elif title == "Pick Lock":
            # AP week ~2-3 weeks after each episode picture lock
            targets = [
                e.picture_lock_date + timedelta(weeks=2)
                for e in eps
                if e.picture_lock_date
            ]
            patterns[title] = mark_multi(targets, ap) if targets else zeros()

        elif title == "Online Editor":
            # Payroll weeks around each episode online date
            targets = [e.online_date for e in eps if e.online_date]
            patterns[title] = mark_multi(targets, payroll) if targets else zeros()

        elif title == "Delivery copies":
            # AP week ~2 weeks before each episode delivery
            targets = [
                e.delivery_date - timedelta(weeks=2)
                for e in eps
                if e.delivery_date
            ]
            patterns[title] = mark_multi(targets, ap) if targets else zeros()

        elif title == "Mix":
            # AP week ~2-3 weeks after each episode mix
            targets = [
                e.mix_date + timedelta(weeks=2)
                for e in eps
                if e.mix_date
            ]
            patterns[title] = mark_multi(targets, ap) if targets else zeros()

        elif title == "Composer":
            # Payroll weeks: edit midpoint and near final picture lock
            edit_len = (params.final_delivery_date - params.edit_start).days
            midpoint = params.edit_start + timedelta(days=edit_len // 2)
            lock = max(
                (e.picture_lock_date for e in eps if e.picture_lock_date),
                default=params.final_delivery_date - timedelta(weeks=2),
            )
            patterns[title] = mark_multi([midpoint, lock], payroll)

        elif title == "Graphics":
            # AP weeks every 2 weeks from edit start through final delivery
            result = zeros()
            ap_count = 0
            for i, w in enumerate(weeks):
                if in_range(w, params.edit_start, params.final_delivery_date) and ap(w):
                    ap_count += 1
                    if ap_count % 2 == 0:
                        result[i] = 1
            patterns[title] = result

        elif title == "After Delivery":
            # AP week ~4 weeks after final delivery
            patterns[title] = mark_nearest(
                params.final_delivery_date + timedelta(weeks=4), ap, window=60
            )

        elif title == "Financing":
            # AP week nearest to end of September for each fiscal year in the schedule
            years = {w.week_commencing.year for w in weeks}
            targets = [date(y, 9, 30) for y in years]
            patterns[title] = mark_multi(targets, ap, window=60)

        elif title == "Insurance":
            # AP week ~2 weeks after prep start
            patterns[title] = mark_nearest(
                params.prep_start + timedelta(weeks=2), ap
            )

        else:
            # Unknown title — spread evenly over all non-hiatus payroll weeks
            patterns[title] = [
                1 if (payroll(w) and "HIATUS" not in w.phase_label.upper()) else 0
                for w in weeks
            ]

    return patterns


# ── Timing sheet writer ───────────────────────────────────────────────────────

def _write_timing_sheet(
    wb,
    output: CashflowOutput,
    params: ProductionParameters,
    title_to_row: dict[str, int],
) -> None:
    """Create the 'Timing' sheet with one pattern row per timing category."""
    ws = wb.create_sheet(_TIMING_SHEET, 1)
    weeks = output.weeks
    num_weeks = len(weeks)
    last_col = FIRST_WEEK_COL + num_weeks - 1
    first_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_col_letter = get_column_letter(last_col)

    patterns = _compute_timing_patterns(list(title_to_row.keys()), output, params)

    # ── Row 1: title ──────────────────────────────────────────────────────────
    ws.cell(row=1, column=1, value="Distribution Timing Patterns").font = TITLE_FONT

    # ── Row 2: instructions ───────────────────────────────────────────────────
    ws.cell(
        row=2, column=1,
        value=(
            "Edit 0 / 1 values in the week columns to change when a cost is distributed.  "
            "Green = active (1), yellow = inactive (0).  "
            "Each Cashflow row computes:  Budget × weight ÷ SUM(all weights for that row)."
        ),
    ).font = Font(name="Calibri", italic=True, size=9, color="555555")

    # ── Row 3: column A/B headers + phase labels live-linked from Cashflow ────
    for col_idx, label in ((1, "Timing Pattern"), (2, "Description")):
        cell = ws.cell(row=3, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = THIN_BORDER

    for i, week in enumerate(weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(row=3, column=col, value=f"=Cashflow!{col_letter}4")
        cell.font = Font(name="Calibri", bold=True, size=8)
        cell.fill = _get_phase_fill(week.phase_label)
        cell.alignment = Alignment(horizontal="center", text_rotation=90)
        cell.border = THIN_BORDER

    # ── Row 4: week dates live-linked from Cashflow ───────────────────────────
    ws.cell(row=4, column=1, value="Week commencing →").font = Font(
        name="Calibri", italic=True, size=8, color="888888"
    )
    for i in range(num_weeks):
        col = FIRST_WEEK_COL + i
        col_letter = get_column_letter(col)
        cell = ws.cell(row=4, column=col, value=f"=Cashflow!{col_letter}5")
        cell.font = Font(name="Calibri", size=7)
        cell.number_format = "DD-MMM"
        cell.alignment = Alignment(horizontal="center")
        cell.border = THIN_BORDER

    # ── Pattern rows ──────────────────────────────────────────────────────────
    for title, excel_row in title_to_row.items():
        pat = patterns.get(title, [0] * num_weeks)

        label_cell = ws.cell(row=excel_row, column=1, value=title)
        label_cell.font = Font(name="Calibri", bold=True, size=10)
        label_cell.border = THIN_BORDER

        desc_cell = ws.cell(
            row=excel_row, column=2,
            value=_TIMING_DESCRIPTIONS.get(title, title),
        )
        desc_cell.font = Font(name="Calibri", size=9, italic=True, color="444444")
        desc_cell.border = THIN_BORDER

        for i, val in enumerate(pat):
            col = FIRST_WEEK_COL + i
            cell = ws.cell(row=excel_row, column=col, value=val)
            cell.number_format = "0"
            cell.alignment = Alignment(horizontal="center")
            cell.border = THIN_BORDER
            cell.fill = _ACTIVE_FILL if val else _INACTIVE_FILL
            cell.font = Font(
                name="Calibri", bold=bool(val), size=9,
                color="1B5E20" if val else "BBBBBB",
            )

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 48
    for i in range(num_weeks):
        ws.column_dimensions[get_column_letter(FIRST_WEEK_COL + i)].width = 5

    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=FIRST_WEEK_COL)


# ── Main entry point ──────────────────────────────────────────────────────────

def write_formula_cashflow_excel(
    output: CashflowOutput,
    params: ProductionParameters,
    distributions: list[LineItemDistribution],
    budget: ParsedBudget | None = None,
) -> BytesIO:
    """Generate a formula-driven cashflow workbook with an editable Timing sheet.

    1. Build the standard workbook (all sheets, full styling).
    2. Insert a 'Timing' sheet (sheet 2) with one editable row per timing
       category derived from the production bible + schedule.
    3. Rewrite the Cashflow sheet data rows so every weekly cell is a formula
       referencing the Timing sheet rather than a hard-coded value.
    4. Add a 'Pattern' column showing which Timing row each Cashflow line uses.
    """
    # ── 1. Standard workbook ──────────────────────────────────────────────────
    standard_buf = write_cashflow_excel(output, params, budget=budget)
    wb = load_workbook(standard_buf)
    ws = wb["Cashflow"]

    # ── 2. Build timing title per row ─────────────────────────────────────────
    dist_by_code: dict[str, PhaseAssignment] = {
        d.budget_code: d.phase for d in distributions
    }

    # Preserve insertion order (first-seen ordering for Timing sheet rows)
    title_to_row: dict[str, int] = {}
    row_timing: list[str] = []
    for row_data in output.rows:
        phase = dist_by_code.get(row_data.code, PhaseAssignment.PRODUCTION)
        title = _get_timing_title(row_data.code, phase)
        row_timing.append(title)
        if title not in title_to_row:
            title_to_row[title] = DATA_START_ROW + len(title_to_row)

    # ── 3. Create Timing sheet ────────────────────────────────────────────────
    _write_timing_sheet(wb, output, params, title_to_row)

    # ── 4. Rewrite Cashflow data rows ─────────────────────────────────────────
    num_weeks = len(output.weeks)
    last_week_col = FIRST_WEEK_COL + num_weeks - 1
    first_col_letter = get_column_letter(FIRST_WEEK_COL)
    last_week_col_letter = get_column_letter(last_week_col)

    # Rename "Total" header → "Budget" (it's now an editable input, not a sum)
    ws.cell(row=5, column=TOTAL_COL).value = "Budget"

    # "Pattern" column header (after the existing summary-code helper column)
    pattern_col = FIRST_WEEK_COL + num_weeks + 3
    pattern_col_letter = get_column_letter(pattern_col)
    phdr = ws.cell(row=5, column=pattern_col, value="Pattern")
    phdr.font = Font(name="Calibri", bold=True, size=8, color="1565C0")
    phdr.alignment = Alignment(horizontal="center")
    phdr.border = THIN_BORDER

    for row_idx, (row_data, title) in enumerate(zip(output.rows, row_timing)):
        excel_row = DATA_START_ROW + row_idx
        prow = title_to_row[title]  # row in the Timing sheet

        # Column C: budget total as a plain value (weekly formulas reference $C{row},
        # so the original SUM formula would create a circular reference).
        total_cell = ws.cell(row=excel_row, column=TOTAL_COL)
        total_cell.value = round(row_data.total, 2)
        total_cell.number_format = CURRENCY_FORMAT
        total_cell.font = Font(name="Calibri", size=10, bold=True)

        # Weekly cells: reference Timing sheet weight, divide by SUM of weights
        sum_range = (
            f"{_TIMING_SHEET}!${first_col_letter}${prow}"
            f":${last_week_col_letter}${prow}"
        )
        for col_offset in range(num_weeks):
            col = FIRST_WEEK_COL + col_offset
            wcol = get_column_letter(col)
            weight = f"{_TIMING_SHEET}!{wcol}${prow}"
            formula = f"=IFERROR($C{excel_row}*{weight}/SUM({sum_range}),0)"
            cell = ws.cell(row=excel_row, column=col)
            cell.value = formula
            cell.number_format = CURRENCY_FORMAT

        # Pattern label column
        pc = ws.cell(row=excel_row, column=pattern_col, value=title)
        pc.font = Font(name="Calibri", size=8, color="1565C0", italic=True)
        pc.alignment = Alignment(horizontal="center")
        pc.border = THIN_BORDER

    ws.column_dimensions[pattern_col_letter].width = 18

    # ── 5. Save ───────────────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
