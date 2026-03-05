from io import BytesIO

import openpyxl

from app.models.budget import BudgetCategory, BudgetDetailRow, BudgetLineItem, ParsedBudget
from app.services.excel_parser import parse_budget_excel
from app.services.tax_credit_writer import write_tax_credit_excel


def test_parse_budget_excel_reads_account_details_rows():
    wb = openpyxl.Workbook()
    ws_categories = wb.active
    ws_categories.title = "Categories"
    ws_categories.append(["Account", "Description", "Total"])
    ws_categories.append(["0201", "Writer(s)", 1000])

    ws_details = wb.create_sheet("Account Details")
    ws_details.append(["Account", "Description", "Amount", "Unit", "x", "Unit 2", "Currency", "Rate", "Unit 3", "Subtotal"])
    ws_details.append(["0201", "Script fee", 2, "wk", "x", 1, "CAD", 500, "ep", 1000])
    ws_details.append(["0201", "Zero row", 1, "wk", "x", 1, "CAD", 0, "ep", 0])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    parsed = parse_budget_excel(buf, filename="test.xlsx")

    assert len(parsed.detail_rows) == 1
    assert parsed.detail_rows[0].account == "0201"
    assert parsed.detail_rows[0].description == "Script fee"
    assert parsed.detail_rows[0].subtotal == 1000


def test_tax_credit_workbook_contains_detail_budget_tab_with_formula_totals():
    budget = ParsedBudget(
        line_items=[
            BudgetLineItem(code="0201", description="Writer(s)", total=1000, category=BudgetCategory.ABOVE_THE_LINE),
            BudgetLineItem(code="0301", description="Development", total=500, category=BudgetCategory.ABOVE_THE_LINE),
            BudgetLineItem(code="1001", description="Cast", total=200, category=BudgetCategory.BELOW_THE_LINE_PRODUCTION),
            BudgetLineItem(code="6001", description="Edit", total=300, category=BudgetCategory.BELOW_THE_LINE_POST),
            BudgetLineItem(code="7001", description="Publicity", total=400, category=BudgetCategory.OTHER),
        ],
        total_budget=2400,
        source_filename="test.xlsx",
        topsheet_totals={"0200": 1000, "0300": 500, "1000": 200, "6000": 300, "7000": 400},
        detail_rows=[
            BudgetDetailRow(account="0201", description="Writer fee", amount=1, unit="wk", unit2="1", currency="CAD", rate=1000, unit3="ep", subtotal=1000),
            BudgetDetailRow(account="0301", description="Script edit", amount=1, unit="wk", unit2="1", currency="CAD", rate=500, unit3="ep", subtotal=500),
            BudgetDetailRow(account="1001", description="Cast fee", amount=1, unit="wk", unit2="1", currency="CAD", rate=200, unit3="ep", subtotal=200),
            BudgetDetailRow(account="6001", description="Post", amount=1, unit="wk", unit2="1", currency="CAD", rate=300, unit3="ep", subtotal=300),
            BudgetDetailRow(account="7001", description="Other", amount=1, unit="wk", unit2="1", currency="CAD", rate=400, unit3="ep", subtotal=400),
            BudgetDetailRow(account="0301", description="Should skip", subtotal=0),
        ],
    )

    output = write_tax_credit_excel(budget, "Test")
    wb = openpyxl.load_workbook(output, data_only=False)

    assert "Detail Budget" in wb.sheetnames
    ws = wb["Detail Budget"]

    values = [row for row in ws.iter_rows(min_row=1, max_col=11, values_only=True)]
    flat_values = [v for row in values for v in row if isinstance(v, str)]

    assert "Should skip" not in flat_values

    # Columns A/B/C should be left-justified for detail rows
    assert ws["A3"].alignment.horizontal == "left"
    assert ws["B3"].alignment.horizontal == "left"
    assert ws["C3"].alignment.horizontal == "left"

    # Section total rows are formulas
    section_total_rows = [i for i, row in enumerate(values, start=1) if isinstance(row[0], str) and row[0].endswith("TOTAL")]
    assert section_total_rows
    for r in section_total_rows:
        assert isinstance(ws.cell(row=r, column=11).value, str)
        assert ws.cell(row=r, column=11).value.startswith("=SUM(")

    # A/B/C/D totals are formulas and inserted after each respective section block
    total_a_row = next(i for i, row in enumerate(values, start=1) if row[0] == 'TOTAL "A" – ABOVE THE LINE')
    total_b_row = next(i for i, row in enumerate(values, start=1) if row[0] == 'TOTAL PRODUCTION "B"')
    total_c_row = next(i for i, row in enumerate(values, start=1) if row[0] == 'TOTAL POST-PRODUCTION "C"')
    total_d_row = next(i for i, row in enumerate(values, start=1) if row[0] == 'TOTAL OTHER "D"')

    assert ws.cell(row=total_a_row, column=11).value.startswith("=SUM(")
    assert ws.cell(row=total_b_row, column=11).value.startswith("=SUM(")
    assert ws.cell(row=total_c_row, column=11).value.startswith("=SUM(")
    assert ws.cell(row=total_d_row, column=11).value.startswith("=SUM(")

    last_a_section_total_row = max(i for i, row in enumerate(values, start=1) if isinstance(row[0], str) and row[0] in {"02.00 TOTAL", "03.00 TOTAL"})
    assert total_a_row > last_a_section_total_row

    # Interior data cells are unbordered; only section outlines are bordered
    assert ws["B3"].border.left is None or ws["B3"].border.left.style is None
    assert ws["B3"].border.right is None or ws["B3"].border.right.style is None
    assert ws["A2"].border.left.style == "thin"
    assert ws["K3"].border.right.style == "thin"
